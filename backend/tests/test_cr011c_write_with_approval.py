"""CR-011-C — write-with-approval action tools + analysis export (§3).

NON-NEGOTIABLE INVARIANT (§0.2.1, §7): the agent never writes a business row
directly. Every action tool only creates a PENDING approval request; the actual
mutation happens ONLY when a human approves it via approvals.apply_request. The
guard test below asserts ZERO direct mutations to any business table.
"""
from datetime import date
from decimal import Decimal

from sqlalchemy import func, select

from app.constants import ROLE_DIRECTOR
from app.models.ai_alert import AIAlert
from app.models.approval_request import ApprovalRequest
from app.models.budget_line_item import BudgetLineItem
from app.models.client_invoice import ClientInvoice
from app.models.cost_entry import CostEntry
from app.models.dashboard import Dashboard
from app.models.notification import Notification
from app.models.report import Report
from app.models.subcontractor import Subcontractor
from app.services import agent as agent_service
from app.services import agent_actions as actions
from app.services import ai as ai_service
from app.services import approvals as approvals_service

# Business tables an action tool must NEVER mutate directly. CR-035: Report/Dashboard
# are included too — authoring a report/pano is propose-only, so a propose call must
# leave these counts unchanged (the row appears only on approval).
BUSINESS_MODELS = [Notification, AIAlert, CostEntry, ClientInvoice, Subcontractor,
                   BudgetLineItem, Report, Dashboard]


def _ids(seed, label="a"):
    return (seed[label]["company"].id,
            seed[label]["users"][ROLE_DIRECTOR].id,
            seed[label]["project"].id)


def _business_counts(db):
    return {m.__name__: db.execute(select(func.count()).select_from(m)).scalar() for m in BUSINESS_MODELS}


def _pending_requests(db, company_id):
    return db.execute(
        select(ApprovalRequest).where(
            ApprovalRequest.company_id == company_id,
            ApprovalRequest.status == "pending",
        )
    ).scalars().all()


def _seed_invoice(db, pid, cid, uid, number="HK-1"):
    inv = ClientInvoice(
        project_id=pid, company_id=cid, invoice_number=number, invoice_date=date(2026, 1, 15),
        amount_try=Decimal("100000"), vat_amount_try=Decimal("20000"),
        total_with_vat_try=Decimal("120000"), net_due_try=Decimal("114000"),
        due_date=date(2026, 2, 15), created_by=uid,
    )
    db.add(inv)
    db.commit()
    return inv


# --------------------------------------------------------------------------- #
# THE invariant: action tools propose only — zero direct mutations
# --------------------------------------------------------------------------- #
def test_action_tools_create_only_pending_requests_zero_mutations(db, seed):
    cid, uid, pid = _ids(seed)
    inv = _seed_invoice(db, pid, cid, uid)

    before = _business_counts(db)

    r1 = actions.propose_reminder(db, cid, uid, title="Hakedişi kontrol et",
                                  note="Ay sonu", due_date="2026-07-01", project_id=pid)
    r2 = actions.propose_flag_invoice(db, cid, uid, target_kind="client_invoice",
                                      target_id=str(inv.id), reason="Tutar yüksek")
    r3 = actions.propose_followup_task(db, cid, uid, title="Müşteriyi ara")
    # CR-035 — authoring a report/dashboard is propose-only too.
    r4 = actions.propose_report(db, cid, uid, title="Proje Kârlılığı",
                                spec={"metrics": ["cost_try"], "dimensions": ["project"], "viz": "table"})
    r5 = actions.propose_dashboard(db, cid, uid, title="Özet Pano", widgets=[
        {"id": "w1", "type": "kpi", "title": "Maliyet", "layout": {"x": 0, "y": 0, "w": 3, "h": 2},
         "spec": {"metrics": ["cost_try"], "viz": "kpi"}},
    ])

    # Every action returned a pending proposal (never "done").
    for r in (r1, r2, r3, r4, r5):
        assert r["status"] == "pending"
        assert r["proposed_action"]["status"] == "pending"
        assert "onayınızı bekliyor" in r["message"]

    # ZERO direct mutations to any business table.
    assert _business_counts(db) == before

    # Exactly five PENDING approval requests, all tagged proposed_by_agent.
    pend = _pending_requests(db, cid)
    assert len(pend) == 5
    assert all(p.proposed_by_agent for p in pend)
    assert {p.kind for p in pend} == {
        "agent_reminder", "agent_flag_invoice", "agent_task",
        "agent_create_report", "agent_create_dashboard",
    }


def test_registries_are_disjoint_actions_not_in_readonly():
    """The propose-only tools are NOT in the read-only registry, so the read-only
    guarantee on TOOL_REGISTRY is untouched."""
    assert agent_service.ACTION_TOOL_NAMES.isdisjoint(set(agent_service.TOOL_REGISTRY))
    assert agent_service.ACTION_TOOL_NAMES == {
        "propose_reminder", "propose_flag_invoice", "propose_followup_task",
        # CR-035 — Report Studio AI authoring.
        "propose_report", "propose_dashboard",
    }


def test_action_schemas_present_and_marked_proposal():
    schemas = {t["name"]: t for t in agent_service.build_tool_schemas()}
    for name in agent_service.ACTION_TOOL_NAMES:
        assert name in schemas
        # The model is told these are proposals, not direct writes.
        assert "ÖNERİ" in schemas[name]["description"]
        assert "company_id" not in schemas[name]["input_schema"].get("properties", {})


# --------------------------------------------------------------------------- #
# Validation
# --------------------------------------------------------------------------- #
def test_propose_flag_invoice_rejects_bad_target_kind(db, seed):
    cid, uid, _ = _ids(seed)
    import pytest
    with pytest.raises(actions.ActionError):
        actions.propose_flag_invoice(db, cid, uid, target_kind="project",
                                     target_id="x", reason="r")


def test_propose_flag_invoice_rejects_missing_record(db, seed):
    import uuid
    cid, uid, _ = _ids(seed)
    import pytest
    with pytest.raises(actions.ActionError):
        actions.propose_flag_invoice(db, cid, uid, target_kind="client_invoice",
                                     target_id=str(uuid.uuid4()), reason="r")


def test_propose_reminder_requires_title(db, seed):
    cid, uid, _ = _ids(seed)
    import pytest
    with pytest.raises(actions.ActionError):
        actions.propose_reminder(db, cid, uid, title="   ")


def test_propose_flag_invoice_company_isolation(db, seed):
    """Company A cannot flag company B's invoice (cross-company → not found)."""
    a_cid, a_uid, _ = _ids(seed, "a")
    b_cid, b_uid, b_pid = _ids(seed, "b")
    b_inv = _seed_invoice(db, b_pid, b_cid, b_uid, number="B-HK")
    import pytest
    with pytest.raises(actions.ActionError):
        actions.propose_flag_invoice(db, a_cid, a_uid, target_kind="client_invoice",
                                     target_id=str(b_inv.id), reason="r")


# --------------------------------------------------------------------------- #
# Approve -> apply round-trip: the mutation happens ONLY on approval
# --------------------------------------------------------------------------- #
def test_reminder_applied_only_on_approval(db, seed):
    cid, uid, pid = _ids(seed)
    actions.propose_reminder(db, cid, uid, title="Hatırlat", project_id=pid)

    # Before approval: no notification exists.
    assert db.execute(select(func.count()).select_from(Notification)).scalar() == 0

    req = _pending_requests(db, cid)[0]
    approvals_service.apply_request(db, req)
    approvals_service.mark_decided(req, user_id=uid, status="approved")
    db.commit()

    notes = db.execute(select(Notification).where(Notification.company_id == cid)).scalars().all()
    assert len(notes) == 1
    assert notes[0].title == "Hatırlat"
    assert notes[0].user_id == uid


def test_flag_invoice_applied_only_on_approval(db, seed):
    cid, uid, pid = _ids(seed)
    inv = _seed_invoice(db, pid, cid, uid)
    actions.propose_flag_invoice(db, cid, uid, target_kind="client_invoice",
                                 target_id=str(inv.id), reason="Mükerrer olabilir")

    assert db.execute(select(func.count()).select_from(AIAlert)).scalar() == 0

    req = next(p for p in _pending_requests(db, cid) if p.kind == "agent_flag_invoice")
    approvals_service.apply_request(db, req)
    approvals_service.mark_decided(req, user_id=uid, status="approved")
    db.commit()

    alerts = db.execute(select(AIAlert).where(AIAlert.company_id == cid)).scalars().all()
    assert len(alerts) == 1
    assert alerts[0].source_type == "client_invoice"
    assert str(alerts[0].source_id) == str(inv.id)
    assert alerts[0].dedup_key.startswith("agent_flag:")


# --------------------------------------------------------------------------- #
# Through the agent loop + executor
# --------------------------------------------------------------------------- #
class _Block:
    def __init__(self, **kw):
        self.__dict__.update(kw)


class _Resp:
    def __init__(self, stop_reason, content):
        self.stop_reason = stop_reason
        self.content = content


class _Messages:
    def __init__(self, responder):
        self._r = responder
        self.calls = 0

    def create(self, **kw):
        r = self._r(self.calls, kw)
        self.calls += 1
        return r


class _Client:
    def __init__(self, responder):
        self.messages = _Messages(responder)


def test_run_agent_surfaces_proposed_actions_without_writing(db, seed, monkeypatch):
    cid, uid, pid = _ids(seed)

    def responder(call, kw):
        if call == 0:
            return _Resp("tool_use", [_Block(type="tool_use", name="propose_reminder",
                                             input={"title": "Ara beni"}, id="a0")])
        return _Resp("end_turn", [_Block(type="text", text="Öneri oluşturuldu, onayınızı bekliyor.")])

    monkeypatch.setattr(ai_service, "_client", lambda: _Client(responder))
    out = agent_service.run_agent(db, cid, [{"role": "user", "content": "bana hatırlat"}], user_id=uid)

    assert out["tools_used"] == ["propose_reminder"]
    assert len(out["proposed_actions"]) == 1
    assert out["proposed_actions"][0]["kind"] == "agent_reminder"
    assert out["proposed_actions"][0]["status"] == "pending"
    # A pending request exists; NO notification was written (not approved).
    assert len(_pending_requests(db, cid)) == 1
    assert db.execute(select(func.count()).select_from(Notification)).scalar() == 0


def test_action_tool_needs_user_context(db, seed):
    """Without a user_id the executor refuses the action (no anonymous proposals)."""
    cid, _, _ = _ids(seed)
    out = agent_service.execute_tool(db, cid, "propose_reminder", {"title": "x"},
                                     [], [], set(), date(2026, 6, 19), None, [])
    assert "error" in out
    assert len(_pending_requests(db, cid)) == 0


# --------------------------------------------------------------------------- #
# Endpoint round-trip: propose via /ai/agent -> approve via /approvals
# --------------------------------------------------------------------------- #
def test_endpoint_propose_then_director_approves(client, seed, monkeypatch):
    director = seed["a"]["users"][ROLE_DIRECTOR]

    def responder(call, kw):
        if call == 0:
            return _Resp("tool_use", [_Block(type="tool_use", name="propose_followup_task",
                                             input={"title": "Teklif hazırla"}, id="a0")])
        return _Resp("end_turn", [_Block(type="text", text="Görev önerisi oluşturuldu.")])

    monkeypatch.setattr(ai_service, "_client", lambda: _Client(responder))
    client.login(director)

    r = client.post("/api/v1/ai/agent",
                    json={"messages": [{"role": "user", "content": "bir görev oluştur"}]})
    assert r.status_code == 200, r.text
    data = r.json()["data"]
    assert len(data["proposed_actions"]) == 1
    req_id = data["proposed_actions"][0]["request_id"]

    # The pending request shows up in /approvals, badged as agent-proposed.
    lst = client.get("/api/v1/approvals").json()["data"]
    match = [i for i in lst if i.get("request_id") == req_id]
    assert match and match[0]["proposed_by_agent"] is True
    assert match[0]["kind"] == "agent_task"

    # Director approves -> the task becomes a notification (applied only now).
    ap = client.put(f"/api/v1/approvals/request/{req_id}/approve")
    assert ap.status_code == 200, ap.text


# --------------------------------------------------------------------------- #
# Analysis export (§3.2)
# --------------------------------------------------------------------------- #
_ANALYSIS = {
    "title": "Tedarikçi Analizi",
    "question": "Akçansa ile ne kadar harcadık?",
    "answer_markdown": "Akçansa ile toplam **4.500 ₺** harcandı.",
    "charts": [{
        "chart_type": "line", "title": "Aylık Harcama", "x_key": "month",
        "series": [{"key": "total", "label": "Toplam", "type": "line"}],
        "data": [{"month": "2026-01", "total": 4000}, {"month": "2026-02", "total": 500}],
        "source_note": "Kaynak: maliyet kayıtları",
    }],
    "citations": [{"type": "cost_entry", "label": "Akçansa — 4.000 ₺",
                   "deep_link": "/projects/x/dashboard?highlight=1"}],
}


def test_export_pdf_renders(db, seed):
    from app.models.company import Company
    from app.services import reports

    company = db.get(Company, seed["a"]["company"].id)
    pdf = reports.render_agent_analysis_pdf(company, _ANALYSIS)
    assert isinstance(pdf, bytes) and pdf[:4] == b"%PDF"
    assert len(pdf) > 1000


def test_export_excel_renders(db, seed):
    from io import BytesIO

    from openpyxl import load_workbook

    from app.models.company import Company
    from app.services import reports

    company = db.get(Company, seed["a"]["company"].id)
    xlsx = reports.render_agent_analysis_excel(company, _ANALYSIS)
    assert isinstance(xlsx, bytes) and len(xlsx) > 0
    wb = load_workbook(BytesIO(xlsx))
    assert "Analiz" in wb.sheetnames
    assert "Kaynaklar" in wb.sheetnames
    # The chart data made it into its own sheet.
    assert any(name not in ("Analiz", "Kaynaklar", "Bilgi") for name in wb.sheetnames)


def test_export_endpoint_pdf_and_excel(client, seed):
    client.login(seed["a"]["users"][ROLE_DIRECTOR])
    r = client.post("/api/v1/ai/agent/export?fmt=pdf", json=_ANALYSIS)
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/pdf"
    assert r.content[:4] == b"%PDF"

    r2 = client.post("/api/v1/ai/agent/export?fmt=excel", json=_ANALYSIS)
    assert r2.status_code == 200, r2.text
    assert "spreadsheetml" in r2.headers["content-type"]


def test_export_endpoint_requires_answer(client, seed):
    client.login(seed["a"]["users"][ROLE_DIRECTOR])
    r = client.post("/api/v1/ai/agent/export?fmt=pdf",
                    json={"answer_markdown": "   ", "charts": [], "citations": []})
    assert r.status_code == 422
