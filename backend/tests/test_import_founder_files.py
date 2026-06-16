"""Regression: import of real-world files shaped like the founder's exports.

Root cause (reported): a regular import showed NOTHING and AI import said "AI şu an
kullanılamıyor". The founder's files are NOT in YAPI's positional template — their
columns are in a different order (e.g. "Harcama Tarihi" / "Gerçek Harcama (TL)")
with title rows above the header and, for the .xlsm, a junk first sheet. The old
positional parser read column 0 (a row index) as the date → every row skipped →
empty preview; the AI path masked any JSON parse/format failure as "AI unavailable".

These tests reproduce the founder's shape and lock in: header-name column mapping
(rows detected), an honest message when nothing imports, and the AI path raising a
DISTINCT parse error (not a fake outage).
"""
import io

import pytest
from openpyxl import Workbook

from app.constants import ROLE_DIRECTOR
from app.services import ai as ai_service
from app.services import excel_import as ei

# The founder's "Data Giris" layout: title rows, then a header whose columns are
# in a different order than YAPI's template.
FOUNDER_HEADER = [
    "#", "Proje / Alt Proje", "İş Kalemi / Açıklama", "Kategori", "Sorumlu",
    "Plan. Başlangıç", "Plan. Bitiş", "Gerçek Bitiş", "İlerleme (0-1)",
    "Bütçe (TL)", "Gerçek Harcama (TL)", "Harcama Tarihi", "USD/TRY Kuru",
]


def _founder_xlsx(rows, *, title_rows=True, category="DIGER HARCAMALAR"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Giris"
    if title_rows:
        ws.append(["✏️  213 ADA — VERİ GİRİŞİ"])  # emoji title row above the header
        ws.append(["Sarı hücrelere veri girin"])
    ws.append(FOUNDER_HEADER)
    for i, (item, cat, budget, spend, spend_date) in enumerate(rows, 1):
        ws.append([i, "213 Ada 1 Parsel", item, cat or category, "—",
                   "12.06.2017", "30.12.2020", spend_date, "1",
                   budget, spend, spend_date, "5.80"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Regular import: header-name mapping detects the founder's rows
# --------------------------------------------------------------------------- #
def test_founder_columns_are_mapped_by_header():
    data = _founder_xlsx([("BAHŞİŞ VB.", None, "4330", "4330", "07.09.2019")])
    res = ei.validate_rows(data)
    cm = res["column_map"]
    assert cm["entry_date"] == 11      # "Harcama Tarihi", not column 0 (the # index)
    assert cm["amount_try"] == 10      # "Gerçek Harcama (TL)"
    assert cm["cost_category"] == 3


def test_founder_rows_detected_not_silently_skipped():
    data = _founder_xlsx([
        ("BAHŞİŞ VB.", "DIGER HARCAMALAR", "4330", "4330", "07.09.2019"),
        ("DIŞ CEPHE", "DIŞ CEPHE", "116000", "116000", "09.08.2019"),
    ])
    res = ei.validate_rows(data)
    nonskipped = [r for r in res["rows"] if not r["skipped"]]
    assert len(nonskipped) == 2                       # the OLD bug skipped all rows
    r0 = nonskipped[0]
    assert str(r0["data"]["entry_date"]) == "2019-09-07"   # parsed from Harcama Tarihi
    assert str(r0["data"]["amount_try"]) == "4330"
    # Unknown category surfaces as an editable per-row error (not a silent drop).
    assert any("Kategori tanınmıyor" in e for e in r0["errors"])


def test_founder_row_with_known_category_is_valid():
    data = _founder_xlsx([("Saha dökümü", "Malzeme — Beton", "150000", "150000", "10.05.2025")])
    res = ei.validate_rows(data)
    row = next(r for r in res["rows"] if not r["skipped"])
    assert row["valid"] is True and row["errors"] == []
    assert row["data"]["cost_category"] == "material_concrete"


def test_title_rows_above_header_are_handled():
    data = _founder_xlsx([("X", "Malzeme — Beton", "1000", "1000", "10.05.2025")], title_rows=True)
    res = ei.validate_rows(data)
    assert res["header_detected"] is True
    assert any(not r["skipped"] for r in res["rows"])


# --------------------------------------------------------------------------- #
# Honest messages instead of a silent empty preview
# --------------------------------------------------------------------------- #
def test_unmappable_sheet_returns_message():
    # A junk sheet (no date/category/amount columns, mostly prose).
    wb = Workbook(); ws = wb.active; ws.title = "Makro Kurulum"
    ws.append(["⚡ VBA MAKRO KURULUM"]); ws.append(["1. Alt+F11"]); ws.append(["2. Yapıştır"])
    buf = io.BytesIO(); wb.save(buf)
    res = ei.validate_rows(buf.getvalue())
    assert not any(not r["skipped"] for r in res["rows"])
    assert res["message"] == ei.ERR_NO_IMPORTABLE


def test_empty_sheet_returns_message():
    wb = Workbook(); buf = io.BytesIO(); wb.save(buf)
    res = ei.validate_rows(buf.getvalue())
    assert res["rows"] == [] and res["message"] == ei.ERR_EMPTY_SHEET


def test_preview_endpoint_surfaces_message_and_rows(client, seed):
    client.login(seed["a"]["users"][ROLE_DIRECTOR])
    pid = seed["a"]["project"].id
    data = _founder_xlsx([("BAHŞİŞ VB.", "DIGER HARCAMALAR", "4330", "4330", "07.09.2019")])
    r = client.post(
        f"/api/v1/projects/{pid}/costs/import/preview",
        files={"file": ("f.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200, r.text
    meta = r.json()["meta"]
    assert meta["total"] >= 1               # rows detected, not an empty preview
    assert "column_map" in meta


# --------------------------------------------------------------------------- #
# AI path: robust JSON extraction + parse-error is NOT a fake outage
# --------------------------------------------------------------------------- #
def test_extract_json_handles_nested_and_fenced():
    assert ai_service._extract_json('{"a": {"b": [1, 2]}, "c": 3}') == {"a": {"b": [1, 2]}, "c": 3}
    assert ai_service._extract_json('Sonuç:\n```json\n{"x": 1}\n```\nBitti') == {"x": 1}
    assert ai_service._extract_json('[{"a": 1}]') == [{"a": 1}]


def test_extract_json_raises_on_truncated():
    with pytest.raises(ValueError):
        ai_service._extract_json('{"maliyet_girisleri": [{"a": 1},')  # cut off mid-JSON


def test_analyze_excel_import_parse_failure_raises_response_error(monkeypatch):
    # The model "answered" but with unparseable text -> AIResponseError, NOT AIUnavailable.
    monkeypatch.setattr(ai_service, "_call_raw_text", lambda *a, **k: "Üzgünüm, yardımcı olamıyorum.")
    with pytest.raises(ai_service.AIResponseError):
        ai_service.analyze_excel_import("### SAYFA: Data\n1 | 2 | 3")


def test_analyze_excel_import_transport_failure_stays_unavailable(monkeypatch):
    def _boom(*a, **k):
        raise ai_service.AIUnavailable(ai_service.AI_UNAVAILABLE_MESSAGE)
    monkeypatch.setattr(ai_service, "_call_raw_text", _boom)
    with pytest.raises(ai_service.AIUnavailable):
        ai_service.analyze_excel_import("text")


def test_analyze_excel_import_valid_json_returns_keys(monkeypatch):
    monkeypatch.setattr(ai_service, "_call_raw_text",
                        lambda *a, **k: '{"maliyet_girisleri": [{"amount_try": 100}]}')
    out = ai_service.analyze_excel_import("text")
    assert out["maliyet_girisleri"] == [{"amount_try": 100}]
    for key in ("faturalar", "alt_yukleniciler", "ekipman", "tanimsiz"):
        assert key in out  # defaulted


def test_ai_import_endpoint_maps_response_error_to_422(client, seed, monkeypatch):
    from app.config import settings

    original_key = settings.anthropic_api_key
    settings.anthropic_api_key = "test-key"  # pass the is_available() gate
    monkeypatch.setattr(
        ai_service, "analyze_excel_import",
        lambda text: (_ for _ in ()).throw(ai_service.AIResponseError(ai_service.AI_RESPONSE_MESSAGE)),
    )
    try:
        client.login(seed["a"]["users"][ROLE_DIRECTOR])
        pid = seed["a"]["project"].id
        data = _founder_xlsx([("BAHŞİŞ VB.", "DIGER HARCAMALAR", "4330", "4330", "07.09.2019")])
        r = client.post(
            f"/api/v1/projects/{pid}/ai-import",
            files={"file": ("f.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 422, r.text          # NOT 503 "AI unavailable"
        assert r.json()["error"]["code"] == "AI_RESPONSE_ERROR"
    finally:
        settings.anthropic_api_key = original_key


def test_ai_import_endpoint_empty_file_is_clear(client, seed, monkeypatch):
    from app.config import settings

    original_key = settings.anthropic_api_key
    settings.anthropic_api_key = "test-key"
    try:
        client.login(seed["a"]["users"][ROLE_DIRECTOR])
        pid = seed["a"]["project"].id
        wb = Workbook(); buf = io.BytesIO(); wb.save(buf)  # empty workbook
        r = client.post(
            f"/api/v1/projects/{pid}/ai-import",
            files={"file": ("f.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 422, r.text
        assert "veri bulunamadı" in r.json()["error"]["message"].lower()
    finally:
        settings.anthropic_api_key = original_key
