"""CR-049 — time-accurate AI authoring + richer preview (backend guards).

Covers the date-window bug behind the "DGN Martı" skill (monthly cash came out
ALL-ZERO because the agent windowed to 2026 while the project's data is 2018–2020):

* FIX A — an explicit ``all_time``/``tum_zamanlar`` preset resolves to the whole-data
  window ``(None, None)``; the engine's cash grain then spans the project's REAL life
  (data-relative), NOT the rolling 18-month ``project_cashflow`` anchored to today
  (which silently empties a project whose data predates today). Recent presets still
  window. The catalog surfaces the presets (incl. all_time) to the agent + picker.
* FIX B — ``list_projects`` exposes ``start_date`` + ``actual_end_date``.
* FIX C — the authoring guidance steers a project lifetime view to all_time.
* Period label — the renderer labels an all-time result from its REAL span.
"""
from datetime import date
from decimal import Decimal

import pytest

from app.constants import ROLE_DIRECTOR
from app.models.cost_entry import CostEntry
from app.responses import APIError
from app.services import agent
from app.services.agent_tools import list_projects, studio_catalog
from app.services.studio import catalog, engine, excel_report

D = Decimal
TODAY = date(2026, 6, 30)


def _uid(seed, label="a"):
    return seed[label]["users"][ROLE_DIRECTOR].id


def _cost(db, p, amount, *, uid, d):
    amt = D(str(amount))
    db.add(CostEntry(
        project_id=p.id, company_id=p.company_id, entry_date=d, cost_category="material_steel",
        amount_try=amt, vat_amount_try=D("0"), total_with_vat_try=amt,
        payment_status="unpaid", entry_type="actual", created_by=uid,
    ))
    db.flush()


def _set_span(db, p, start, end):
    p.start_date = start
    p.planned_end_date = end
    db.add(p)
    db.flush()


def _run(db, cid, spec):
    return engine.run_spec(db, cid, spec, today=TODAY)


# --------------------------------------------------------------------------- #
# FIX A — all_time resolution + validation
# --------------------------------------------------------------------------- #
def test_all_time_preset_resolves_to_all_data(db, seed):
    p = seed["a"]["project"]
    uid = _uid(seed)
    _cost(db, p, "100000", uid=uid, d=date(2018, 6, 10))
    _cost(db, p, "50000", uid=uid, d=date(2026, 1, 10))
    db.commit()
    for preset in ("all_time", "tum_zamanlar"):
        res = _run(db, p.company_id, {"metrics": ["cost_try"], "date_range": {"preset": preset}})
        assert res["meta"]["date_range"] == {"from": None, "to": None}
        assert res["totals"]["metrics"]["cost_try"] == 150000.0  # every row, both years
    # Parity: an explicit all_time == omitting date_range entirely.
    none_total = _run(db, p.company_id, {"metrics": ["cost_try"]})["totals"]["metrics"]["cost_try"]
    assert none_total == 150000.0


def test_all_time_preset_is_validated(db):
    catalog.validate_spec({"metrics": ["cost_try"], "date_range": {"preset": "all_time"}})
    catalog.validate_spec({"metrics": ["cost_try"], "date_range": {"preset": "tum_zamanlar"}})
    assert engine.is_known_preset("all_time") and engine.is_known_preset("tum_zamanlar")
    with pytest.raises(APIError):
        catalog.validate_spec({"metrics": ["cost_try"], "date_range": {"preset": "never_ever"}})


# --------------------------------------------------------------------------- #
# The repro: monthly cashflow populated across the real span, not all-zero-2026
# --------------------------------------------------------------------------- #
def test_all_time_cashflow_spans_real_project_life(db, seed):
    p = seed["a"]["project"]
    uid = _uid(seed)
    _set_span(db, p, date(2018, 1, 1), date(2020, 12, 31))
    _cost(db, p, "300000", uid=uid, d=date(2018, 6, 15))
    _cost(db, p, "400000", uid=uid, d=date(2019, 9, 20))
    _cost(db, p, "500000", uid=uid, d=date(2020, 3, 5))
    db.commit()

    # all-time (no date_range) → monthly cash rows span 2018–2020 and are non-zero.
    res = _run(db, p.company_id, {"metrics": ["cash_out", "net_cash"], "dimensions": ["month"]})
    months = {r["dims"]["month"] for r in res["rows"]}
    assert {"2018-06", "2019-09", "2020-03"} <= months
    assert round(sum(r["metrics"]["cash_out"] for r in res["rows"]), 2) == 1200000.0
    assert len([r for r in res["rows"] if r["metrics"]["cash_out"]]) >= 3  # the 3 data months

    # The explicit all_time preset behaves the same.
    res2 = _run(db, p.company_id, {"metrics": ["cash_out"], "dimensions": ["month"],
                                   "date_range": {"preset": "tum_zamanlar"}})
    assert round(sum(r["metrics"]["cash_out"] for r in res2["rows"]), 2) == 1200000.0


def test_recent_window_still_empties_a_past_project(db, seed):
    # The OLD (buggy) symptom proven intact: a 2026 window over 2018–2020 data is
    # empty — recent presets still window correctly (no over-correction to all-time).
    p = seed["a"]["project"]
    uid = _uid(seed)
    _set_span(db, p, date(2018, 1, 1), date(2020, 12, 31))
    _cost(db, p, "300000", uid=uid, d=date(2018, 6, 15))
    db.commit()
    res = _run(db, p.company_id, {"metrics": ["cash_out"], "dimensions": ["month"],
                                  "date_range": {"preset": "bu_yil"}})  # 2026
    assert round(sum((r["metrics"]["cash_out"] or 0) for r in res["rows"]), 2) == 0.0
    rc = _run(db, p.company_id, {"metrics": ["cost_try"], "date_range": {"preset": "son_3_ay"}})
    assert (rc["totals"]["metrics"]["cost_try"] or 0) == 0.0


def test_recent_preset_still_narrows(db, seed):
    p = seed["a"]["project"]
    uid = _uid(seed)
    _cost(db, p, "100000", uid=uid, d=date(2026, 6, 10))   # inside son_3_ay
    _cost(db, p, "70000", uid=uid, d=date(2018, 1, 10))    # outside
    db.commit()
    res = _run(db, p.company_id, {"metrics": ["cost_try"], "date_range": {"preset": "son_3_ay"}})
    assert res["totals"]["metrics"]["cost_try"] == 100000.0
    assert res["meta"]["date_range"]["to"] == "2026-06-30"


def test_dgn_marti_cashflow_repro_via_skill(db, seed):
    # End-to-end: a project-scoped skill deck with a monthly Nakit Akışı widget windowed
    # all_time → the canonical batch-run populates 2018–2020 (the CR-049 acceptance);
    # a 2026 preset would have been all-zero.
    from app.api.studio import _run_dashboard_batch
    from app.services.skills import _plan_dashboard

    p = seed["a"]["project"]
    uid = _uid(seed)
    p.name = "DGN Martı"
    _set_span(db, p, date(2018, 1, 1), date(2020, 12, 31))
    _cost(db, p, "300000", uid=uid, d=date(2018, 6, 15))
    _cost(db, p, "500000", uid=uid, d=date(2020, 3, 5))
    db.commit()

    widget = {"id": "cf", "type": "chart", "title": "Aylık Nakit Akışı",
              "layout": {"x": 0, "y": 0, "w": 6, "h": 4},
              "spec": {"metrics": ["cash_out", "net_cash"], "dimensions": ["month"],
                       "viz": "line", "date_range": {"preset": "tum_zamanlar"}}}
    deck = _plan_dashboard({"widgets": [widget], "project_scope": str(p.id),
                            "date_range": {"preset": "tum_zamanlar"}})
    res = _run_dashboard_batch(db, seed["a"]["users"][ROLE_DIRECTOR], deck)["cf"]
    assert round(sum((r["metrics"]["cash_out"] or 0) for r in res["rows"]), 2) == 800000.0
    assert {"2018-06", "2020-03"} <= {r["dims"]["month"] for r in res["rows"]}


# --------------------------------------------------------------------------- #
# FIX A (catalog) — date presets surfaced to the agent + the manual picker
# --------------------------------------------------------------------------- #
def test_catalog_surfaces_date_presets_with_all_time(db, seed):
    cat = catalog.get_catalog_public()
    ids = {p["id"] for p in cat["date_presets"]}
    assert {"all_time", "last_3_months", "ytd"} <= ids
    all_time = next(p for p in cat["date_presets"] if p["id"] == "all_time")
    assert "Tüm zamanlar" in all_time["label"]
    # the agent's studio_catalog tool relays them
    tool = studio_catalog(db, seed["a"]["company"].id)
    assert any(p["id"] == "all_time" for p in tool["date_presets"])


# --------------------------------------------------------------------------- #
# FIX B — list_projects timeframe
# --------------------------------------------------------------------------- #
def test_list_projects_exposes_timeframe(db, seed):
    p = seed["a"]["project"]
    p.start_date = date(2018, 1, 1)
    p.actual_end_date = date(2020, 12, 31)
    db.add(p)
    db.commit()
    rec = next(r for r in list_projects(db, seed["a"]["company"].id)["records"] if r["id"] == str(p.id))
    assert rec["start_date"] == "2018-01-01"
    assert rec["actual_end_date"] == "2020-12-31"


def test_list_projects_actual_end_date_nullable(db, seed):
    p = seed["a"]["project"]   # actual_end_date is None by default
    db.commit()
    rec = next(r for r in list_projects(db, seed["a"]["company"].id)["records"] if r["id"] == str(p.id))
    assert rec["start_date"] is not None
    assert rec["actual_end_date"] is None


# --------------------------------------------------------------------------- #
# Period label — labelled from the real span (renderer)
# --------------------------------------------------------------------------- #
def test_period_label_from_real_span():
    span = excel_report._period_label({
        "meta": {"date_range": {"from": None, "to": None}},
        "columns": [{"id": "month", "kind": "dimension", "type": "date"}],
        "rows": [{"dims": {"month": "2018-06"}}, {"dims": {"month": "2020-03"}}],
    })
    assert span == "2018 – 2020"
    single = excel_report._period_label({
        "meta": {"date_range": {"from": None, "to": None}},
        "columns": [{"id": "year", "kind": "dimension", "type": "date"}],
        "rows": [{"dims": {"year": "2019"}}],
    })
    assert single == "2019"
    none = excel_report._period_label({
        "meta": {"date_range": {"from": None, "to": None}},
        "columns": [{"id": "cost_try", "kind": "metric", "type": "currency"}],
        "rows": [{"dims": {}}],
    })
    assert none == "Tüm zamanlar"
    explicit = excel_report._period_label({"meta": {"date_range": {"from": "2026-01-01", "to": "2026-03-31"}}})
    assert explicit == "Ocak 2026 – Mart 2026"


def test_excel_header_prefers_concrete_span_over_all_time():
    # A multi-widget all-time workbook whose FIRST widget is a snapshot KPI (no time
    # dim → "Tüm zamanlar") must still headline the REAL span a later time-series
    # widget covers ("2018 – 2020"), not the bare all-time label.
    import io

    from openpyxl import load_workbook

    from app.services.studio.excel_report import build_workbook

    meta = {"date_range": {"from": None, "to": None}, "basis": {"currency": "try"}}
    kpi_res = {"columns": [{"id": "revenue", "label": "Gelir", "kind": "metric", "type": "currency"}],
               "rows": [], "totals": {"metrics": {"revenue": 100.0}}, "meta": meta}
    cash_res = {"columns": [{"id": "month", "label": "Ay", "kind": "dimension", "type": "date"},
                            {"id": "cash_out", "label": "Nakit çıkış", "kind": "metric", "type": "currency"}],
                "rows": [{"dims": {"month": "2018-06"}, "metrics": {"cash_out": 300000.0}},
                         {"dims": {"month": "2020-03"}, "metrics": {"cash_out": 500000.0}}],
                "totals": {"metrics": {"cash_out": 800000.0}}, "meta": meta}
    widgets = [{"id": "k", "type": "kpi", "title": "Özet", "spec": {"viz": "kpi"}},
               {"id": "c", "type": "table", "title": "Nakit", "spec": {"viz": "table"}}]
    wb = load_workbook(io.BytesIO(build_workbook(widgets, {"k": kpi_res, "c": cash_res}, "Test")))
    cells = [c.value for s in wb.worksheets for row in s.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("2018 – 2020" in c for c in cells)
    assert not any(c.strip() == "Tüm zamanlar" for c in cells)


# --------------------------------------------------------------------------- #
# FIX C — the steering text the model receives (agent behaviour is not LLM-testable)
# --------------------------------------------------------------------------- #
def test_authoring_guidance_steers_lifetime_to_all_time():
    g = agent._ACTION_GUIDANCE
    assert "all_time" in g and "tum_zamanlar" in g
    assert "ZAMAN PENCERESİ" in g
    skill = next(s for s in agent.build_tool_schemas() if s["name"] == "propose_skill")
    assert "tum_zamanlar" in skill["description"]
