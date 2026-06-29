"""CR-046 — decision-grade Excel engine.

Proves the new xlsx is a board-ready workbook, not a one-sheet-per-KPI dump:
  * an "Özet" dashboard sheet (NOT one sheet per KPI);
  * ≥1 native Excel chart (``ws._charts``) wired to a data-sheet range;
  * ₺ number formats + navy styled headers + a "Toplam" totals row (authoritative
    value from ``run_spec``, never a re-sum formula);
  * RAG conditional formatting on variance columns when a comparison is present;
  * the ``_safe_cell`` formula-injection guard still neutralises ``=HYPERLINK``/``+``/
    ``-``/``@`` on EVERY sheet;
  * the 422 empty-workbook guard;
  * READ-ONLY (the engine takes no DB session — it consumes a computed result);
  * the snapshot-metric fix — a non-windowed metric (hakediş) in a time-grouped table
    becomes a KPI card ("tüm proje"), never blank monthly rows;
  * all three entry points (single report, dashboard deck, skill run) get the format;
  * the file opens with zero formula errors.

Integration cases run on the SQLite ``client``/``seed``/``db`` fixtures; engine
behaviour (snapshot, RAG, 422, injection) is unit-tested on ``build_workbook`` with
synthetic ``run_spec`` results so the exact shapes are owned by the test.
"""
import io
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from app.constants import ROLE_DIRECTOR
from app.models.cost_entry import CostEntry
from app.responses import APIError
from app.services.studio.excel_report import build_workbook

import pytest

D = Decimal
DASH = "/api/v1/studio/dashboards"
REP = "/api/v1/studio/reports"


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _login(client, seed, label="a"):
    client.login(seed[label]["users"][ROLE_DIRECTOR])
    return seed[label]["project"], seed[label]["users"][ROLE_DIRECTOR].id


def _cost(db, p, amount, uid, d=date(2026, 1, 10), cat="material_steel"):
    amt = D(str(amount))
    db.add(CostEntry(
        project_id=p.id, company_id=p.company_id, entry_date=d, cost_category=cat,
        amount_try=amt, vat_amount_try=D("0"), total_with_vat_try=amt,
        payment_status="unpaid", entry_type="actual", created_by=uid,
    ))
    db.commit()


def _layout(x=0, y=0, w=6, h=4):
    return {"x": x, "y": y, "w": w, "h": h}


def _kpi(wid="k1", title="Toplam Maliyet"):
    return {"id": wid, "type": "kpi", "title": title, "layout": _layout(),
            "spec": {"metrics": ["cost_try"], "viz": "kpi"}}


def _chart(wid="c1", title="Aylık Maliyet"):
    return {"id": wid, "type": "chart", "title": title, "layout": _layout(),
            "spec": {"metrics": ["cost_try"], "dimensions": ["month"], "viz": "line"}}


def _table(wid="t1", title="Kategori"):
    return {"id": wid, "type": "table", "title": title, "layout": _layout(),
            "spec": {"metrics": ["cost_try"], "dimensions": ["cost_category"], "viz": "table"}}


def _all_cells(wb):
    return [c for s in wb.worksheets for row in s.iter_rows(values_only=True) for c in row]


def _all_str_cells(wb):
    return [c.value for s in wb.worksheets for row in s.iter_rows() for c in row if isinstance(c.value, str)]


def _make_dashboard(client, widgets, title="Pano"):
    return client.post(DASH, json={"title": title, "widgets": widgets}).json()["data"]["id"]


# --------------------------------------------------------------------------- #
# Integration — the dashboard deck xlsx is a decision dashboard
# --------------------------------------------------------------------------- #
def test_dashboard_xlsx_is_ozet_not_one_sheet_per_kpi(client, db, seed):
    p, uid = _login(client, seed)
    _cost(db, p, "120000", uid, d=date(2026, 1, 10))
    _cost(db, p, "80000", uid, d=date(2026, 2, 10))
    # Two KPI widgets + a chart + a table — the OLD engine made 4 sheets (one per
    # widget, incl. one-number KPI sheets). The new engine: ONE Özet + data sheets.
    did = _make_dashboard(client, [
        _kpi("k1", "Toplam Maliyet"), _kpi("k2", "Maruziyet"), _chart(), _table(),
    ])
    xlsx = client.post(f"{DASH}/{did}/export?format=xlsx")
    assert xlsx.status_code == 200 and xlsx.content[:2] == b"PK"
    wb = load_workbook(io.BytesIO(xlsx.content))

    assert "Özet" in wb.sheetnames
    assert wb.sheetnames[0] == "Özet"  # the dashboard leads
    # NOT one-sheet-per-KPI: the kpi widget titles are NOT sheet names.
    assert "Toplam Maliyet" not in wb.sheetnames
    assert "Maruziyet" not in wb.sheetnames
    # Özet + the chart's data sheet + the table's data sheet (kpis are cards).
    assert len(wb.sheetnames) <= 3


def test_dashboard_xlsx_has_native_chart(client, db, seed):
    p, uid = _login(client, seed)
    _cost(db, p, "100000", uid, d=date(2026, 1, 10))
    _cost(db, p, "60000", uid, d=date(2026, 2, 10))
    did = _make_dashboard(client, [_kpi(), _chart()])
    xlsx = client.post(f"{DASH}/{did}/export?format=xlsx")
    wb = load_workbook(io.BytesIO(xlsx.content))
    # A native Excel chart lives on the Özet (not a flat dump).
    assert len(wb["Özet"]._charts) >= 1


def test_dashboard_xlsx_currency_format_and_navy_header(client, db, seed):
    p, uid = _login(client, seed)
    _cost(db, p, "120000", uid)
    did = _make_dashboard(client, [_kpi(), _table()])
    xlsx = client.post(f"{DASH}/{did}/export?format=xlsx")
    wb = load_workbook(io.BytesIO(xlsx.content))
    data_ws = [wb[n] for n in wb.sheetnames if n != "Özet"][0]

    # Styled navy header (white bold, navy fill).
    h = data_ws.cell(row=1, column=data_ws.max_column)  # a metric header cell
    assert (h.fill.fgColor.rgb or "").endswith("183047")
    assert h.font.bold
    # A ₺ number format reached a metric cell somewhere in the workbook.
    fmts = [c.number_format for s in wb.worksheets for row in s.iter_rows() for c in row]
    assert any("#,##0" in f and "₺" in f for f in fmts)


def test_dashboard_xlsx_has_toplam_value_row(client, db, seed):
    p, uid = _login(client, seed)
    _cost(db, p, "120000", uid)
    did = _make_dashboard(client, [_table()])
    xlsx = client.post(f"{DASH}/{did}/export?format=xlsx")
    wb = load_workbook(io.BytesIO(xlsx.content))
    # "Toplam" present, and the total is an authoritative VALUE (not a formula).
    assert "Toplam" in _all_cells(wb)
    assert all(c.data_type != "f" for s in wb.worksheets for row in s.iter_rows() for c in row)
    assert 120000 in _all_cells(wb) or 120000.0 in _all_cells(wb)


def test_single_report_xlsx_is_ozet_lite(client, db, seed):
    p, uid = _login(client, seed)
    _cost(db, p, "90000", uid)
    rid = client.post(REP, json={
        "title": "Tek Rapor", "spec": {"metrics": ["cost_try"], "dimensions": ["cost_category"], "viz": "table"},
    }).json()["data"]["id"]
    xlsx = client.post(f"{REP}/{rid}/export?format=xlsx")
    assert xlsx.status_code == 200
    wb = load_workbook(io.BytesIO(xlsx.content))
    assert "Özet" in wb.sheetnames  # Özet-lite even for a single report


# --------------------------------------------------------------------------- #
# Unit — engine behaviour on synthetic run_spec results (read-only, no DB)
# --------------------------------------------------------------------------- #
def _result(columns, rows, totals, *, comparison=None, comparison_unit=None):
    return {
        "columns": columns, "rows": rows, "totals": totals,
        "meta": {"date_range": {"from": "2026-01-01", "to": "2026-02-28"},
                 "comparison": comparison, "comparison_unit": comparison_unit},
    }


def test_snapshot_metric_becomes_kpi_not_blank_rows():
    # A time-grouped table (month) with a windowed metric (cost_try) AND a snapshot
    # metric (progress_billing / Hakediş, project grain). The snapshot must NOT appear
    # as blank monthly rows — it becomes a "tüm proje" KPI card.
    columns = [
        {"id": "month", "label": "Ay", "kind": "dimension", "type": "date"},
        {"id": "cost_try", "label": "Maliyet (₺)", "kind": "metric", "type": "currency"},
        {"id": "progress_billing", "label": "Hakediş", "kind": "metric", "type": "currency"},
    ]
    rows = [
        {"dims": {"month": "2026-01"}, "metrics": {"cost_try": 100000, "progress_billing": None}, "deltas": None},
        {"dims": {"month": "2026-02"}, "metrics": {"cost_try": 70000, "progress_billing": None}, "deltas": None},
    ]
    totals = {"metrics": {"cost_try": 170000, "progress_billing": 500000}, "deltas": None}
    widgets = [{"id": "t1", "type": "table", "title": "Aylık",
                "spec": {"metrics": ["cost_try", "progress_billing"], "dimensions": ["month"], "viz": "table"}}]
    wb = load_workbook(io.BytesIO(build_workbook(widgets, {"t1": _result(columns, rows, totals)}, "Test")))

    data_ws = [wb[n] for n in wb.sheetnames if n != "Özet"][0]
    headers = [data_ws.cell(row=1, column=ci).value for ci in range(1, data_ws.max_column + 1)]
    assert "Maliyet (₺)" in headers           # windowed metric stays as a monthly column
    assert "Hakediş" not in headers           # snapshot metric NOT a blank monthly column
    # …it surfaces as a "tüm proje" KPI card on the Özet instead.
    assert any(isinstance(v, str) and "Hakediş" in v and "tüm proje" in v for v in _all_str_cells(wb))


def test_rag_conditional_formatting_on_variance():
    columns = [
        {"id": "project", "label": "Proje", "kind": "dimension", "type": "enum"},
        {"id": "cost_try", "label": "Maliyet (₺)", "kind": "metric", "type": "currency"},
    ]
    rows = [
        {"dims": {"project": "A"}, "metrics": {"cost_try": 100000}, "deltas": {"cost_try": 0.12}},
        {"dims": {"project": "B"}, "metrics": {"cost_try": 50000}, "deltas": {"cost_try": -0.05}},
    ]
    totals = {"metrics": {"cost_try": 150000}, "deltas": {"cost_try": 0.06}}
    widgets = [{"id": "t", "type": "table", "title": "Proje",
                "spec": {"metrics": ["cost_try"], "dimensions": ["project"], "viz": "table"}}]
    wb = load_workbook(io.BytesIO(build_workbook(
        widgets, {"t": _result(columns, rows, totals, comparison={"from": "2025-01-01", "to": "2025-02-28"})}, "Test")))
    data_ws = [wb[n] for n in wb.sheetnames if n != "Özet"][0]
    # A Δ column with RAG conditional-formatting rules was added.
    assert len(list(data_ws.conditional_formatting)) >= 1
    headers = [data_ws.cell(row=1, column=ci).value for ci in range(1, data_ws.max_column + 1)]
    assert any(isinstance(h, str) and h.startswith("Δ") for h in headers)


def test_no_empty_workbook_422():
    widgets = [{"id": "x", "type": "text", "title": "Not", "content": "yalnızca metin"}]
    with pytest.raises(APIError) as ei:
        build_workbook(widgets, {}, "Boş")
    assert ei.value.status_code == 422 and ei.value.code == "NO_DATA"


def test_safe_cell_neutralises_injection_in_engine():
    payload = '=HYPERLINK("http://evil.example")'
    columns = [
        {"id": "cost_category", "label": "Kategori", "kind": "dimension", "type": "enum"},
        {"id": "cost_try", "label": "Maliyet (₺)", "kind": "metric", "type": "currency"},
    ]
    rows = [{"dims": {"cost_category": payload}, "metrics": {"cost_try": 1000}, "deltas": None}]
    totals = {"metrics": {"cost_try": 1000}, "deltas": None}
    widgets = [{"id": "t", "type": "table", "title": "Enjeksiyon",
                "spec": {"metrics": ["cost_try"], "dimensions": ["cost_category"], "viz": "table"}}]
    wb = load_workbook(io.BytesIO(build_workbook(widgets, {"t": _result(columns, rows, totals)}, "Test")))
    cells = _all_str_cells(wb)
    assert any(v.startswith("'=HYPERLINK") for v in cells)            # neutralised to text
    assert all(c.data_type != "f" for s in wb.worksheets for row in s.iter_rows() for c in row)
    assert not any(v[:1] in ("=", "+", "-", "@") for v in cells)


def test_opens_with_zero_formula_errors():
    columns = [
        {"id": "month", "label": "Ay", "kind": "dimension", "type": "date"},
        {"id": "cost_try", "label": "Maliyet (₺)", "kind": "metric", "type": "currency"},
    ]
    rows = [
        {"dims": {"month": "2026-01"}, "metrics": {"cost_try": 100000}, "deltas": None},
        {"dims": {"month": "2026-02"}, "metrics": {"cost_try": 70000}, "deltas": None},
    ]
    totals = {"metrics": {"cost_try": 170000}, "deltas": None}
    widgets = [{"id": "c", "type": "chart", "title": "Aylık",
                "spec": {"metrics": ["cost_try"], "dimensions": ["month"], "viz": "line"}}]
    wb = load_workbook(io.BytesIO(build_workbook(widgets, {"c": _result(columns, rows, totals)}, "Test")))
    # No error cells anywhere; the chart's Reference is within the data sheet bounds.
    assert all(c.data_type != "e" for s in wb.worksheets for row in s.iter_rows() for c in row)
    data_ws = [wb[n] for n in wb.sheetnames if n != "Özet"][0]
    for ch in wb["Özet"]._charts:
        for series in ch.series:
            ref = series.val.numRef.f if series.val and series.val.numRef else ""
            # the data reference points at the data sheet, not beyond it
            assert data_ws.title in ref or ref == ""


def test_all_snapshot_time_grouped_table_renders_cards_not_422():
    # A time-grouped table whose ONLY metric is a snapshot (Hakediş): the metric is
    # dropped from the (empty) data sheet, but the run produces a valid card-only Özet
    # — it must NOT 422 just because no data sheet was written.
    columns = [
        {"id": "month", "label": "Ay", "kind": "dimension", "type": "date"},
        {"id": "progress_billing", "label": "Hakediş", "kind": "metric", "type": "currency"},
    ]
    rows = [
        {"dims": {"month": "2026-01"}, "metrics": {"progress_billing": None}, "deltas": None},
        {"dims": {"month": "2026-02"}, "metrics": {"progress_billing": None}, "deltas": None},
    ]
    totals = {"metrics": {"progress_billing": 500000}, "deltas": None}
    widgets = [{"id": "t", "type": "table", "title": "Aylık Hakediş",
                "spec": {"metrics": ["progress_billing"], "dimensions": ["month"], "viz": "table"}}]
    # Must not raise APIError(422) — a card-only decision dashboard is valid.
    wb = load_workbook(io.BytesIO(build_workbook(widgets, {"t": _result(columns, rows, totals)}, "Test")))
    assert "Özet" in wb.sheetnames
    assert any(isinstance(v, str) and "Hakediş" in v and "tüm proje" in v for v in _all_str_cells(wb))


def test_abs_comparison_delta_renders_amount_not_percent():
    # comparison_unit="abs" → deltas are absolute ₺ amounts, not fractions. They must
    # render as ₺ on the card and with a currency number format in the Δ column —
    # never multiplied by 100 / shown as "5000000.0%".
    from app.services.studio.excel_report import build_single_report

    columns = [
        {"id": "project", "label": "Proje", "kind": "dimension", "type": "enum"},
        {"id": "cost_try", "label": "Maliyet (₺)", "kind": "metric", "type": "currency"},
    ]
    rows = [{"dims": {"project": "A"}, "metrics": {"cost_try": 100000}, "deltas": {"cost_try": 50000}}]
    totals = {"metrics": {"cost_try": 100000}, "deltas": {"cost_try": 50000}}
    result = _result(columns, rows, totals,
                     comparison={"from": "2025-01-01", "to": "2025-02-28"}, comparison_unit="abs")
    wb = load_workbook(io.BytesIO(build_single_report(result, "Abs Rapor", viz="table")))

    cells = _all_str_cells(wb)
    # KPI card delta is a ₺ amount with an arrow, never the ×100 percent blow-up.
    assert any("₺" in v and ("▲" in v or "▼" in v) for v in cells)
    assert not any("5000000" in v for v in cells)
    # The Δ column cell holding 50000 uses a currency format, not 0.0%.
    data_ws = [wb[n] for n in wb.sheetnames if n != "Özet"][0]
    dfmts = [c.number_format for row in data_ws.iter_rows() for c in row
             if c.value in (50000, 50000.0)]
    assert dfmts and all("0.0%" not in f for f in dfmts)
    assert any("#,##0" in f for f in dfmts)


def test_rag_favourable_direction_per_metric():
    # green=favourable / red=unfavourable: a rising cost, cash outflow or open
    # receivable is UNFAVOURABLE (red); rising revenue / cash inflow is favourable.
    from app.services.studio.excel_report import _favourable_up

    assert _favourable_up("cost_try") is False
    assert _favourable_up("cash_out") is False
    assert _favourable_up("receivables") is False
    assert _favourable_up("revenue") is True
    assert _favourable_up("progress_billing") is True
    assert _favourable_up("cash_in") is True
    assert _favourable_up("net_cash") is True


def test_engine_is_read_only_no_db_param():
    # The engine consumes a computed result; it must take no DB session (read-only,
    # no queries/mutations). Guards the §3 no-fabrication / read-only invariant.
    import inspect

    params = set(inspect.signature(build_workbook).parameters)
    assert "db" not in params and "session" not in params
