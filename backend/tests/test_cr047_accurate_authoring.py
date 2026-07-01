"""CR-047 — accurate AI authoring: project-scoped + revenue-model-aware.

Guards the three fixes behind the "DGN Martı" bug (a project-named report that
silently reported the whole company and showed hakediş for a sell-side project):

* FIX B (engine model-awareness — the accuracy GUARANTEE): model-inapplicable
  metrics return ``None`` (rendered "–") regardless of what the plan asks —
  ``progress_billing``/``billing_vs_contract`` null for sell-side; ``unit_sales_revenue``
  / ``*_per_m2`` / ``irr`` / ``roi`` null for non-sell-side. The model-aware
  ``revenue`` is unchanged.
* FIX A (project scoping GUARANTEED at run time): a skill carrying
  ``plan.project_scope`` merges a ``{field:project,op:=,value:scope}`` filter into
  every data widget at run time, so the totals are THAT project's, not company-wide;
  a widget's own project filter still wins; no scope ⇒ unchanged.
* FIX C (model-aware compile): ``list_projects`` returns ``revenue_model``;
  ``propose_skill`` validates + stores ``project_scope``.

Plus the DGN Martı repro end-to-end.
"""
import uuid
from datetime import date
from decimal import Decimal

import pytest
from sqlalchemy import func, select

from app.api.studio import _global_merge, _run_dashboard_batch
from app.constants import ROLE_DIRECTOR
from app.models.client_invoice import ClientInvoice
from app.models.cost_entry import CostEntry
from app.models.dashboard import Dashboard
from app.models.landowner_payment import LandownerPayment
from app.models.project import Project
from app.models.unit_sale import UnitSale
from app.services import agent_actions as actions
from app.services.skills import _plan_dashboard
from app.services.studio import engine

D = Decimal
TODAY = date(2026, 6, 30)


# --------------------------------------------------------------------------- #
# Builders (mirror test_cr032)
# --------------------------------------------------------------------------- #
def _uid(seed, label="a"):
    return seed[label]["users"][ROLE_DIRECTOR].id


def _user(seed, label="a"):
    return seed[label]["users"][ROLE_DIRECTOR]


def _cost(db, p, amount, *, uid, d=date(2026, 1, 10), cat="material_steel"):
    amt = D(str(amount))
    db.add(CostEntry(
        project_id=p.id, company_id=p.company_id, entry_date=d, cost_category=cat,
        amount_try=amt, vat_amount_try=D("0"), total_with_vat_try=amt,
        payment_status="unpaid", entry_type="actual", created_by=uid,
    ))
    db.flush()


def _sale(db, p, label, price, *, unit_type="2+1", net_m2="80"):
    db.add(UnitSale(
        project_id=p.id, company_id=p.company_id, unit_label=label, unit_type=unit_type,
        net_m2=D(str(net_m2)), sale_price_try=D(str(price)), sale_date=date(2026, 1, 5),
    ))
    db.flush()


def _invoice(db, p, amount, *, uid):
    amt = D(str(amount))
    db.add(ClientInvoice(
        project_id=p.id, company_id=p.company_id, invoice_number=f"HK-{p.project_code}-{amount}",
        invoice_date=date(2026, 1, 15), amount_try=amt, vat_amount_try=D("0"),
        total_with_vat_try=amt, net_due_try=amt, amount_received_try=D("0"),
        retention_amount_try=D("0"), due_date=date(2026, 2, 15), created_by=uid,
    ))
    db.flush()


def _set_sell_side(db, p, net_m2="200"):
    p.revenue_model = "kat_karsiligi"
    p.construction_net_m2 = D(net_m2)
    db.add(p)
    db.flush()


def _set_hakedis(db, p):
    p.revenue_model = "hakedis"
    db.add(p)
    db.flush()


def _project(db, cid, name, uid, *, code, revenue_model="hakedis"):
    p = Project(
        company_id=cid, name=name, project_code=code, project_type="road",
        client_name="İşveren", contract_value_try=1_000_000, original_budget_try=800_000,
        start_date=date(2025, 1, 1), planned_end_date=date(2025, 12, 31),
        project_manager_id=uid, revenue_model=revenue_model,
    )
    db.add(p)
    db.flush()
    return p


def _run(db, cid, spec):
    return engine.run_spec(db, cid, spec, today=TODAY)


def _proj_metrics(db, cid, metrics):
    """Run a project-grouped spec and return the single project's metrics dict."""
    res = _run(db, cid, {"metrics": metrics, "dimensions": ["project"]})
    return res["rows"][0]["metrics"]


# --------------------------------------------------------------------------- #
# FIX B — engine model-aware null map
# --------------------------------------------------------------------------- #
def test_progress_billing_null_for_sell_side_real_for_hakedis(db, seed):
    p = seed["a"]["project"]
    _set_hakedis(db, p)
    _invoice(db, p, "70000", uid=_uid(seed))
    db.commit()

    m = _proj_metrics(db, p.company_id, ["progress_billing", "billing_vs_contract"])
    assert m["progress_billing"] == 70000.0                  # hakediş project → real
    assert m["billing_vs_contract"] == 7.0                   # 70000 / 1,000,000 contract

    _set_sell_side(db, p)                                    # → kat_karsiligi (sell-side)
    db.commit()
    m2 = _proj_metrics(db, p.company_id, ["progress_billing", "billing_vs_contract"])
    assert m2["progress_billing"] is None                    # hakediş meaningless → "–"
    assert m2["billing_vs_contract"] is None


def test_unit_sales_and_m2_null_for_hakedis_real_for_sell_side(db, seed):
    p = seed["a"]["project"]
    _set_sell_side(db, p)
    _sale(db, p, "A1", "1000000", net_m2="80")
    _cost(db, p, "100000", uid=_uid(seed))
    db.commit()

    sell = ["unit_sales_revenue", "cost_per_m2", "revenue_per_m2", "profit_per_m2", "roi", "irr"]
    m = _proj_metrics(db, p.company_id, sell)
    assert m["unit_sales_revenue"] == 1000000.0              # sell-side → real
    assert m["cost_per_m2"] is not None and m["revenue_per_m2"] is not None

    _set_hakedis(db, p)
    db.commit()
    m2 = _proj_metrics(db, p.company_id, sell)
    for mid in sell:
        assert m2[mid] is None, f"{mid} must be null for a hakediş project"


def test_revenue_unchanged_for_both_models(db, seed):
    # FIX B must NOT touch the model-aware `revenue` (sales.py) — it is non-null for
    # both a sell-side (unit sales) and a hakediş (invoices) project.
    p = seed["a"]["project"]
    _set_sell_side(db, p)
    _sale(db, p, "A1", "1000000", net_m2="80")
    db.commit()
    assert _proj_metrics(db, p.company_id, ["revenue"])["revenue"] == 1000000.0

    _set_hakedis(db, p)
    _invoice(db, p, "70000", uid=_uid(seed))
    db.commit()
    assert _proj_metrics(db, p.company_id, ["revenue"])["revenue"] is not None


# --------------------------------------------------------------------------- #
# FIX A — project scoping guaranteed at run time
# --------------------------------------------------------------------------- #
def test_scoped_skill_run_merges_project_filter(db, seed):
    cid, uid = seed["a"]["company"].id, _uid(seed)
    p1 = seed["a"]["project"]
    p2 = _project(db, cid, "İkinci Proje", uid, code="P2")
    _cost(db, p1, "100000", uid=uid)
    _cost(db, p2, "777000", uid=uid)
    db.commit()

    widget = {"id": "w1", "type": "kpi", "title": "Maliyet", "layout": {"x": 0, "y": 0, "w": 3, "h": 2},
              "spec": {"metrics": ["cost_try"], "viz": "kpi"}}

    # Scoped to p1 (widget has NO project filter) → only p1's 100000, not 877000.
    scoped = _run_dashboard_batch(db, _user(seed), _plan_dashboard({"widgets": [widget], "project_scope": str(p1.id)}))
    assert scoped["w1"]["totals"]["metrics"]["cost_try"] == 100000.0

    # No scope → company-wide (both projects).
    unscoped = _run_dashboard_batch(db, _user(seed), _plan_dashboard({"widgets": [widget]}))
    assert unscoped["w1"]["totals"]["metrics"]["cost_try"] == 877000.0


def test_widget_own_project_filter_wins_over_scope(db, seed):
    cid, uid = seed["a"]["company"].id, _uid(seed)
    p1 = seed["a"]["project"]
    p2 = _project(db, cid, "İkinci Proje", uid, code="P2")
    _cost(db, p1, "100000", uid=uid)
    _cost(db, p2, "777000", uid=uid)
    db.commit()

    # The widget explicitly filters p2, but the plan scope is p1 → the widget WINS.
    widget = {"id": "w1", "type": "kpi", "title": "Maliyet", "layout": {"x": 0, "y": 0, "w": 3, "h": 2},
              "spec": {"metrics": ["cost_try"], "viz": "kpi",
                       "filters": [{"field": "project", "op": "=", "value": str(p2.id)}]}}
    res = _run_dashboard_batch(db, _user(seed), _plan_dashboard({"widgets": [widget], "project_scope": str(p1.id)}))
    assert res["w1"]["totals"]["metrics"]["cost_try"] == 777000.0   # p2 (widget), not p1 (scope)


def test_global_merge_no_scope_is_unchanged():
    deck = Dashboard(widgets=[], date_range=None, comparison=None, filters=None)
    merged = _global_merge({"metrics": ["cost_try"]}, deck, None)
    assert not any(f.get("field") == "project" for f in (merged.get("filters") or []))
    # With a scope it injects exactly one project filter.
    pid = str(uuid.uuid4())
    merged2 = _global_merge({"metrics": ["cost_try"]}, deck, pid)
    proj_filters = [f for f in merged2["filters"] if f.get("field") == "project"]
    assert proj_filters == [{"field": "project", "op": "=", "value": pid}]


# --------------------------------------------------------------------------- #
# FIX C — model-aware compile
# --------------------------------------------------------------------------- #
def test_list_projects_returns_revenue_model(db, seed):
    from app.services.agent_tools import list_projects

    p = seed["a"]["project"]
    _set_sell_side(db, p)
    db.commit()
    out = list_projects(db, seed["a"]["company"].id)
    rec = next(r for r in out["records"] if r["id"] == str(p.id))
    assert rec["revenue_model"] == "kat_karsiligi"


def test_propose_skill_validates_and_stores_project_scope(db, seed):
    cid, uid, p = seed["a"]["company"].id, _uid(seed), seed["a"]["project"]
    widget = {"id": "w1", "type": "kpi", "title": "Maliyet", "layout": {"x": 0, "y": 0, "w": 3, "h": 2},
              "spec": {"metrics": ["cost_try"], "viz": "kpi"}}

    r = actions.propose_skill(db, cid, uid, name="DGN", widgets=[widget], project_scope=str(p.id))
    assert r["proposed_action"]["plan"]["project_scope"] == str(p.id)

    # No scope → None in the plan (company-wide skill stays valid).
    r2 = actions.propose_skill(db, cid, uid, name="Genel", widgets=[widget])
    assert r2["proposed_action"]["plan"]["project_scope"] is None

    # A bad / cross-company project id is rejected (the agent must resolve or ask).
    with pytest.raises(actions.ActionError):
        actions.propose_skill(db, cid, uid, name="Kötü", widgets=[widget], project_scope=str(uuid.uuid4()))


def test_unit_sales_revenue_null_on_unit_grain_for_hakedis(db, seed):
    # The dual/unit grain (grouped by unit_type) must ALSO null unit_sales_revenue for
    # a non-sell-side project — the accuracy guarantee can't have a hole on the exact
    # metric the CR calls out, even if a hakediş project carries a stray UnitSale.
    p = seed["a"]["project"]
    _set_hakedis(db, p)
    _sale(db, p, "A1", "1000000", net_m2="80", unit_type="2+1")
    db.commit()
    rows = _run(db, p.company_id, {"metrics": ["unit_sales_revenue"], "dimensions": ["unit_type"]})["rows"]
    for row in rows:
        assert row["metrics"]["unit_sales_revenue"] is None

    # Sell-side: the same per-unit_type grouping yields a real value.
    _set_sell_side(db, p)
    db.commit()
    rows2 = _run(db, p.company_id, {"metrics": ["unit_sales_revenue"], "dimensions": ["unit_type"]})["rows"]
    assert any(r["metrics"]["unit_sales_revenue"] for r in rows2)


def test_non_equals_widget_project_filter_still_scoped(db, seed):
    # A widget's NON-equals project filter must NOT escape the skill scope: the scope
    # is still ANDed on, so the report can only narrow within the scoped project.
    cid, uid = seed["a"]["company"].id, _uid(seed)
    p1 = seed["a"]["project"]
    p2 = _project(db, cid, "İkinci Proje", uid, code="P2")
    _cost(db, p1, "100000", uid=uid)
    _cost(db, p2, "777000", uid=uid)
    db.commit()
    widget = {"id": "w1", "type": "kpi", "title": "Maliyet", "layout": {"x": 0, "y": 0, "w": 3, "h": 2},
              "spec": {"metrics": ["cost_try"], "viz": "kpi",
                       "filters": [{"field": "project", "op": "!=", "value": str(p2.id)}]}}
    res = _run_dashboard_batch(db, _user(seed), _plan_dashboard({"widgets": [widget], "project_scope": str(p1.id)}))
    # scope(p1) AND project!=p2 → p1 only (100000), never widened to company-wide.
    assert res["w1"]["totals"]["metrics"]["cost_try"] == 100000.0


_WIDGET = {"id": "w1", "type": "kpi", "title": "Maliyet", "layout": {"x": 0, "y": 0, "w": 3, "h": 2},
           "spec": {"metrics": ["cost_try"], "viz": "kpi"}}


def _plan(scope=None):
    return {"format": "xlsx", "title": "X", "widgets": [_WIDGET], "project_scope": scope}


def test_post_skill_validates_project_scope(client, db, seed):
    # The direct persistence path re-validates project_scope (same as the agent draft),
    # so a hand-rolled POST can't store a foreign/garbage scope.
    director = seed["a"]["users"][ROLE_DIRECTOR]
    client.login(director)
    body = lambda plan: {"name": "X", "instruction": "y", "plan": plan, "format": "xlsx"}

    assert client.post("/api/v1/skills", json=body(_plan(str(uuid.uuid4())))).status_code == 422   # not a real project
    assert client.post("/api/v1/skills", json=body(_plan("not-a-uuid"))).status_code == 422          # non-UUID
    # A real, in-company project id is accepted.
    ok = client.post("/api/v1/skills", json=body(_plan(str(seed["a"]["project"].id))))
    assert ok.status_code == 200
    # PUT re-validates too.
    sid = ok.json()["data"]["id"]
    assert client.put(f"/api/v1/skills/{sid}", json={"plan": _plan(str(uuid.uuid4()))}).status_code == 422


def test_null_metric_renders_dash_in_xlsx(client, db, seed):
    # The CR §4 acceptance: a sell-side hakediş shows "–" in the deliverable, not blank.
    import io

    from openpyxl import load_workbook

    from app.services.studio.export import studio_export_dashboard

    p = seed["a"]["project"]
    _set_sell_side(db, p)            # kat_karsiligi → progress_billing is null
    _invoice(db, p, "70000", uid=_uid(seed))
    db.commit()
    widget = {"id": "w1", "type": "table", "title": "Hakediş", "layout": {"x": 0, "y": 0, "w": 6, "h": 4},
              "spec": {"metrics": ["progress_billing"], "dimensions": ["project"], "viz": "table"}}
    deck = _plan_dashboard({"widgets": [widget], "project_scope": str(p.id)})
    results = _run_dashboard_batch(db, _user(seed), deck)
    resp = studio_export_dashboard(deck.widgets, results, "Test", "xlsx", company="Şirket")
    wb = load_workbook(io.BytesIO(resp.body))
    cells = [c.value for s in wb.worksheets for row in s.iter_rows() for c in row]
    assert "–" in cells   # the null progress_billing renders the en-dash, not a blank


# --------------------------------------------------------------------------- #
# The DGN Martı repro (end-to-end)
# --------------------------------------------------------------------------- #
def test_dgn_marti_repro_scoped_and_model_correct(db, seed):
    cid, uid = seed["a"]["company"].id, _uid(seed)
    dgn = seed["a"]["project"]
    dgn.name = "DGN Martı"
    _set_sell_side(db, dgn)                                   # kat_karsiligi (sell-side)
    karadeniz = _project(db, cid, "Karadeniz Atıksu", uid, code="KAR", revenue_model="hakedis")
    _sale(db, dgn, "Daire 1", "5800000", net_m2="80")        # DGN's real revenue = unit sales
    _invoice(db, karadeniz, "168000000", uid=uid)            # the cross-project hakediş (the 168M)
    db.commit()

    widget = {"id": "w1", "type": "kpi", "title": "Gelir/Hakediş",
              "layout": {"x": 0, "y": 0, "w": 3, "h": 2},
              "spec": {"metrics": ["revenue", "unit_sales_revenue", "progress_billing"], "viz": "kpi"}}
    res = _run_dashboard_batch(db, _user(seed), _plan_dashboard({"widgets": [widget], "project_scope": str(dgn.id)}))
    t = res["w1"]["totals"]["metrics"]

    assert t["unit_sales_revenue"] == 5800000.0    # DGN's unit sales (~5.8M), scoped
    assert t["revenue"] == 5800000.0               # model-aware revenue = unit sales
    assert t["progress_billing"] is None           # sell-side → hakediş "–", NOT 213.95M company-wide
