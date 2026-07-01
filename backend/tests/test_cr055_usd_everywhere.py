"""CR-055 — USD-at-date alongside ₺, export currency, chart auto-sizing.

Part A (USD everywhere):
  * ``run_spec_usd`` is the USD companion of a spec: cost_try is swapped for its cost_usd
    sibling (₺-locked → real USD), cash metrics are nulled (no engine USD), and sibling
    ids are relabelled back so it pairs column-for-column with the ₺ run. USD figures
    EQUAL a real ``basis='usd'`` run — never fabricated.
  * ``build_workbook``/``build_single_report`` with ``results_usd`` render a ₺ headline +
    USD secondary on currency KPI cards, a paired "($)" column after each ₺ column, an
    Özet footnote (rate-at-date), and "–" for a figure with no USD (+ a missing note).
  * With no ``results_usd`` the output is ₺-only — identical to pre-CR-055 (no regression).

Part C (chart auto-sizing):
  * chart width grows with the category count (clamped); stacked charts don't overlap
    (the Özet anchor advances by each chart's actual height).

Rendering is unit-tested on synthetic results (read-only), and the USD companion +
end-to-end xlsx are integration-tested on the SQLite ``client``/``seed``/``db`` fixtures.
"""
import io
from datetime import date
from decimal import Decimal

from openpyxl import Workbook, load_workbook

from app.constants import ROLE_DIRECTOR
from app.models.cost_entry import CostEntry
from app.services.studio.excel_report import (
    add_chart, build_single_report, build_workbook, data_sheet,
)

D = Decimal


# --------------------------------------------------------------------------- #
# Synthetic result helpers (mirror CR-052/054)
# --------------------------------------------------------------------------- #
def _col(cid, kind, ctype, label):
    return {"id": cid, "kind": kind, "type": ctype, "label": label}


def _result(columns, rows, totals, *, currency="try", usd_missing=0, comparison=None, comparison_unit=None):
    return {
        "columns": columns, "rows": rows, "totals": totals,
        "meta": {"date_range": {"from": "2026-01-01", "to": "2026-02-28"},
                 "comparison": comparison, "comparison_unit": comparison_unit,
                 "currency": currency, "usd_missing_count": usd_missing},
    }


def _table(wid, title, metrics, dims, viz="table"):
    return {"id": wid, "type": "table", "title": title,
            "spec": {"metrics": metrics, "dimensions": dims, "viz": viz}}


def _data_ws(wb):
    return [wb[n] for n in wb.sheetnames if n != "Özet"][0]


def _headers(ws):
    return [ws.cell(row=1, column=ci).value for ci in range(1, ws.max_column + 1)]


def _fmts_for(ws, value):
    return [c.number_format for row in ws.iter_rows() for c in row if c.value == value]


def _ozet_strings(wb):
    return [c.value for row in wb["Özet"].iter_rows() for c in row if isinstance(c.value, str)]


# --------------------------------------------------------------------------- #
# Part A rendering — dual ₺ + USD (synthetic results)
# --------------------------------------------------------------------------- #
def _cost_cols():
    return [_col("cat", "dimension", "enum", "Kategori"),
            _col("cost_try", "metric", "currency", "Maliyet (₺)")]


def test_currency_kpi_shows_try_and_usd():
    cols = _cost_cols()
    try_res = _result(cols, [{"dims": {"cat": "A"}, "metrics": {"cost_try": 1_000_000}, "deltas": None}],
                      {"metrics": {"cost_try": 1_000_000}, "deltas": None})
    usd_res = _result(cols, [{"dims": {"cat": "A"}, "metrics": {"cost_try": 40_000}, "deltas": None}],
                      {"metrics": {"cost_try": 40_000}, "deltas": None}, currency="usd")
    wb = load_workbook(io.BytesIO(build_single_report(try_res, "Maliyet", viz="table", result_usd=usd_res)))
    ozet = wb["Özet"]
    # KPI card: ₺ headline (₺ format) + USD secondary (│ $ format) — both present on the Özet.
    try_fmts = [c.number_format for row in ozet.iter_rows() for c in row if c.value == 1_000_000]
    usd_fmts = [c.number_format for row in ozet.iter_rows() for c in row if c.value == 40_000]
    assert try_fmts and any("₺" in f for f in try_fmts)
    assert usd_fmts and any("$" in f for f in usd_fmts)   # USD secondary line rendered


def test_table_currency_column_gets_paired_usd_column():
    cols = _cost_cols()
    try_rows = [{"dims": {"cat": f"K{i}"}, "metrics": {"cost_try": 1000 * (5 - i)}, "deltas": None} for i in range(4)]
    usd_rows = [{"dims": {"cat": f"K{i}"}, "metrics": {"cost_try": 40 * (5 - i)}, "deltas": None} for i in range(4)]
    try_res = _result(cols, try_rows, {"metrics": {"cost_try": 10000}, "deltas": None})
    usd_res = _result(cols, usd_rows, {"metrics": {"cost_try": 400}, "deltas": None}, currency="usd")
    wb = load_workbook(io.BytesIO(build_single_report(try_res, "Maliyet", viz="table", result_usd=usd_res)))
    ws = _data_ws(wb)
    headers = _headers(ws)
    assert "Maliyet (₺)" in headers          # ₺ column kept primary
    assert "Maliyet ($)" in headers          # paired USD column added right after it
    # the paired USD values carry the $ format and EQUAL the companion values (no fabrication).
    for uv in (200, 160, 120, 80):
        fmts = _fmts_for(ws, uv)
        assert fmts and all("$" in f for f in fmts)


def test_missing_usd_renders_dash_and_footnote():
    cols = _cost_cols()
    try_rows = [{"dims": {"cat": "A"}, "metrics": {"cost_try": 1000}, "deltas": None},
                {"dims": {"cat": "B"}, "metrics": {"cost_try": 2000}, "deltas": None}]
    usd_rows = [{"dims": {"cat": "A"}, "metrics": {"cost_try": 40}, "deltas": None},
                {"dims": {"cat": "B"}, "metrics": {"cost_try": None}, "deltas": None}]  # B: no amount_usd
    try_res = _result(cols, try_rows, {"metrics": {"cost_try": 3000}, "deltas": None})
    usd_res = _result(cols, usd_rows, {"metrics": {"cost_try": 40}, "deltas": None},
                      currency="usd", usd_missing=1)
    wb = load_workbook(io.BytesIO(build_single_report(try_res, "Maliyet", viz="table", result_usd=usd_res)))
    strings = _ozet_strings(wb)
    assert any("USD değerleri" in s and "CR-014" in s for s in strings)   # honesty footnote
    assert any("USD kuru yok" in s for s in strings)                       # missing note (1 row)
    # the B row's USD cell is the en-dash, never a fabricated number.
    ws = _data_ws(wb)
    usd_col = _headers(ws).index("Maliyet ($)") + 1
    usd_cells = [ws.cell(row=r, column=usd_col).value for r in range(2, ws.max_row + 1)]
    assert "–" in usd_cells


def test_no_results_usd_is_try_only_no_regression():
    cols = _cost_cols()
    rows = [{"dims": {"cat": f"K{i}"}, "metrics": {"cost_try": 1000 * (5 - i)}, "deltas": None} for i in range(4)]
    wb = load_workbook(io.BytesIO(build_single_report(_result(cols, rows, {"metrics": {"cost_try": 10000}}),
                                                      "Maliyet", viz="table")))
    assert "Maliyet ($)" not in _headers(_data_ws(wb))       # no USD column
    assert not any("USD değerleri" in s for s in _ozet_strings(wb))  # no footnote


def test_non_currency_metric_has_no_usd_pair():
    # A percent metric gets NO paired USD column (USD pairs are for ₺ currency figures).
    cols = [_col("proje", "dimension", "enum", "Proje"),
            _col("roi", "metric", "percent", "ROI")]
    try_res = _result(cols, [{"dims": {"proje": "A"}, "metrics": {"roi": 39.7}, "deltas": None}],
                      {"metrics": {"roi": 39.7}, "deltas": None})
    usd_res = _result(cols, [{"dims": {"proje": "A"}, "metrics": {"roi": 39.7}, "deltas": None}],
                      {"metrics": {"roi": 39.7}, "deltas": None}, currency="usd")
    wb = load_workbook(io.BytesIO(build_single_report(try_res, "ROI", viz="table", result_usd=usd_res)))
    assert not any(h and "($)" in h for h in _headers(_data_ws(wb)))


# --------------------------------------------------------------------------- #
# Part C — chart auto-sizing
# --------------------------------------------------------------------------- #
def _inmem_bar_width(n):
    wb = Workbook(); wb.remove(wb.active)
    cols = [_col("cat", "dimension", "enum", "Kategori"), _col("cnt", "metric", "number", "Adet")]
    rows = [{"dims": {"cat": f"K{i:02d}"}, "metrics": {"cnt": 100 * (n - i)}, "deltas": None} for i in range(n)]
    ref = data_sheet(wb, _result(cols, rows, {"metrics": {"cnt": 1}}), "V", set())
    ozet = wb.create_sheet("Özet", 0)
    add_chart(ozet, ref, "bar", "K", "A10")
    return ozet._charts[0].width


def test_chart_width_scales_with_category_count():
    w_small = _inmem_bar_width(5)
    w_large = _inmem_bar_width(12)
    assert w_large > w_small                       # more categories → wider chart
    assert 16.0 <= w_small <= 40.0 and 16.0 <= w_large <= 40.0   # within the clamp


def test_stacked_charts_do_not_overlap():
    cols_a = [_col("cat", "dimension", "enum", "Kategori"), _col("cnt", "metric", "number", "Adet")]
    rows_a = [{"dims": {"cat": f"K{i}"}, "metrics": {"cnt": 10 - i}, "deltas": None} for i in range(6)]
    cols_b = [_col("ay", "dimension", "date", "Ay"), _col("cost_try", "metric", "currency", "Maliyet")]
    rows_b = [{"dims": {"ay": f"2026-{m:02d}"}, "metrics": {"cost_try": m * 1000}, "deltas": None} for m in range(1, 13)]
    wb = load_workbook(io.BytesIO(build_workbook(
        [_table("a", "Kat", ["cnt"], ["cat"]), _table("b", "Ay", ["cost_try"], ["ay"])],
        {"a": _result(cols_a, rows_a, {"metrics": {"cnt": 1}}),
         "b": _result(cols_b, rows_b, {"metrics": {"cost_try": 1}})}, "T")))
    charts = wb["Özet"]._charts
    assert len(charts) == 2
    rows = sorted(c.anchor._from.row for c in charts)
    # the second chart anchors well below the first (a chart is ~15 rows tall) → no overlap.
    assert rows[1] - rows[0] >= 14


# --------------------------------------------------------------------------- #
# Part A — USD companion correctness on real data (integration)
# --------------------------------------------------------------------------- #
def _login(client, seed, label="a"):
    client.login(seed[label]["users"][ROLE_DIRECTOR])
    return seed[label]["project"], seed[label]["users"][ROLE_DIRECTOR].id


def _cost_usd(db, p, amount_try, amount_usd, uid, d=date(2026, 1, 10), cat="material_steel"):
    at = D(str(amount_try))
    db.add(CostEntry(
        project_id=p.id, company_id=p.company_id, entry_date=d, cost_category=cat,
        amount_try=at, vat_amount_try=D("0"), total_with_vat_try=at,
        amount_usd=D(str(amount_usd)), fx_rate_usd=D("33.3333"),
        payment_status="unpaid", entry_type="actual", created_by=uid,
    ))
    db.commit()


def test_run_spec_usd_cost_try_uses_cost_usd_not_try(client, db, seed):
    from app.services.studio.engine import run_spec, run_spec_usd
    p, uid = _login(client, seed)
    _cost_usd(db, p, "100000", "3000", uid, cat="material_steel")
    _cost_usd(db, p, "50000", "1500", uid, cat="labor")
    spec = {"metrics": ["cost_try"], "dimensions": ["cost_category"], "viz": "table"}
    try_res = run_spec(db, p.company_id, spec)
    usd_res = run_spec_usd(db, p.company_id, spec)
    assert usd_res["meta"]["currency"] == "usd"
    # keyed by the SAME id (cost_try) but carrying the real USD-at-date sum (4500), NOT ₺.
    assert float(try_res["totals"]["metrics"]["cost_try"]) == 150000.0
    assert float(usd_res["totals"]["metrics"]["cost_try"]) == 4500.0


def test_run_spec_usd_both_cost_try_and_cost_usd_no_double_count(client, db, seed):
    # A report can select BOTH "Maliyet (₺)" and "Maliyet ($)". The companion must keep
    # both ids with the REAL USD (3000) — never double-count (6000) or drop a column.
    from app.services.studio.engine import run_spec_usd
    p, uid = _login(client, seed)
    _cost_usd(db, p, "100000", "3000", uid, cat="material_steel")
    spec = {"metrics": ["cost_try", "cost_usd"], "dimensions": ["cost_category"], "viz": "table"}
    usd = run_spec_usd(db, p.company_id, spec)
    tot = usd["totals"]["metrics"]
    assert float(tot["cost_try"]) == 3000.0    # ₺-locked cost_try's USD = its cost_usd sibling
    assert float(tot["cost_usd"]) == 3000.0     # real cost_usd kept, NOT collapsed
    assert "cost_try" in tot and "cost_usd" in tot   # both distinct, neither 6000


def test_already_usd_column_not_paired_again(client, db, seed):
    # An already-USD column (cost_usd "Maliyet ($)") must NOT get a redundant "($) ($)"
    # pair; only the ₺ cost_try column is paired.
    from app.services.studio.engine import run_both_currencies
    p, uid = _login(client, seed)
    _cost_usd(db, p, "100000", "3000", uid, cat="material_steel")
    _cost_usd(db, p, "50000", "1500", uid, cat="labor")
    spec = {"metrics": ["cost_try", "cost_usd"], "dimensions": ["cost_category"], "viz": "table"}
    t, u = run_both_currencies(db, p.company_id, spec)
    wb = load_workbook(io.BytesIO(build_single_report(t, "Maliyet", viz="table", result_usd=u)))
    headers = [h for h in _headers(_data_ws(wb)) if isinstance(h, str)]
    assert "Maliyet (₺)" in headers and "Maliyet ($)" in headers
    assert not any("($) ($)" in h for h in headers)          # no redundant double-pair
    assert headers.count("Maliyet ($)") == 2                 # cost_usd primary + cost_try's pair


def test_run_both_currencies_pairs_try_and_usd(client, db, seed):
    from app.services.studio.engine import run_both_currencies
    p, uid = _login(client, seed)
    _cost_usd(db, p, "100000", "3000", uid)
    t, u = run_both_currencies(db, p.company_id, {"metrics": ["cost_try"], "dimensions": ["cost_category"]})
    assert t["meta"]["currency"] == "try" and u["meta"]["currency"] == "usd"
    assert float(t["totals"]["metrics"]["cost_try"]) == 100000.0
    assert float(u["totals"]["metrics"]["cost_try"]) == 3000.0


def test_run_spec_usd_nulls_cash_metrics(client, db, seed):
    # Cash has no USD in the engine → the companion nulls it (renders "–"), never ₺-as-$.
    from app.services.studio.engine import run_spec_usd
    p, uid = _login(client, seed)
    _cost_usd(db, p, "100000", "3000", uid)   # produces a cashflow outflow month
    usd_res = run_spec_usd(db, p.company_id, {"metrics": ["net_cash"], "dimensions": ["month"], "viz": "table"})
    assert usd_res["totals"]["metrics"].get("net_cash") is None
    assert all(r["metrics"].get("net_cash") is None for r in usd_res["rows"])


def test_end_to_end_xlsx_shows_real_usd(client, db, seed):
    from app.services.studio.engine import run_both_currencies
    p, uid = _login(client, seed)
    _cost_usd(db, p, "100000", "3000", uid, cat="material_steel")
    _cost_usd(db, p, "50000", "1500", uid, cat="labor")
    spec = {"metrics": ["cost_try"], "dimensions": ["cost_category"], "viz": "table"}
    t, u = run_both_currencies(db, p.company_id, spec)
    wb = load_workbook(io.BytesIO(build_single_report(t, "Maliyet", viz="table", result_usd=u)))
    ws = _data_ws(wb)
    assert any(h and "($)" in h for h in _headers(ws))          # paired USD column
    all_vals = [c.value for s in wb.worksheets for row in s.iter_rows() for c in row]
    assert 4500 in all_vals or 4500.0 in all_vals              # the REAL USD total (not ₺ 150000)
    assert any("USD değerleri" in s and "CR-014" in s for s in _ozet_strings(wb))
