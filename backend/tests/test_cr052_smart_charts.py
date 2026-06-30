"""CR-052 — smart charts by data shape.

Proves the Excel engine adds a native chart to a report table ONLY when the data
shape earns it (trend / comparison / composition), never one-per-table:

  * ``pick_chart`` selects ``line|area|bar|clustered_bar|pie|None`` by the result's
    columns + rows (date → line/clustered-bar; category → bar/pie; ≤3-row / snapshot /
    multi-dim → None); an explicit ``viz`` is honoured unless the shape forbids it;
  * a monthly table earns a line (1 metric) / clustered bar (≥2 metrics); a category
    breakdown earns a horizontal bar (or a pie when ≤6 part-to-whole ₺ slices);
  * a 2-row table earns NO chart; a chart widget and its twin table are not both
    charted (de-dup by metric+dim signature);
  * charts reference the data sheet's CHRONOLOGICAL ranges (CR-050) so lines read
    left→right in time; a category chart caps its categories (~12) for readability;
  * ``_safe_cell`` still neutralises injection on every label that feeds a chart, and
    the workbook opens with zero chart/formula errors;
  * the agent's authoring guidance lightly steers viz choice + a few high-value charts.

Engine behaviour is unit-tested on synthetic ``run_spec`` results (read-only, no DB),
exactly as CR-046 does, so the shapes are owned by the test.
"""
import io

import pytest
from openpyxl import load_workbook

from app.services.studio.excel_report import (
    build_single_report, build_workbook, pick_chart,
)


# --------------------------------------------------------------------------- #
# Column / result helpers
# --------------------------------------------------------------------------- #
def _col(cid, kind, ctype, label):
    return {"id": cid, "kind": kind, "type": ctype, "label": label}


def _date(cid="month", ctype="date"):
    return _col(cid, "dimension", ctype, "Ay")


def _cat(cid="cat", ctype="enum"):
    return _col(cid, "dimension", ctype, "Kategori")


def _metric(cid="cost_try", ctype="currency"):
    return _col(cid, "metric", ctype, cid)


def _result(columns, rows, totals, *, frm="2026-01-01", to="2026-12-31"):
    return {
        "columns": columns, "rows": rows, "totals": totals,
        "meta": {"date_range": {"from": frm, "to": to},
                 "comparison": None, "comparison_unit": None},
    }


def _table(wid, title, metrics, dims, viz="table"):
    return {"id": wid, "type": "table", "title": title,
            "spec": {"metrics": metrics, "dimensions": dims, "viz": viz}}


def _chart_w(wid, title, metrics, dims, viz="line"):
    return {"id": wid, "type": "chart", "title": title,
            "spec": {"metrics": metrics, "dimensions": dims, "viz": viz}}


def _build(widgets, results, title="Test"):
    return load_workbook(io.BytesIO(build_workbook(widgets, results, title)))


def _charts(wb):
    return wb["Özet"]._charts


def _classes(wb):
    return [type(c).__name__ for c in _charts(wb)]


def _cat_ref(chart):
    cat = chart.series[0].cat
    if cat is None:
        return None
    if cat.strRef:
        return cat.strRef.f
    if cat.numRef:
        return cat.numRef.f
    return None


# --------------------------------------------------------------------------- #
# 1) pick_chart — selection by data shape
# --------------------------------------------------------------------------- #
def _rows(n):
    return [{} for _ in range(n)]  # only the COUNT matters to the selector


@pytest.mark.parametrize("cols,n,viz,expected", [
    # date dimension → trend
    ([_date(), _metric()], 5, "line", "line"),
    ([_date(), _metric()], 5, None, "line"),
    ([_date(), _metric()], 5, "table", "line"),
    ([_date(), _metric()], 6, "area", "area"),                 # honour cumulative/area
    ([_date(), _metric()], 6, "bar", "clustered_bar"),         # honour column-over-time
    ([_date(), _metric("gelir"), _metric("gider")], 6, "line", "clustered_bar"),
    ([_date(), _metric()], 1, "line", None),                   # single point → none
    # one category dimension → comparison / composition
    ([_cat(), _metric()], 8, "table", "bar"),                  # >6 ₺ rows → ranking bar
    ([_cat(), _metric()], 5, "table", "pie"),                  # ≤6 part-to-whole → pie
    ([_cat(), _metric()], 6, "bar", "pie"),                    # upgrade bar → pie when it fits
    ([_cat(), _metric("margin", "percent")], 5, "bar", "bar"),  # percent ≠ part-to-whole
    ([_cat(), _metric("butce"), _metric("gercek")], 4, "table", "clustered_bar"),
    ([_cat(), _metric()], 3, "table", None),                   # ≤3 rows → none
    ([_cat(), _metric()], 2, "line", None),                    # shape forbids explicit viz
    # shapes a chart only muddles → none
    ([_metric()], 1, "kpi", None),                             # snapshot single value, 0 dim
    ([_date(), _cat(), _metric()], 8, "table", None),          # cross-tab (2 dims)
    ([_date()], 8, "line", None),                              # no metric
])
def test_pick_chart_by_shape(cols, n, viz, expected):
    assert pick_chart(cols, _rows(n), viz) == expected


def test_pick_chart_tolerates_none_rows():
    assert pick_chart([_date(), _metric()], None, "line") is None


# --------------------------------------------------------------------------- #
# 2) A table earns the right chart by shape
# --------------------------------------------------------------------------- #
def test_monthly_table_earns_line_chart():
    cols = [_date(), _metric("cost_try")]
    rows = [{"dims": {"month": f"2026-{m:02d}"}, "metrics": {"cost_try": m * 1000}}
            for m in range(1, 6)]
    wb = _build([_table("t", "Aylık Maliyet", ["cost_try"], ["month"])],
                {"t": _result(cols, rows, {"metrics": {"cost_try": 15000}})})
    assert "LineChart" in _classes(wb)


def test_monthly_two_metrics_earns_clustered_bar():
    cols = [_date(), _metric("gelir"), _metric("gider")]
    rows = [{"dims": {"month": f"2026-{m:02d}"}, "metrics": {"gelir": m * 200, "gider": m * 150}}
            for m in range(1, 6)]
    wb = _build([_table("t", "Aylık Gelir-Gider", ["gelir", "gider"], ["month"])],
                {"t": _result(cols, rows, {"metrics": {"gelir": 3000, "gider": 2250}})})
    bars = [c for c in _charts(wb) if type(c).__name__ == "BarChart"]
    assert bars and bars[0].type == "col" and bars[0].grouping == "clustered"
    assert len(bars[0].series) == 2                       # gelir + gider side-by-side
    assert bars[0].legend is not None                     # multi-series keeps a legend


def test_category_table_earns_horizontal_bar():
    cols = [_cat("vendor"), _metric("cost_try")]
    rows = [{"dims": {"vendor": f"V{i}"}, "metrics": {"cost_try": 1000 * (12 - i)}}
            for i in range(8)]                            # 8 ₺ rows → ranking bar
    wb = _build([_table("t", "Tedarikçi Harcaması", ["cost_try"], ["vendor"])],
                {"t": _result(cols, rows, {"metrics": {"cost_try": 99999}})})
    bars = [c for c in _charts(wb) if type(c).__name__ == "BarChart"]
    assert bars and bars[0].type == "bar"                 # horizontal ranking


def test_small_part_to_whole_earns_pie():
    cols = [_cat("kategori"), _metric("cost_try")]
    rows = [{"dims": {"kategori": k}, "metrics": {"cost_try": v}}
            for k, v in [("Beton", 5), ("Çelik", 4), ("İşçilik", 3), ("Tesisat", 2), ("Diğer", 1)]]
    wb = _build([_table("t", "Gider Dağılımı", ["cost_try"], ["kategori"])],
                {"t": _result(cols, rows, {"metrics": {"cost_try": 15}})})
    pies = [c for c in _charts(wb) if type(c).__name__ == "PieChart"]
    assert pies
    # Every slice carries a Heneka palette fill (guards the DataPoint spPr kwarg —
    # passing graphicalProperties= silently drops all slice colors on openpyxl 3.x).
    pts = pies[0].series[0].data_points
    assert len(pts) == 5
    fills = {dp.graphicalProperties.solidFill.srgbClr for dp in pts
             if dp.graphicalProperties and dp.graphicalProperties.solidFill}
    assert "183047" in fills and len(fills) >= 3       # NAVY + ≥2 other Heneka colors


def test_two_row_table_gets_no_chart():
    # A "Daire Satış tek tip" 2-row table: a chart restates the table → none.
    cols = [_cat("tip"), _metric("cost_try")]
    rows = [{"dims": {"tip": "A"}, "metrics": {"cost_try": 5}},
            {"dims": {"tip": "B"}, "metrics": {"cost_try": 3}}]
    wb = _build([_table("t", "Daire Satış (tek tip)", ["cost_try"], ["tip"])],
                {"t": _result(cols, rows, {"metrics": {"cost_try": 8}})})
    assert _charts(wb) == []


def test_explicit_chart_widget_overridden_to_none_by_shape():
    # An explicit chart widget is honoured ONLY when the shape supports it.
    cols = [_cat("tip"), _metric("cost_try")]
    rows = [{"dims": {"tip": "A"}, "metrics": {"cost_try": 5}},
            {"dims": {"tip": "B"}, "metrics": {"cost_try": 3}}]
    wb = _build([_chart_w("c", "Çok Az", ["cost_try"], ["tip"], viz="bar")],
                {"c": _result(cols, rows, {"metrics": {"cost_try": 8}})})
    assert _charts(wb) == []


# --------------------------------------------------------------------------- #
# 3) De-dup — a chart widget and its twin table are not both charted
# --------------------------------------------------------------------------- #
def test_no_widget_double_charted():
    cols = [_date(), _metric("cost_try")]
    rows = [{"dims": {"month": f"2026-{m:02d}"}, "metrics": {"cost_try": m * 1000}}
            for m in range(1, 6)]
    r = _result(cols, rows, {"metrics": {"cost_try": 15000}})
    widgets = [
        _chart_w("c", "Aylık Grafik", ["cost_try"], ["month"], viz="line"),
        _table("t", "Aylık Tablo", ["cost_try"], ["month"]),
    ]
    wb = _build(widgets, {"c": r, "t": r})
    assert len(_charts(wb)) == 1                           # the twin table is not re-charted


def test_same_metric_dim_different_filter_both_charted():
    # Two chart widgets, same metric+dim but DIFFERENT filters (project A vs B) are NOT
    # twins — both must be charted (the de-dup only collapses identical-data widgets).
    cols = [_date(), _metric("cost_try")]
    rows = [{"dims": {"month": f"2026-{m:02d}"}, "metrics": {"cost_try": m * 1000}}
            for m in range(1, 6)]
    r = _result(cols, rows, {"metrics": {"cost_try": 15000}})
    a = _chart_w("a", "Proje A Aylık", ["cost_try"], ["month"], viz="line")
    a["spec"]["filters"] = [{"field": "project", "op": "=", "value": "AAA"}]
    b = _chart_w("b", "Proje B Aylık", ["cost_try"], ["month"], viz="line")
    b["spec"]["filters"] = [{"field": "project", "op": "=", "value": "BBB"}]
    wb = _build([a, b], {"a": r, "b": r})
    assert len(_charts(wb)) == 2


# --------------------------------------------------------------------------- #
# 4) Charts sit on chronological ranges (CR-050) + category cap
# --------------------------------------------------------------------------- #
def test_chart_references_chronological_range():
    cols = [_date(), _metric("cost_try")]
    months = [f"2026-{m:02d}" for m in range(1, 6)]        # ascending (CR-050)
    rows = [{"dims": {"month": mo}, "metrics": {"cost_try": i * 1000}}
            for i, mo in enumerate(months, 1)]
    wb = _build([_table("t", "Aylık", ["cost_try"], ["month"], viz="line")],
                {"t": _result(cols, rows, {"metrics": {"cost_try": 15000}})})
    data_ws = [wb[n] for n in wb.sheetnames if n != "Özet"][0]
    dim_vals = [data_ws.cell(row=r, column=1).value for r in range(2, 2 + len(months))]
    assert dim_vals == months == sorted(months)           # the line walks time ascending
    ref = _cat_ref(_charts(wb)[0])
    assert f"$A$2:$A${1 + len(months)}" in ref            # full span, in order, no Toplam row


def test_category_chart_caps_categories():
    cols = [_cat(), _metric("cost_try")]
    rows = [{"dims": {"cat": f"K{i:02d}"}, "metrics": {"cost_try": 1000 * (40 - i)}}
            for i in range(20)]                           # 20 categories → bar, capped
    wb = _build([_table("t", "Kategori", ["cost_try"], ["cat"])],
                {"t": _result(cols, rows, {"metrics": {"cost_try": 1}})})
    ref = _cat_ref(_charts(wb)[0])
    assert ref.endswith("$A$13")                          # capped to 12 categories (rows 2..13)


# --------------------------------------------------------------------------- #
# 5) The DGN-shaped report — decision-first, opens clean
# --------------------------------------------------------------------------- #
def test_dgn_report_mixed_charts_open_clean():
    # cumulative cash → ascending line; monthly gelir-vs-gider → clustered bar;
    # category gider (>6) → horizontal bar. One curated set, zero chart/formula errors.
    cash_cols = [_date(), _metric("cum_cash")]
    cash_rows = [{"dims": {"month": f"2019-{m:02d}"}, "metrics": {"cum_cash": m * 100000}}
                 for m in range(1, 8)]
    gg_cols = [_date(), _metric("gelir"), _metric("gider")]
    gg_rows = [{"dims": {"month": f"2019-{m:02d}"}, "metrics": {"gelir": m * 200000, "gider": m * 150000}}
               for m in range(1, 6)]
    cat_cols = [_cat("kategori"), _metric("cost_try")]
    cat_rows = [{"dims": {"kategori": f"K{i:02d}"}, "metrics": {"cost_try": 1000 * (20 - i)}}
                for i in range(9)]                        # 9 categories → bar
    widgets = [
        _table("cash", "Kümülatif Nakit", ["cum_cash"], ["month"], viz="line"),
        _table("gg", "Aylık Gelir-Gider", ["gelir", "gider"], ["month"]),
        _table("cat", "Gider Kategorileri", ["cost_try"], ["kategori"]),
    ]
    results = {
        "cash": _result(cash_cols, cash_rows, {"metrics": {"cum_cash": 700000}}, frm="2019-01-01", to="2019-07-31"),
        "gg": _result(gg_cols, gg_rows, {"metrics": {"gelir": 3000000, "gider": 2250000}}, frm="2019-01-01", to="2019-05-31"),
        "cat": _result(cat_cols, cat_rows, {"metrics": {"cost_try": 99999}}, frm="2019-01-01", to="2019-12-31"),
    }
    wb = _build(widgets, results, "DGN Aylık Rapor")

    classes = _classes(wb)
    assert "LineChart" in classes                         # cumulative cash, ascending
    cluster = [c for c in _charts(wb)
               if type(c).__name__ == "BarChart" and c.grouping == "clustered"
               and c.type == "col" and len(c.series) == 2]
    rankbar = [c for c in _charts(wb) if type(c).__name__ == "BarChart" and c.type == "bar"]
    assert cluster, "monthly gelir-vs-gider should be a clustered bar"
    assert rankbar, "category gider should be a horizontal bar"

    # Opens with zero chart/formula errors; every series points into a data sheet.
    assert all(c.data_type != "e" for s in wb.worksheets for row in s.iter_rows() for c in row)
    assert all(c.data_type != "f" for s in wb.worksheets for row in s.iter_rows() for c in row)
    data_titles = {n for n in wb.sheetnames if n != "Özet"}
    for ch in _charts(wb):
        for series in ch.series:
            ref = series.val.numRef.f if series.val and series.val.numRef else ""
            assert ref == "" or any(t in ref for t in data_titles)


def test_safe_cell_applied_to_chart_labels():
    # A formula-injection slice label must still be neutralised even though the table
    # now earns a (pie) chart whose categories reference that very cell.
    payload = "=cmd|'/c calc'!A1"
    cols = [_cat("kategori"), _metric("cost_try")]
    rows = [{"dims": {"kategori": payload if i == 0 else f"K{i}"},
             "metrics": {"cost_try": 1000 * (8 - i)}} for i in range(6)]  # 6 ₺ rows → pie
    wb = _build([_table("t", "Enjeksiyon", ["cost_try"], ["kategori"])],
                {"t": _result(cols, rows, {"metrics": {"cost_try": 21000}})})
    cells = [c.value for s in wb.worksheets for row in s.iter_rows()
             for c in row if isinstance(c.value, str)]
    assert any(v.startswith("'=cmd") for v in cells)      # neutralised to text
    assert not any(v[:1] in ("=", "+", "-", "@") for v in cells)
    assert "PieChart" in _classes(wb)                     # …and the chart was still earned


def test_single_report_table_earns_chart_by_shape():
    # CR-052 extends "tables earn charts" to the single-report Özet-lite too.
    cols = [_cat("kategori"), _metric("cost_try")]
    rows = [{"dims": {"kategori": f"K{i}"}, "metrics": {"cost_try": 1000 * (10 - i)}}
            for i in range(8)]                            # 8 ₺ rows → bar
    wb = load_workbook(io.BytesIO(build_single_report(
        _result(cols, rows, {"metrics": {"cost_try": 1}}), "Kategori Raporu", viz="table")))
    assert "BarChart" in [type(c).__name__ for c in wb["Özet"]._charts]


# --------------------------------------------------------------------------- #
# 6) The agent's authoring guidance lightly steers chart choice
# --------------------------------------------------------------------------- #
def test_authoring_guidance_steers_chart_choice():
    from app.services import agent

    g = agent._ACTION_GUIDANCE
    assert "GÖRSEL" in g or "GRAFİK" in g
    assert "line" in g and "bar" in g
    assert "AZ SAYIDA" in g                               # few high-value charts, not one per section

    rep = next(s for s in agent.build_tool_schemas() if s["name"] == "propose_report")
    spec_desc = rep["input_schema"]["properties"]["spec"]["description"]
    assert "line/area" in spec_desc or "şekl" in spec_desc.lower()
