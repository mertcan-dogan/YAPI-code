"""CR-054 — report output fixes: percent format, chart axes, semantic colors.

Proves the three rendering fixes on the decision-grade Excel engine:

  * **Percent format (%2842 bug):** a percent-type metric VALUE (28.42, percent-units)
    renders with the literal ``0.0"%"`` (NO Excel ×100), so a margin KPI reads %28,4 —
    never %2842; a pct DELTA (0.15, a fraction) keeps ``0.0%`` (Excel ×100 → %15,0).
  * **Chart axes (blank-axis bug):** every generated chart un-hides both axes
    (``x_axis.delete is False`` / ``y_axis.delete is False`` — openpyxl's ``delete=None``
    trap that made Excel hide the ticks), titles both axes (unit on the value axis, the
    dimension on the category axis), and the value axis carries the type-correct numFmt —
    ₺ for currency, ``0.0"%"`` for percent (asserted on the ``area`` path too).
  * **Semantic colors:** a single-series SIGNED chart (net nakit / a variance, dipping
    below zero) colors points green ≥0 / red <0; a categorical / all-positive single
    series keeps the Heneka palette (navy). The signed colors are asserted so they can't
    silently regress (the CR-052 pie lesson).

Axis details are asserted on the engine's IN-MEMORY chart (via ``add_chart``) because
openpyxl's *reader* cannot reattach an AreaChart's axes on reload (the axes ARE written
to the XML Excel reads — proven by the title round-trip — but ``load_workbook`` returns
a fresh default axis). Colours are asserted on the round-tripped workbook, exactly as
CR-052 does. All shapes are synthetic ``run_spec`` results (read-only, no DB).
"""
import io

from openpyxl import Workbook, load_workbook

from app.services.studio.excel_report import (
    add_chart, build_single_report, build_workbook, data_sheet,
)

# Heneka hex (report_theme) — no alpha, uppercased (openpyxl solidFill.srgbClr form).
GREEN, RED, NAVY = "27815F", "B94D45", "183047"
FMT_TRY = '#,##0" ₺"'
FMT_PCT_VALUE = '0.0"%"'
FMT_PCT_FRACTION = "0.0%"


# --------------------------------------------------------------------------- #
# Result / widget helpers (mirror CR-052)
# --------------------------------------------------------------------------- #
def _col(cid, kind, ctype, label):
    return {"id": cid, "kind": kind, "type": ctype, "label": label}


def _date(cid="month"):
    return _col(cid, "dimension", "date", "Ay")


def _cat(cid="cat", label="Kategori"):
    return _col(cid, "dimension", "enum", label)


def _metric(cid, ctype, label=None):
    return _col(cid, "metric", ctype, label or cid)


def _result(columns, rows, totals, *, comparison=None, comparison_unit=None):
    return {
        "columns": columns, "rows": rows, "totals": totals,
        "meta": {"date_range": {"from": "2026-01-01", "to": "2026-12-31"},
                 "comparison": comparison, "comparison_unit": comparison_unit},
    }


def _table(wid, title, metrics, dims, viz="table"):
    return {"id": wid, "type": "table", "title": title,
            "spec": {"metrics": metrics, "dimensions": dims, "viz": viz}}


def _inmem_chart(result, kind, title="Grafik"):
    """The engine's IN-MEMORY chart for a given kind — exactly what add_chart builds
    (and what Excel reads), independent of openpyxl's reader quirks."""
    wb = Workbook(); wb.remove(wb.active)
    ref = data_sheet(wb, result, "Veri", set())
    ozet = wb.create_sheet("Özet", 0)
    add_chart(ozet, ref, kind, title, "A10")
    return ozet._charts[0]


def _charts(widgets, results):
    wb = load_workbook(io.BytesIO(build_workbook(widgets, results, "Test")))
    return wb, wb["Özet"]._charts


def _fmt_code(axis):
    nf = axis.numFmt
    return getattr(nf, "formatCode", nf)


def _title_text(axis):
    try:
        return axis.title.tx.rich.p[0].r[0].t
    except Exception:  # pragma: no cover - defensive
        return None


def _assert_axes(ch, *, value_fmt, value_title, cat_title=None):
    # CR-054 — both axes un-hidden (else Excel hides the ticks), titled, and the value
    # axis (always y_axis / NumericAxis) carries the type-correct number format.
    assert ch.x_axis.delete is False, "category axis must be un-hidden"
    assert ch.y_axis.delete is False, "value axis must be un-hidden"
    assert _title_text(ch.y_axis) == value_title       # unit on the value axis
    assert _title_text(ch.x_axis) is not None          # dimension on the category axis
    if cat_title is not None:
        assert _title_text(ch.x_axis) == cat_title
    assert _fmt_code(ch.y_axis) == value_fmt


def _numeric_fmts(wb, value):
    return [c.number_format for s in wb.worksheets for row in s.iter_rows()
            for c in row if c.value == value]


# --------------------------------------------------------------------------- #
# 1) Percent format — no ×100 double-scaling on VALUES; fractions keep ×100
# --------------------------------------------------------------------------- #
def test_percent_value_uses_literal_percent_not_x100():
    # A percent-type metric VALUE (28.42, already percent-units) must format 0.0"%",
    # NEVER 0.0% (which Excel multiplies by 100 → the live %2842 bug).
    cols = [_cat("proje", "Proje"), _metric("margin_pct_current", "percent", "Güncel kâr marjı")]
    rows = [{"dims": {"proje": "A"}, "metrics": {"margin_pct_current": 28.42}, "deltas": None}]
    totals = {"metrics": {"margin_pct_current": 28.42}, "deltas": None}
    wb = load_workbook(io.BytesIO(build_single_report(_result(cols, rows, totals), "Marj", viz="table")))

    fmts = _numeric_fmts(wb, 28.42)
    assert fmts, "the 28.42 percent value should appear (KPI card + table cell)"
    assert all(f == FMT_PCT_VALUE for f in fmts)         # literal % — renders %28,4
    assert not any(f == FMT_PCT_FRACTION for f in fmts)  # never the ×100 format → %2842


def test_pct_delta_still_uses_x100_percent_format():
    # A pct DELTA is a FRACTION (0.15) → keep 0.0% so Excel ×100 → %15,0 (unchanged).
    cols = [_cat("proje", "Proje"), _metric("revenue", "currency", "Gelir")]
    rows = [
        {"dims": {"proje": "A"}, "metrics": {"revenue": 100}, "deltas": {"revenue": 0.15}},
        {"dims": {"proje": "B"}, "metrics": {"revenue": 80}, "deltas": {"revenue": 0.15}},
    ]
    totals = {"metrics": {"revenue": 180}, "deltas": {"revenue": 0.15}}
    result = _result(cols, rows, totals,
                     comparison={"from": "2025-01-01", "to": "2025-02-28"}, comparison_unit="pct")
    wb = load_workbook(io.BytesIO(build_single_report(result, "Gelir", viz="table")))

    dfmts = _numeric_fmts(wb, 0.15)
    assert dfmts, "the 0.15 pct delta should appear in the Δ column"
    assert all(f == FMT_PCT_FRACTION for f in dfmts)     # fraction → 0.0% (×100)
    assert not any(f == FMT_PCT_VALUE for f in dfmts)


def test_percent_kpi_value_is_literal_percent():
    # The specific live symptom on the ROI/marj KPI: a percent-unit value (39.7) must
    # carry the literal-% format on its cell (→ %39,7, not %3970).
    cols = [_cat("proje", "Proje"), _metric("roi", "percent", "ROI")]
    rows = [{"dims": {"proje": "A"}, "metrics": {"roi": 39.7}, "deltas": None}]
    totals = {"metrics": {"roi": 39.7}, "deltas": None}
    wb = load_workbook(io.BytesIO(build_single_report(_result(cols, rows, totals), "ROI", viz="kpi")))
    fmts = _numeric_fmts(wb, 39.7)
    assert fmts and all(f == FMT_PCT_VALUE for f in fmts)


# --------------------------------------------------------------------------- #
# 2) Chart axes — un-hidden, titled, type-correct value numFmt (incl. area)
# --------------------------------------------------------------------------- #
def test_line_chart_axes_currency():
    cols = [_date(), _metric("cost_try", "currency", "Maliyet (₺)")]
    rows = [{"dims": {"month": f"2026-{m:02d}"}, "metrics": {"cost_try": m * 1000}} for m in range(1, 6)]
    ch = _inmem_chart(_result(cols, rows, {"metrics": {"cost_try": 15000}}), "line")
    assert type(ch).__name__ == "LineChart"
    _assert_axes(ch, value_fmt=FMT_TRY, value_title="₺", cat_title="Ay")


def test_area_chart_axes_currency():
    # The CR calls out the AREA path specifically — it lacked a value-axis format before.
    cols = [_date(), _metric("cum_cash", "currency", "Kümülatif nakit")]
    rows = [{"dims": {"month": f"2026-{m:02d}"}, "metrics": {"cum_cash": m * 1000}} for m in range(1, 7)]
    ch = _inmem_chart(_result(cols, rows, {"metrics": {"cum_cash": 21000}}), "area")
    assert type(ch).__name__ == "AreaChart"
    _assert_axes(ch, value_fmt=FMT_TRY, value_title="₺", cat_title="Ay")


def test_horizontal_bar_value_axis_is_y_not_x():
    # The latent bug: a horizontal bar set numFmt on x_axis (the CATEGORY axis). The value
    # axis is ALWAYS y_axis — even when the bar is drawn horizontally. Assert ₺ on y_axis.
    cols = [_cat("vendor", "Tedarikçi"), _metric("cost_try", "currency", "Maliyet (₺)")]
    rows = [{"dims": {"vendor": f"V{i}"}, "metrics": {"cost_try": 1000 * (12 - i)}} for i in range(8)]
    ch = _inmem_chart(_result(cols, rows, {"metrics": {"cost_try": 1}}), "bar")
    assert type(ch).__name__ == "BarChart" and ch.type == "bar"
    assert ch.x_axis.delete is False and ch.y_axis.delete is False
    assert _fmt_code(ch.y_axis) == FMT_TRY             # ₺ on the VALUE axis (y)
    assert _title_text(ch.y_axis) == "₺"
    assert _title_text(ch.x_axis) == "Tedarikçi"        # category axis titled by dim


def test_clustered_bar_axes_currency():
    cols = [_date(), _metric("gelir", "currency"), _metric("gider", "currency")]
    rows = [{"dims": {"month": f"2026-{m:02d}"}, "metrics": {"gelir": m * 200, "gider": m * 150}}
            for m in range(1, 6)]
    ch = _inmem_chart(_result(cols, rows, {"metrics": {"gelir": 3000, "gider": 2250}}), "clustered_bar")
    assert type(ch).__name__ == "BarChart" and ch.grouping == "clustered"
    _assert_axes(ch, value_fmt=FMT_TRY, value_title="₺", cat_title="Ay")


def test_percent_chart_value_axis_is_percent():
    # A genuine % chart shows % ticks (0.0"%"), titled "%", not ₺.
    cols = [_cat("proje", "Proje"), _metric("roi", "percent", "ROI")]
    rows = [{"dims": {"proje": f"P{i}"}, "metrics": {"roi": 10.0 + i}} for i in range(5)]
    ch = _inmem_chart(_result(cols, rows, {"metrics": {"roi": 60.0}}), "bar")
    assert type(ch).__name__ == "BarChart"
    assert _fmt_code(ch.y_axis) == FMT_PCT_VALUE
    assert _title_text(ch.y_axis) == "%"
    assert ch.y_axis.delete is False and ch.x_axis.delete is False


def test_axes_survive_save_and_open_clean():
    # The full pipeline: a line + a percent bar in one workbook open with zero chart/
    # formula errors, both charts present. (openpyxl reads back line/bar axes fine; the
    # per-kind detail is asserted in-memory above where the area reader falls short.)
    line_cols = [_date(), _metric("cost_try", "currency", "Maliyet")]
    line_rows = [{"dims": {"month": f"2026-{m:02d}"}, "metrics": {"cost_try": m * 1000}} for m in range(1, 6)]
    pct_cols = [_cat("proje", "Proje"), _metric("roi", "percent", "ROI")]
    pct_rows = [{"dims": {"proje": f"P{i}"}, "metrics": {"roi": 5.0 * i}} for i in range(5)]
    widgets = [_table("l", "Aylık", ["cost_try"], ["month"]), _table("r", "ROI", ["roi"], ["proje"])]
    results = {"l": _result(line_cols, line_rows, {"metrics": {"cost_try": 15000}}),
               "r": _result(pct_cols, pct_rows, {"metrics": {"roi": 50.0}})}
    wb, chs = _charts(widgets, results)
    assert len(chs) == 2
    assert all(c.data_type != "e" for s in wb.worksheets for row in s.iter_rows() for c in row)
    assert all(c.data_type != "f" for s in wb.worksheets for row in s.iter_rows() for c in row)
    for ch in chs:  # line + bar reload their axes cleanly
        assert ch.x_axis.delete is False and ch.y_axis.delete is False


# --------------------------------------------------------------------------- #
# 3) Semantic colors — green/red by sign for signed single-series
# --------------------------------------------------------------------------- #
def test_signed_single_series_line_colored_by_sign():
    # Net nakit over months with deficit months → per-point green ≥0 / red <0.
    cols = [_date(), _metric("net_cash", "currency", "Net nakit")]
    vals = [300, -150, 200, -50, 400]
    rows = [{"dims": {"month": f"2026-{m:02d}"}, "metrics": {"net_cash": v}}
            for m, v in enumerate(vals, 1)]
    _, chs = _charts([_table("t", "Net Nakit", ["net_cash"], ["month"], viz="line")],
                     {"t": _result(cols, rows, {"metrics": {"net_cash": 700}})})
    assert chs and type(chs[0]).__name__ == "LineChart"
    dpts = chs[0].series[0].data_points
    ordered = [dp.graphicalProperties.solidFill.srgbClr for dp in dpts
               if dp.graphicalProperties and dp.graphicalProperties.solidFill]
    assert ordered == [GREEN, RED, GREEN, RED, GREEN]   # +,-,+,-,+
    assert NAVY not in ordered                          # NOT the categorical palette


def test_signed_single_series_bar_colored_by_sign():
    # The same on a horizontal bar (a ± variance ranking) — the mechanism is per-point.
    cols = [_cat("proje", "Proje"), _metric("variance", "currency", "Sapma")]
    vals = [500, -300, 200, -100, 400, -50, 150, 250]    # 8 rows → bar, mixed signs
    rows = [{"dims": {"proje": f"P{i}"}, "metrics": {"variance": v}} for i, v in enumerate(vals)]
    _, chs = _charts([_table("t", "Sapma", ["variance"], ["proje"])],
                     {"t": _result(cols, rows, {"metrics": {"variance": 1050}})})
    bars = [c for c in chs if type(c).__name__ == "BarChart"]
    assert bars and bars[0].type == "bar"
    ordered = [dp.graphicalProperties.solidFill.srgbClr for dp in bars[0].series[0].data_points
               if dp.graphicalProperties and dp.graphicalProperties.solidFill]
    assert ordered == [GREEN, RED, GREEN, RED, GREEN, RED, GREEN, GREEN]


def test_categorical_single_series_keeps_palette():
    # An all-positive category ranking is NOT signed → keep the Heneka palette (navy),
    # never a misleading all-green.
    cols = [_cat("vendor", "Tedarikçi"), _metric("cost_try", "currency")]
    rows = [{"dims": {"vendor": f"V{i}"}, "metrics": {"cost_try": 1000 * (10 - i)}} for i in range(8)]
    _, chs = _charts([_table("t", "Tedarikçi", ["cost_try"], ["vendor"])],
                     {"t": _result(cols, rows, {"metrics": {"cost_try": 1}})})
    bars = [c for c in chs if type(c).__name__ == "BarChart"]
    assert bars
    ser = bars[0].series[0]
    assert ser.graphicalProperties.solidFill.srgbClr == NAVY      # series-level palette
    fills = {dp.graphicalProperties.solidFill.srgbClr for dp in ser.data_points
             if dp.graphicalProperties and dp.graphicalProperties.solidFill}
    assert GREEN not in fills and RED not in fills                # no per-sign override


def test_multi_series_keeps_palette_not_signed():
    # A multi-series chart (gelir vs gider) keeps the palette even with a negative — the
    # per-sign colouring only applies to a SINGLE signed series.
    cols = [_date(), _metric("gelir", "currency"), _metric("gider", "currency")]
    rows = [{"dims": {"month": f"2026-{m:02d}"}, "metrics": {"gelir": m * 200, "gider": -(m * 150)}}
            for m in range(1, 6)]
    _, chs = _charts([_table("t", "Gelir-Gider", ["gelir", "gider"], ["month"])],
                     {"t": _result(cols, rows, {"metrics": {"gelir": 3000, "gider": -2250}})})
    bars = [c for c in chs if type(c).__name__ == "BarChart" and c.grouping == "clustered"]
    assert bars and len(bars[0].series) == 2
    for s in bars[0].series:
        no_sign = {dp.graphicalProperties.solidFill.srgbClr for dp in s.data_points
                   if dp.graphicalProperties and dp.graphicalProperties.solidFill}
        assert GREEN not in no_sign and RED not in no_sign        # palette, not per-sign
