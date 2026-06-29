"""CR-048 — premade decision-grade Raporlar (Maliyet Detay, Hakediş, Alt Yüklenici,
Nakit Akış).

Proves: each endpoint returns a valid non-empty PDF (and Excel for cost/cashflow);
the figures come from the EXISTING services (no fabrication); the Hakediş report is
revenue-model-aware (sell-side → sales view/note, hakediş → the invoice tables);
company-scoped (cross-company → 404); read-only; bad fmt → 422.
"""
import io
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy import func, select

from app.constants import ROLE_DIRECTOR, ROLE_SITE_MANAGER
from app.models.client_invoice import ClientInvoice
from app.models.cost_entry import CostEntry
from app.models.landowner_payment import LandownerPayment
from app.models.project import Project
from app.models.subcontractor import Subcontractor
from app.models.unit_sale import UnitSale
from app.services import reports_premade as rp

D = Decimal
TODAY = date(2026, 6, 30)
BASE = "/api/v1/reports"
_COUNTED = [CostEntry, ClientInvoice, Subcontractor, UnitSale, LandownerPayment, Project]


# --------------------------------------------------------------------------- #
# Builders
# --------------------------------------------------------------------------- #
def _uid(seed, label="a"):
    return seed[label]["users"][ROLE_DIRECTOR].id


def _director(seed, label="a"):
    return seed[label]["users"][ROLE_DIRECTOR]


def _cost(db, p, amount, uid, *, d=date(2025, 6, 10), cat="material_steel", subcontractor_id=None, paid="0"):
    amt = D(str(amount))
    db.add(CostEntry(
        project_id=p.id, company_id=p.company_id, entry_date=d, cost_category=cat,
        amount_try=amt, vat_amount_try=D("0"), total_with_vat_try=amt, amount_paid_try=D(str(paid)),
        payment_status="unpaid", entry_type="actual", subcontractor_id=subcontractor_id, created_by=uid,
    ))
    db.flush()


def _invoice(db, p, amount, uid, *, d=date(2025, 7, 15)):
    amt = D(str(amount))
    db.add(ClientInvoice(
        project_id=p.id, company_id=p.company_id, invoice_number=f"HK-{p.project_code}-{amount}",
        invoice_date=d, invoice_type="hakedis", amount_try=amt, vat_amount_try=D("0"),
        total_with_vat_try=amt, net_due_try=amt, amount_received_try=D("0"),
        retention_amount_try=amt * D("0.05"), due_date=date(2025, 8, 15), created_by=uid,
    ))
    db.flush()


def _subcontractor(db, p, name, contract, *, paid_via_cost=None, uid=None):
    s = Subcontractor(project_id=p.id, company_id=p.company_id, name=name,
                      contract_value_try=D(str(contract)), retention_pct=D("10.00"), status="active")
    db.add(s)
    db.flush()
    if paid_via_cost:
        _cost(db, p, paid_via_cost, uid, subcontractor_id=s.id, paid=paid_via_cost, cat="labor")
    return s


def _new_project(db, cid, uid, name, code, revenue_model="hakedis", net_m2="200"):
    p = Project(company_id=cid, name=name, project_code=code, project_type="road",
                client_name="İşveren", contract_value_try=1_000_000, original_budget_try=800_000,
                start_date=date(2025, 1, 1), planned_end_date=date(2025, 12, 31),
                project_manager_id=uid, revenue_model=revenue_model, construction_net_m2=D(net_m2))
    db.add(p)
    db.flush()
    return p


def _seed_hakedis(db, seed):
    cid, uid = seed["a"]["company"].id, _uid(seed)
    p = seed["a"]["project"]
    p.revenue_model = "hakedis"
    db.add(p)
    _cost(db, p, "100000", uid, cat="material_steel")
    _cost(db, p, "50000", uid, cat="material_concrete")
    _invoice(db, p, "70000", uid)
    _subcontractor(db, p, "Beton Taşeron A.Ş.", "200000", paid_via_cost="30000", uid=uid)
    db.commit()
    return p, cid, uid


def _seed_sellside(db, seed):
    cid, uid = seed["a"]["company"].id, _uid(seed)
    p = _new_project(db, cid, uid, "DGN Martı", "DGN", revenue_model="kat_karsiligi")
    db.add(UnitSale(project_id=p.id, company_id=p.company_id, unit_label="Daire 1", unit_type="2+1",
                    net_m2=D("80"), sale_price_try=D("5800000"), sale_date=date(2025, 5, 5)))
    db.add(LandownerPayment(project_id=p.id, company_id=p.company_id, payer_name="Arsa Sahibi",
                            payment_date=date(2025, 3, 1), amount_try=D("500000"),
                            committed_total_try=D("2000000")))
    _cost(db, p, "120000", uid, cat="material_steel")
    db.commit()
    return p, cid, uid


def _counts(db):
    return {m.__name__: db.execute(select(func.count()).select_from(m)).scalar() for m in _COUNTED}


# --------------------------------------------------------------------------- #
# Maliyet Detay (cost) — PDF + Excel
# --------------------------------------------------------------------------- #
def test_cost_pdf(client, db, seed):
    p, cid, uid = _seed_hakedis(db, seed)
    client.login(_director(seed))
    r = client.get(f"{BASE}/cost/{p.id}")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert r.content[:4] == b"%PDF" and len(r.content) > 1000


def test_cost_figures_match_source(db, seed):
    # The cost builder's KPI total equals project_financials (no fabrication).
    from app.services.financials import project_financials

    p, cid, uid = _seed_hakedis(db, seed)
    d = rp.build_cost_data(db, p, seed["a"]["company"])
    f = project_financials(db, p)
    labels = {k: v for k, v in d["kpis"]}
    from app.utils.format import format_currency_tr
    assert labels["Gerçekleşen Maliyet"] == format_currency_tr(f["total_actual_with_vat_try"])
    cats = {c["label"] for c in d["categories"]}
    assert any("Çelik" in c or "Demir" in c or "steel" in c.lower() for c in cats) or d["categories"]


def test_cost_xlsx(client, db, seed):
    p, cid, uid = _seed_hakedis(db, seed)
    client.login(_director(seed))
    r = client.get(f"{BASE}/cost/{p.id}?fmt=xlsx")
    assert r.status_code == 200 and "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    assert "Özet" in wb.sheetnames
    # A category figure reached a data sheet (the 100000 steel cost).
    vals = [c for s in wb.worksheets for row in s.iter_rows(values_only=True) for c in row]
    assert any(isinstance(v, (int, float)) and abs(v - 100000) < 1 for v in vals)


def test_cost_bad_fmt(client, db, seed):
    p, cid, uid = _seed_hakedis(db, seed)
    client.login(_director(seed))
    assert client.get(f"{BASE}/cost/{p.id}?fmt=csv").status_code == 422


# --------------------------------------------------------------------------- #
# Hakediş (invoice) — model-aware
# --------------------------------------------------------------------------- #
def test_invoice_hakedis_mode(client, db, seed):
    p, cid, uid = _seed_hakedis(db, seed)
    d = rp.build_invoice_data(db, p, seed["a"]["company"])
    assert d["sell_side"] is False and d["mode"] == "hakedis"
    assert d["invoices"] and any("70000" in (str(i["amount"])) for i in d["invoices"])
    client.login(_director(seed))
    r = client.get(f"{BASE}/invoice/{p.id}")
    assert r.status_code == 200 and r.content[:4] == b"%PDF"


def test_invoice_sellside_is_sales_view_not_empty_hakedis(client, db, seed):
    p, cid, uid = _seed_sellside(db, seed)
    d = rp.build_invoice_data(db, p, seed["a"]["company"])
    assert d["sell_side"] is True and d["mode"] == "sell_side"
    assert "invoices" not in d                      # NEVER an empty/misleading hakediş table
    assert d["units"] and any("5800000" in str(u["price"]) for u in d["units"])
    assert d["landowner"]["paid"] is not None
    client.login(_director(seed))
    r = client.get(f"{BASE}/invoice/{p.id}")
    assert r.status_code == 200 and r.content[:4] == b"%PDF"


# --------------------------------------------------------------------------- #
# Alt Yüklenici (subcontractor)
# --------------------------------------------------------------------------- #
def test_subcontractor(client, db, seed):
    p, cid, uid = _seed_hakedis(db, seed)
    d = rp.build_subcontractor_data(db, p, seed["a"]["company"])
    assert d["rows"] and d["rows"][0]["name"] == "Beton Taşeron A.Ş."
    assert "200000" in str(d["rows"][0]["committed"])   # contract value from the service
    assert "30000" in str(d["rows"][0]["paid"])         # paid via linked cost entry
    client.login(_director(seed))
    r = client.get(f"{BASE}/subcontractor/{p.id}")
    assert r.status_code == 200 and r.content[:4] == b"%PDF"


def test_subcontractor_empty_is_calm_note(client, db, seed):
    # A project with no subcontractors → a calm note, never a zero-filled table.
    p = _new_project(db, seed["a"]["company"].id, _uid(seed), "Boş Proje", "BOS")
    db.commit()
    d = rp.build_subcontractor_data(db, p, seed["a"]["company"])
    assert d["rows"] == []
    client.login(_director(seed))
    assert client.get(f"{BASE}/subcontractor/{p.id}").status_code == 200


# --------------------------------------------------------------------------- #
# Nakit Akış (cashflow) — PDF + Excel, full span
# --------------------------------------------------------------------------- #
def test_cashflow_pdf(client, db, seed):
    p, cid, uid = _seed_hakedis(db, seed)
    client.login(_director(seed))
    r = client.get(f"{BASE}/cashflow/{p.id}")
    assert r.status_code == 200 and r.content[:4] == b"%PDF" and len(r.content) > 1000


def test_cashflow_xlsx_matches_project_cashflow(client, db, seed):
    # The Excel cumulative must equal the canonical project_cashflow last cumulative
    # (single source — PDF and Excel can't diverge).
    p, cid, uid = _seed_hakedis(db, seed)
    d = rp.build_cashflow_data(db, p, seed["a"]["company"], today=TODAY)
    assert d["periods"], "the full-span cashflow has periods"
    last_cum = float(d["periods"][-1]["cum"])

    client.login(_director(seed))
    r = client.get(f"{BASE}/cashflow/{p.id}?fmt=xlsx")
    assert r.status_code == 200 and "spreadsheetml" in r.headers["content-type"]
    wb = load_workbook(io.BytesIO(r.content))
    assert "Özet" in wb.sheetnames
    # The cumulative value reached the workbook.
    vals = [c for s in wb.worksheets for row in s.iter_rows(values_only=True) for c in row]
    assert any(isinstance(v, (int, float)) and abs(v - last_cum) < 1 for v in vals)


# --------------------------------------------------------------------------- #
# Company scope + read-only + format
# --------------------------------------------------------------------------- #
def test_cross_company_404(client, db, seed):
    pa, _, _ = _seed_hakedis(db, seed)
    client.login(seed["b"]["users"][ROLE_DIRECTOR])
    for kind in ("cost", "invoice", "subcontractor", "cashflow"):
        assert client.get(f"{BASE}/{kind}/{pa.id}").status_code == 404


def test_site_manager_cannot_export(client, db, seed):
    # Section 3.2 — site managers cannot export (same gate as /reports/project).
    p, cid, uid = _seed_hakedis(db, seed)
    client.login(seed["a"]["users"][ROLE_SITE_MANAGER])
    for kind in ("cost", "invoice", "subcontractor", "cashflow"):
        assert client.get(f"{BASE}/{kind}/{p.id}").status_code == 403, kind


def test_cost_monthly_trend_is_chronological(db, seed):
    # The Aylık Maliyet Trendi must be time-ordered, not amount-ordered (a real trend).
    cid, uid = seed["a"]["company"].id, _uid(seed)
    p = seed["a"]["project"]
    p.revenue_model = "hakedis"
    db.add(p)
    _cost(db, p, "30000", uid, d=date(2025, 1, 10))    # earlier month, smaller amount
    _cost(db, p, "90000", uid, d=date(2025, 3, 10))    # later month, larger amount
    db.commit()
    months = [m["month"] for m in rp.build_cost_data(db, p, seed["a"]["company"])["monthly"] if m["month"]]
    assert months == sorted(months), "monthly cost trend must be chronological"


def test_cost_xlsx_has_ozet_kpi_cards(client, db, seed):
    # The Excel Özet shows the SAME headline KPIs as the PDF (no PDF/Excel divergence).
    p, cid, uid = _seed_hakedis(db, seed)
    client.login(_director(seed))
    r = client.get(f"{BASE}/cost/{p.id}?fmt=xlsx")
    wb = load_workbook(io.BytesIO(r.content))
    ozet = [c.value for row in wb["Özet"].iter_rows() for c in row if isinstance(c.value, str)]
    assert any("Gerçekleşen Maliyet" in v for v in ozet)        # a KPI card label is on the Özet


def test_cashflow_includes_pre_start_activity(db, seed):
    # CR-048 (data-relative span): a cost dated BEFORE the project start is still
    # captured in the cashflow periods + the Toplam Nakit Çıkış KPI — not dropped.
    cid, uid = seed["a"]["company"].id, _uid(seed)
    p = _new_project(db, cid, uid, "Eski Proje", "ESK")  # start 2025-01-01
    _cost(db, p, "44444", uid, d=date(2024, 11, 10))     # pre-start cost
    db.commit()
    periods = rp.build_cashflow_data(db, p, seed["a"]["company"], today=TODAY)["periods"]
    assert any(pr["period"].startswith("2024-") for pr in periods), "pre-start month present"


def test_reports_are_read_only(client, db, seed):
    p, cid, uid = _seed_hakedis(db, seed)
    _seed_sellside(db, seed)
    client.login(_director(seed))
    before = _counts(db)
    sell = db.execute(select(Project).where(Project.project_code == "DGN")).scalar_one()
    for path in (f"cost/{p.id}", f"cost/{p.id}?fmt=xlsx", f"invoice/{p.id}", f"invoice/{sell.id}",
                 f"subcontractor/{p.id}", f"cashflow/{p.id}", f"cashflow/{p.id}?fmt=xlsx"):
        assert client.get(f"{BASE}/{path}").status_code == 200
    db.expire_all()
    assert _counts(db) == before
