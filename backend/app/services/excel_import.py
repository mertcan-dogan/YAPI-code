"""Excel cost import — template, sheet listing, validation (Section 9.1 + CR-002-F)."""
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.constants import COST_CATEGORIES
from app.schemas.common import MIN_DATE, max_future_date

# Eski binary .xls formatı openpyxl tarafından okunamaz — kullanıcıya net mesaj göster.
LEGACY_XLS_MESSAGE = (
    "Eski .xls formatı desteklenmiyor. "
    "Lütfen dosyayı .xlsx veya .xlsm olarak kaydedip tekrar yükleyin."
)


def is_legacy_xls(filename: str | None, exc: Exception | None = None) -> bool:
    """True if the upload looks like a legacy binary .xls (openpyxl cannot read it).

    Detected by filename extension (.xls but not .xlsx/.xlsm) or an
    InvalidFileException raised by openpyxl while opening the bytes.
    """
    if filename and filename.lower().endswith(".xls"):
        return True
    if isinstance(exc, InvalidFileException):
        return True
    return False

# Template columns in order (Section 9.1).
COLUMNS = [
    "Tarih", "Kategori", "Alt Kategori", "Tedarikçi Adı", "Açıklama",
    "Fatura No", "Tutar TRY", "KDV Oranı", "Vade Tarihi", "Ödeme Durumu", "Ödeme Tarihi",
]
LABEL_TO_KEY = {label.lower(): key for key, label in COST_CATEGORIES.items()}

# CR-002-F detection keywords.
HEADER_KEYWORDS = {"tarih", "kategori", "tutar"}
TOTAL_KEYWORDS = ("toplam", "total", "sum", "genel")

# CR-002-F++ : map a file's columns to our logical fields BY HEADER NAME, so an
# upload whose columns are in a different order or use synonyms (e.g. "Harcama
# Tarihi" → tarih, "Gerçek Harcama (TL)" → tutar) still has its rows detected
# instead of being silently skipped. Order matters: the specific date columns
# (vade / ödeme tarihi) claim their header first so the generic "tarih" left for
# the entry date doesn't grab them; the amount field is resolved last and never
# matches a date column. Each field claims the left-most not-yet-claimed header
# containing one of its synonyms.
FIELD_SYNONYMS: list[tuple[str, tuple[str, ...]]] = [
    ("payment_due_date", ("vade",)),
    ("date_paid", ("ödeme tarih", "odeme tarih", "tahsilat tarih")),
    ("entry_date", ("harcama tarih", "fatura tarih", "işlem tarih", "gerçekleşme tarih",
                    "gider tarih", "belge tarih", "tarih")),
    ("cost_category", ("kategori", "gider tür", "masraf tür", "gider kalemi tür")),
    ("subcategory", ("alt kategori", "iş kalemi", "alt başlık", "kalem")),
    ("supplier_name", ("tedarikçi", "satıcı", "firma", "cari")),
    ("invoice_number", ("fatura no", "fatura numara", "belge no", "fiş no")),
    ("description", ("açıklama", "detay", "izah", "not")),
    ("vat_rate", ("kdv",)),
    ("payment_status", ("ödeme durumu", "odeme durumu", "durum")),
    # Amount last + never a date column (a "tarih" header is excluded).
    ("amount_try", ("gerçek harcama", "harcama (tl", "tutar", "harcama", "gider",
                    "maliyet", "bedel", "ödenen", "bütçe")),
]

# Positional template order (Section 9.1) — the fallback when headers don't map.
POSITIONAL_MAP = {
    "entry_date": 0, "cost_category": 1, "subcategory": 2, "supplier_name": 3,
    "description": 4, "invoice_number": 5, "amount_try": 6, "vat_rate": 7,
    "payment_due_date": 8, "payment_status": 9, "date_paid": 10,
}

# Shown when a file has rows but none are importable (wrong schema / sheet).
ERR_NO_IMPORTABLE = (
    "İçe aktarılabilir satır bulunamadı. Sütun başlıkları beklenen biçimle "
    "eşleşmiyor olabilir (gerekli sütunlar: Tarih, Kategori, Tutar). Şablonu "
    "indirip kullanın ya da düzensiz dosyalar için 'AI ile İçe Aktar'ı deneyin."
)
ERR_EMPTY_SHEET = "Seçilen sayfa boş — içe aktarılacak veri yok."


def _build_column_map(header_cells: list) -> dict[str, int]:
    """Map logical field -> column index by matching header text against synonyms.

    Date columns are matched before the amount column so a "...Tarih" header is
    never taken as the amount. Returns only the fields that were found.
    """
    headers = [str(c).strip().lower() if c not in (None, "") else "" for c in header_cells]
    colmap: dict[str, int] = {}
    claimed: set[int] = set()
    for field, synonyms in FIELD_SYNONYMS:
        for syn in synonyms:
            hit = None
            for i, h in enumerate(headers):
                if i in claimed or not h:
                    continue
                # The amount column must not be a date column.
                if field == "amount_try" and "tarih" in h:
                    continue
                if syn in h:
                    hit = i
                    break
            if hit is not None:
                colmap[field] = hit
                claimed.add(hit)
                break
    return colmap


def build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Maliyetler"
    ws.append(COLUMNS)
    ws.append([
        "01.05.2025", "Malzeme — Beton", "C30 hazır beton", "Akçansa", "Saha dökümü",
        "FAT-2025-001", "150000", "20", "31.05.2025", "Ödenmedi", "",
    ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def list_sheet_names(file_bytes: bytes) -> list[str]:
    """CR-002-F: return the sheet names of an uploaded workbook."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    return list(wb.sheetnames)


MAX_AI_ROWS = 500


def excel_to_text(file_bytes: bytes) -> tuple[str, bool, int]:
    """CR-002-H: flatten every sheet to text for the AI analyser.

    Returns (text, truncated, row_count). At most MAX_AI_ROWS data rows across all
    sheets; formula cells are read as their resolved values (data_only=True).
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    lines: list[str] = []
    count = 0
    truncated = False
    for name in wb.sheetnames:
        ws = wb[name]
        lines.append(f"### SAYFA: {name}")
        for row in ws.iter_rows(values_only=True):
            if count >= MAX_AI_ROWS:
                truncated = True
                break
            cells = ["" if c is None else str(c) for c in row]
            if not any(cells):
                continue
            lines.append(" | ".join(cells))
            count += 1
        if truncated:
            break
    return "\n".join(lines), truncated, count


def _parse_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError("Geçersiz tarih")


# CR-006-F: kullanıcıya görünen Türkçe doğrulama mesajları.
ERR_DATE_FORMAT = "Geçersiz tarih formatı — DD.MM.YYYY kullanın"
ERR_AMOUNT_NUMERIC = "Tutar sayısal olmalı — para birimi sembolü girmeyin"
ERR_DATE_ORDER = "Vade tarihi fatura tarihinden önce olamaz"


def _parse_amount(value) -> Decimal:
    if value in (None, ""):
        raise ValueError("Tutar zorunludur")
    s = str(value).replace(".", "").replace(",", ".") if isinstance(value, str) else value
    try:
        d = Decimal(str(s))
    except (InvalidOperation, ValueError):
        raise ValueError(ERR_AMOUNT_NUMERIC)
    if d <= 0:
        raise ValueError("Tutar 0'dan büyük olmalıdır")
    return d


def _looks_like_header(cells: list) -> int:
    """Count how many header keywords appear in a row's cells."""
    text = " ".join(str(c).lower() for c in cells if c not in (None, ""))
    return sum(1 for kw in HEADER_KEYWORDS if kw in text)


def _numeric_ratio(cells: list) -> float:
    """CR-006-F: oran of cells that parse as a number (header rows are mostly text)."""
    values = [c for c in cells if c not in (None, "")]
    if not values:
        return 1.0  # boş satır başlık olamaz
    numeric = 0
    for c in values:
        if isinstance(c, (int, float)):
            numeric += 1
            continue
        try:
            Decimal(str(c).replace(".", "").replace(",", "."))
            numeric += 1
        except (InvalidOperation, ValueError):
            pass
    return numeric / len(values)


def _is_total_row(cells: list) -> bool:
    text = " ".join(str(c).lower() for c in cells if c not in (None, ""))
    return any(kw in text for kw in TOTAL_KEYWORDS)


def validate_rows(file_bytes: bytes, sheet_name: str | None = None) -> dict:
    """Parse + validate a sheet (CR-002-F).

    Returns {header_detected, header_row, column_map, message, rows: [...]}. Each
    row dict has: row, valid, skipped, errors, data. Columns are matched to logical
    fields by HEADER NAME (any order/synonyms) and fall back to the positional
    template when headers don't map. ``message`` carries a human reason when a sheet
    is empty or nothing is importable, so the UI never shows a silent empty preview.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    raw = [list(r) for r in ws.iter_rows(values_only=True)]
    if not any(any(c not in (None, "") for c in r) for r in raw):
        return {"header_detected": False, "header_row": 1, "column_map": {},
                "message": ERR_EMPTY_SHEET, "rows": []}

    # --- Header detection (CR-002-F + CR-006-F) ---
    # 1) scan first 5 rows for >=2 header keywords; 2) fall back to the first row
    #    whose numeric-cell ratio is < %30 (header rows are mostly text labels).
    header_row = None
    for i, cells in enumerate(raw[:5]):
        if _looks_like_header(cells) >= 2:
            header_row = i
            break
    if header_row is None:
        for i, cells in enumerate(raw[:5]):
            if any(c not in (None, "") for c in cells) and _numeric_ratio(cells) < 0.30:
                header_row = i
                break
    header_detected = header_row is not None
    if header_row is None:
        header_row = 0  # assume the first row is the header

    # --- Column mapping (header-name based, positional fallback) ---
    colmap = _build_column_map(raw[header_row])
    # If neither the date nor the amount column could be named, the file is almost
    # certainly in the positional template layout (or has opaque headers): use it.
    if "entry_date" not in colmap and "amount_try" not in colmap:
        colmap = dict(POSITIONAL_MAP)

    def cell(cells: list, field: str):
        i = colmap.get(field)
        return cells[i] if i is not None and i < len(cells) else None

    width = max([len(COLUMNS)] + [i + 1 for i in colmap.values()])

    rows: list[dict] = []
    for idx in range(header_row + 1, len(raw)):
        cells = list(raw[idx]) + [None] * (width - len(raw[idx]))
        # Blank line if the fields we actually read are all empty.
        if all(cell(cells, f) in (None, "") for f in ("entry_date", "cost_category", "amount_try", "description")):
            continue

        # CR-002-F: skip total/summary rows.
        if _is_total_row(cells):
            rows.append({"row": idx + 1, "valid": False, "skipped": True,
                         "errors": ["Bu satır otomatik olarak atlandı: Toplam satırı"], "data": {}})
            continue

        errors: list[str] = []
        parsed: dict = {}

        # Tarih — empty/invalid means this is not a data row -> skip (grey).
        try:
            d = _parse_date(cell(cells, "entry_date"))
        except ValueError:
            d = None
        if d is None:
            rows.append({"row": idx + 1, "valid": False, "skipped": True,
                         "errors": ["Bu satır otomatik olarak atlandı: Tarih boş/geçersiz"], "data": {}})
            continue
        if d < MIN_DATE or d > max_future_date(1):
            errors.append("Geçersiz tarih")
        else:
            parsed["entry_date"] = d

        cat_cell = cell(cells, "cost_category")
        cat_label = (str(cat_cell).strip().lower() if cat_cell else "")
        key = LABEL_TO_KEY.get(cat_label)
        if not key:
            shown = str(cat_cell).strip() if cat_cell else "(boş)"
            valid = ", ".join(list(COST_CATEGORIES.values())[:8]) + "…"
            errors.append(f"Kategori tanınmıyor: {shown} — Geçerli kategoriler: {valid}")
        else:
            parsed["cost_category"] = key

        sub = cell(cells, "subcategory")
        parsed["subcategory"] = str(sub).strip() if sub else None
        sup = cell(cells, "supplier_name")
        parsed["supplier_name"] = str(sup).strip() if sup else None
        desc = cell(cells, "description")
        parsed["description"] = str(desc).strip() if desc else None
        inv = cell(cells, "invoice_number")
        parsed["invoice_number"] = str(inv).strip() if inv else None

        try:
            parsed["amount_try"] = _parse_amount(cell(cells, "amount_try"))
        except ValueError as e:
            errors.append(str(e))

        vat = cell(cells, "vat_rate")
        try:
            parsed["vat_rate"] = Decimal(str(vat)) if vat not in (None, "") else Decimal("20")
        except (InvalidOperation, ValueError):
            errors.append("Geçersiz KDV oranı")

        try:
            parsed["payment_due_date"] = _parse_date(cell(cells, "payment_due_date"))
        except ValueError:
            errors.append(ERR_DATE_FORMAT)
        # CR-006-F: vade tarihi fatura (giriş) tarihinden önce olamaz.
        if parsed.get("entry_date") and parsed.get("payment_due_date") \
                and parsed["payment_due_date"] < parsed["entry_date"]:
            errors.append(ERR_DATE_ORDER)

        status_cell = cell(cells, "payment_status")
        status_raw = (str(status_cell).strip().lower() if status_cell else "ödenmedi")
        parsed["payment_status"] = "paid" if status_raw in ("ödendi", "odendi", "paid") else "unpaid"

        try:
            parsed["date_paid"] = _parse_date(cell(cells, "date_paid"))
        except ValueError:
            errors.append(ERR_DATE_FORMAT)
        if parsed["payment_status"] == "paid" and not parsed.get("date_paid"):
            errors.append("Ödendi durumunda ödeme tarihi zorunludur")

        rows.append({"row": idx + 1, "valid": not errors, "skipped": False, "errors": errors, "data": parsed})

    # CR-015-fix: an honest reason when there is nothing to import, so the preview
    # is never a silent empty screen.
    importable = any(not r["skipped"] for r in rows)
    message = None if importable else ERR_NO_IMPORTABLE

    return {"header_detected": header_detected, "header_row": header_row + 1,
            "column_map": colmap, "message": message, "rows": rows}
