"""CR-046 — decision-grade Excel report engine.

Turns already-computed ``run_spec`` results into a board-ready workbook:
  * Sheet 1 **"Özet"** — a Heneka header band, KPI cards (most-important top-left),
    and native Excel charts (``openpyxl.chart`` Bar/Line/Pie) that reference the
    data-sheet ranges, so they stay live if the user edits.
  * **Data sheets** (consolidated per table/section, NOT one-per-KPI) — navy headers,
    banded rows, autofilter, frozen header, ₺/% number formats, a "Toplam" totals row
    (the authoritative ``run_spec`` total, written as a value), and variance columns
    with RAG conditional formatting.

STRICTLY READ-ONLY: it consumes a result dict (no DB session, no queries, no writes).
**No fabrication** — every figure comes from ``run_spec``; the engine only adds
layout/charts/formatting, never numbers. The single ``_safe_cell`` formula-injection
guard (from ``export.py``) is applied to every user-authored string.

Both xlsx entry points route here: ``export._dashboard_xlsx`` (skills + dashboard
deck) and the xlsx branch of ``export.studio_export`` (single report). The PDF
(CR-037) and CSV (422) paths are untouched.
"""
import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.responses import APIError
from app.services.report_theme import (
    BG, CARD, GOLD, GREEN, HAIR, INK, MUT, NAVY, PETROL, RED, TINT,
)
from app.services.studio.catalog import GRAIN_CASH, GRAIN_COST_LINE, METRICS
from app.services.studio.export import _is_renderable, _safe_cell, _sheet_name

# --------------------------------------------------------------------------- #
# Palette → openpyxl ARGB + reusable styles
# --------------------------------------------------------------------------- #
def _argb(hex_str: str) -> str:
    """'#183047' → 'FF183047' (openpyxl wants ARGB, alpha first)."""
    return "FF" + hex_str.lstrip("#").upper()


_FONT = "Calibri"
_navy_fill = PatternFill("solid", fgColor=_argb(NAVY))
_gold_fill = PatternFill("solid", fgColor=_argb(GOLD))
_card_fill = PatternFill("solid", fgColor=_argb(BG))
_band_fill = PatternFill("solid", fgColor="FFF3F4F0")  # subtle zebra (BG, lighter)
_white_bold = Font(name=_FONT, bold=True, color="FFFFFFFF")
_white = Font(name=_FONT, color="FFFFFFFF")
_white_title = Font(name=_FONT, bold=True, size=16, color="FFFFFFFF")
_muted = Font(name=_FONT, size=9, color=_argb(MUT))
_value_font = Font(name=_FONT, bold=True, size=18, color=_argb(NAVY))
_ink = Font(name=_FONT, color=_argb(INK))
_bold = Font(name=_FONT, bold=True, color=_argb(INK))
_hair = Side(style="thin", color=_argb(HAIR))
_thin_border = Border(left=_hair, right=_hair, top=_hair, bottom=_hair)
_green_tint = PatternFill("solid", fgColor=_argb(TINT["g"]))
_red_tint = PatternFill("solid", fgColor=_argb(TINT["r"]))
_green_font = Font(name=_FONT, bold=True, color=_argb(GREEN))
_red_font = Font(name=_FONT, bold=True, color=_argb(RED))
_grey_font = Font(name=_FONT, color=_argb(MUT))
_usd_font = Font(name=_FONT, size=11, color=_argb(MUT))       # CR-055 — KPI USD secondary line
_footnote_font = Font(name=_FONT, italic=True, size=8, color=_argb(MUT))

# Number formats — ₺ with negatives in parens (red) and zeros as an en-dash.
_FMT_TRY = '#,##0" ₺";[Red](#,##0)" ₺";"–"'
# CR-055 — USD-at-date companion (mirrors _FMT_TRY). USD is NEVER fabricated: a cell
# with no amount_usd renders the en-dash, never a converted-at-today number.
_FMT_USD = '#,##0" $";[Red](#,##0)" $";"–"'
# CR-054 — percent has TWO formats, because the two kinds of percent carry different
# scales and Excel's "%" token multiplies by 100:
#   * VALUE cells hold percent-UNITS (28.42, 39.7 — the engine already ×100 via HUNDRED),
#     so a LITERAL "%" that does NOT re-scale, else 0.0% would render 28.42 as %2842.
#   * DELTA cells hold FRACTIONS (0.15), so 0.0% (Excel ×100) → %15,0 (unchanged).
_FMT_PCT_VALUE = '0.0"%"'      # percent-unit metric values (KPI cards + table cells)
_FMT_PCT_FRACTION = "0.0%"     # pct deltas (fractions) — Excel multiplies by 100
_FMT_NUM = '#,##0;[Red](#,##0);"–"'
# CR-047 — a model-inapplicable metric is None; show the en-dash (a numeric format's
# zero-section only fires for an actual 0, so a None/blank cell would otherwise be empty).
_DASH = "–"

_TR_MONTHS = ["", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
              "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık"]


def _num_fmt(col_type: str) -> str:
    if col_type == "currency":
        return _FMT_TRY
    if col_type == "percent":
        return _FMT_PCT_VALUE   # CR-054 — percent-unit VALUE (literal %, no ×100)
    return _FMT_NUM


# CR-055 — the Özet footnote making the USD basis honest (rate-at-date, CR-014).
_USD_FOOTNOTE = "USD değerleri her işlemin kendi tarihindeki kur ile hesaplanmıştır (CR-014)."


def _usd_label(label: str) -> str:
    """Derive a paired USD column header from a ₺ one: 'Maliyet (₺)' → 'Maliyet ($)',
    otherwise append ' ($)' (e.g. 'Gelir' → 'Gelir ($)')."""
    if "(₺)" in (label or ""):
        return label.replace("(₺)", "($)")
    if "₺" in (label or ""):
        return label.replace("₺", "$")
    return f"{label or ''} ($)".strip()


# --------------------------------------------------------------------------- #
# Metric metadata (read-only, from the catalog) — snapshot + favourable direction
# --------------------------------------------------------------------------- #
def _is_snapshot(metric_id: str) -> bool:
    """A non-windowed / whole-project snapshot metric (e.g. hakediş, forecast):
    its grain is not a time-series grain, so grouping it by month yields blank rows.
    Mirrors ``export._has_unwindowed_metric``."""
    m = METRICS.get(metric_id)
    return bool(m) and m["grain"] not in (GRAIN_COST_LINE, GRAIN_CASH)


# Non-cost metrics where an INCREASE is unfavourable (more outflow / more unpaid /
# slower collection). Cost-group metrics are already caught by the group check below.
_UNFAVOURABLE_UP = {"cash_out", "receivables", "dso"}


def _favourable_up(metric_id: str) -> bool:
    """RAG direction: for cost metrics, cash outflow, open receivables and DSO an
    increase is UNFAVOURABLE (red); for revenue/profit/margin/inflow an increase is
    favourable (green)."""
    if metric_id in _UNFAVOURABLE_UP:
        return False
    m = METRICS.get(metric_id)
    return not (m and m.get("group") == "Maliyet")


def _is_time_grouped(columns: list) -> bool:
    return any(c.get("kind") == "dimension" and c.get("type") == "date" for c in columns)


# --------------------------------------------------------------------------- #
# Period label (Turkish) from the result meta — no query
# --------------------------------------------------------------------------- #
def _period_label(result: dict) -> str:
    dr = ((result or {}).get("meta") or {}).get("date_range") or {}
    f, t = dr.get("from"), dr.get("to")
    if not f and not t:
        # CR-049 — an all-time window: label from the REAL span the result covers
        # (its date-bucket rows), e.g. "2018 – 2020"; "Tüm zamanlar" when the report
        # carries no time dimension. Never a hardcoded recent range.
        return _all_time_label(result)
    try:
        fd = datetime.fromisoformat(f) if f else None
        td = datetime.fromisoformat(t) if t else None
    except (TypeError, ValueError):
        return f"{f or '—'} – {t or '—'}"
    if fd and td and fd.year == td.year and fd.month == td.month:
        return f"{_TR_MONTHS[fd.month]} {fd.year}"
    fs = f"{_TR_MONTHS[fd.month]} {fd.year}" if fd else "—"
    ts = f"{_TR_MONTHS[td.month]} {td.year}" if td else "—"
    return f"{fs} – {ts}"


def _all_time_label(result: dict) -> str:
    """CR-049 — derive the span an all-time result actually covers from any date-typed
    dimension buckets in its rows (month '2018-01' / quarter '2018-Q1' / year '2018' —
    the leading 4 digits are the year). Returns '2018 – 2020' (or a single '2018'), or
    'Tüm zamanlar' when the report has no time dimension to read a span from."""
    date_ids = [c.get("id") for c in (result.get("columns") or [])
                if c.get("kind") == "dimension" and c.get("type") == "date"]
    years: set[int] = set()
    for row in (result.get("rows") or []) if date_ids else []:
        dims = row.get("dims") or {}
        for did in date_ids:
            v = dims.get(did)
            if isinstance(v, str) and len(v) >= 4 and v[:4].isdigit():
                years.add(int(v[:4]))
    if not years:
        return "Tüm zamanlar"
    lo, hi = min(years), max(years)
    return str(lo) if lo == hi else f"{lo} – {hi}"


# --------------------------------------------------------------------------- #
# Header band
# --------------------------------------------------------------------------- #
def header_band(ws, title: str, company: str | None, period: str, start_row: int = 1) -> int:
    """Navy band: company / title / period / 'Yapı AI ile üretildi · {tarih}' + a gold
    rule. Returns the next free row. Every string is _safe_cell-guarded."""
    span = 8
    last_col = get_column_letter(span)
    rows = [
        (_safe_cell(company or "YAPI"), _white_bold, 13),
        (_safe_cell(title or "Rapor"), _white_title, 18),
        (_safe_cell(period or ""), _white, 11),
        (_safe_cell(f"Yapı AI ile üretildi · {datetime.now(timezone.utc):%d.%m.%Y}"), _white, 9),
    ]
    r = start_row
    for text, font, height in rows:
        ws.merge_cells(f"A{r}:{last_col}{r}")
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for ci in range(1, span + 1):
            ws.cell(row=r, column=ci).fill = _navy_fill
        ws.row_dimensions[r].height = height
        r += 1
    # gold rule
    ws.merge_cells(f"A{r}:{last_col}{r}")
    for ci in range(1, span + 1):
        ws.cell(row=r, column=ci).fill = _gold_fill
    ws.row_dimensions[r].height = 4
    return r + 2


# --------------------------------------------------------------------------- #
# KPI cards
# --------------------------------------------------------------------------- #
def kpi_cards(ws, cards: list, start_row: int) -> int:
    """A 2–4-across grid of cards: muted label / large ₺ value / RAG delta. ``cards``
    are dicts {label, value, value_type, delta, favourable_up}. Returns next free row.

    CR-055 — when any card carries a ``usd_value`` (a currency card with a USD
    companion), the whole grid gains a 4th row: the ₺ headline gets a muted USD
    secondary line beneath it (``≈ 312.400 $``), so every ₺ figure shows its USD-at-date
    equivalent. A missing USD renders "–" (never a fabricated number)."""
    if not cards:
        return start_row
    dual = any("usd_value" in c for c in cards)   # a USD secondary line is present
    body_rows = 4 if dual else 3                    # label, value, [usd], delta
    per_row = 4
    col_pairs = [(1, 2), (3, 4), (5, 6), (7, 8)]  # each card spans 2 cols
    r = start_row
    for i in range(0, len(cards), per_row):
        band = cards[i:i + per_row]
        for j, card in enumerate(band):
            c0, c1 = col_pairs[j]
            cl0, cl1 = get_column_letter(c0), get_column_letter(c1)
            usd_row = (r + 2) if dual else None
            delta_row = (r + 3) if dual else (r + 2)
            # label
            ws.merge_cells(f"{cl0}{r}:{cl1}{r}")
            lc = ws.cell(row=r, column=c0, value=_safe_cell(card["label"]))
            lc.font = _muted
            lc.alignment = Alignment(horizontal="left", indent=1)
            # value — a model-inapplicable metric is None → render the CR's "–", not blank.
            ws.merge_cells(f"{cl0}{r + 1}:{cl1}{r + 1}")
            vc = ws.cell(row=r + 1, column=c0)
            val = card.get("value")
            if val is None:
                vc.value = _DASH
            else:
                vc.value = val
                vc.number_format = _num_fmt(card.get("value_type") or "currency")
            vc.font = _value_font
            vc.alignment = Alignment(horizontal="left", indent=1)
            # CR-055 — USD secondary line (currency cards only; blank for %/count cards).
            if dual:
                ws.merge_cells(f"{cl0}{usd_row}:{cl1}{usd_row}")
                uc = ws.cell(row=usd_row, column=c0)
                if "usd_value" in card:
                    uv = card.get("usd_value")
                    if uv is None:
                        uc.value = _DASH   # no amount_usd for this figure → honest en-dash
                    else:
                        uc.value = uv
                        uc.number_format = _FMT_USD
                    uc.font = _usd_font
                    uc.alignment = Alignment(horizontal="left", indent=1)
            # delta — a styled string (arrow + %), RAG-coloured by favourable direction.
            ws.merge_cells(f"{cl0}{delta_row}:{cl1}{delta_row}")
            dc = ws.cell(row=delta_row, column=c0)
            delta = card.get("delta")
            if delta is None:
                dc.value = "—"
                dc.font = _grey_font
            else:
                d = float(delta)
                arrow = "▲" if d > 0 else ("▼" if d < 0 else "■")
                if card.get("delta_unit") == "abs":
                    # An absolute ₺/number difference, not a fraction.
                    unit_suffix = " ₺" if card.get("value_type") == "currency" else ""
                    dc.value = f"{arrow} {abs(d):,.0f}{unit_suffix}"
                else:
                    dc.value = f"{arrow} %{abs(d) * 100:.1f}"
                fav = card.get("favourable_up", True)
                dc.font = _grey_font if d == 0 else (_green_font if (d > 0) == fav else _red_font)
            dc.alignment = Alignment(horizontal="left", indent=1)
            # card border on the body rows
            for rr in range(r, r + body_rows):
                for cc in (c0, c1):
                    ws.cell(row=rr, column=cc).border = _thin_border
                    if rr != r:
                        ws.cell(row=rr, column=cc).fill = _card_fill
        r += body_rows + 1  # body rows + 1 gap
    return r


# --------------------------------------------------------------------------- #
# Data sheet (consolidated table) — returns chart-range info
# --------------------------------------------------------------------------- #
def data_sheet(wb, result: dict, sheet_title: str, used_names: set,
               result_usd: dict | None = None) -> dict | None:
    """Write a styled data sheet from a result. Snapshot metrics in a time-grouped
    table are dropped here (rerouted to KPI cards by the caller). Returns a dict with
    the chart-range info ({sheet, cat_col, metric_cols, header_row, last_data_row}) or
    None when there is nothing tabular to render.

    CR-055 — when ``result_usd`` (the same spec run with basis.currency='usd') is given,
    each ₺ currency column gains a paired USD-at-date column ('Maliyet (₺)' → 'Maliyet
    ($)'), placed right after the ₺ metric block so the chart's ₺ range is unchanged. A
    figure with no ``amount_usd`` renders "–" (never fabricated)."""
    all_columns = result.get("columns") or []
    rows = result.get("rows") or []
    time_grouped = _is_time_grouped(all_columns)
    # Drop snapshot metric columns from a time-grouped table (the Hakediş bug).
    columns = [
        c for c in all_columns
        if not (c.get("kind") == "metric" and time_grouped and _is_snapshot(c["id"]))
    ]
    metric_cols = [c for c in columns if c.get("kind") == "metric"]
    dim_cols = [c for c in columns if c.get("kind") == "dimension"]
    if not columns or not rows or not metric_cols:
        return None

    has_delta = bool(((result.get("meta") or {}).get("comparison")))
    delta_unit = ((result.get("meta") or {}).get("comparison_unit")) or "pct"

    def _delta_fmt(col: dict) -> str:
        # pct deltas are FRACTIONS → 0.0% (Excel ×100); abs deltas are ₺/number diffs.
        return _FMT_PCT_FRACTION if delta_unit == "pct" else _num_fmt(col.get("type"))

    # CR-055 — USD companion: a $ column per ₺ currency metric. Match USD rows/totals by
    # dimension key (order-independent — same spec, same groups) so a re-sorted USD run
    # still lines up. usd_missing_count flags rows the FX snapshot couldn't value.
    # A $ pair is added only for ₺ currency columns; a column already in USD (its label
    # carries "$", e.g. cost_usd "Maliyet ($)") is not paired again with a redundant column.
    usd_metric_cols = ([c for c in metric_cols
                        if c.get("type") == "currency" and "$" not in (c.get("label") or "")]
                       if result_usd else [])
    usd_by_key: dict = {}
    usd_totals: dict = {}
    if usd_metric_cols:
        for ur in (result_usd.get("rows") or []):
            key = tuple((ur.get("dims") or {}).get(d["id"]) for d in dim_cols)
            usd_by_key[key] = ur.get("metrics") or {}
        usd_totals = (result_usd.get("totals") or {}).get("metrics") or {}

    name = _sheet_name(sheet_title, used_names)
    used_names.add(name)
    ws = wb.create_sheet(title=name)

    # Column plan: [dims...] [₺ metrics...] [$ USD metrics...] [Δ metrics... (if compare)]
    plan: list = []  # (col_index, col_dict, role)  role: dim|metric|usd|delta
    ci = 1
    for c in dim_cols:
        plan.append((ci, c, "dim")); ci += 1
    metric_first_col = ci
    for c in metric_cols:
        plan.append((ci, c, "metric")); ci += 1
    metric_last_col = ci - 1
    # $ block sits AFTER the ₺ block so metric_first/last_col (the chart's ₺ range) never
    # shift — charts stay ₺-only and CR-052/054 chart behaviour is untouched.
    for c in usd_metric_cols:
        plan.append((ci, c, "usd")); ci += 1
    delta_cols: list = []
    if has_delta:
        for c in metric_cols:
            plan.append((ci, c, "delta")); delta_cols.append((ci, c)); ci += 1
    ncols = ci - 1

    # Header row
    for col_idx, c, role in plan:
        if role == "delta":
            label = f"Δ {c['label']}"
        elif role == "usd":
            label = _usd_label(c["label"])
        else:
            label = c["label"]
        cell = ws.cell(row=1, column=col_idx, value=_safe_cell(label))
        cell.font = _white_bold
        cell.fill = _navy_fill
        cell.alignment = Alignment(horizontal=("left" if role == "dim" else "right"), vertical="center")

    # Data rows
    for ri, row in enumerate(rows, start=2):
        dims = row.get("dims") or {}
        metrics = row.get("metrics") or {}
        deltas = row.get("deltas") or {}
        rkey = tuple(dims.get(d["id"]) for d in dim_cols)   # CR-055 — USD row match key
        banded = (ri % 2 == 0)
        for col_idx, c, role in plan:
            if role == "dim":
                v = dims.get(c["id"])
                cell = ws.cell(row=ri, column=col_idx, value=_safe_cell(v))
            elif role == "metric":
                mv = metrics.get(c["id"])
                if mv is None:  # CR-047 — model-inapplicable metric → "–", not blank
                    cell = ws.cell(row=ri, column=col_idx, value=_DASH)
                else:
                    cell = ws.cell(row=ri, column=col_idx, value=mv)
                    cell.number_format = _num_fmt(c["type"])
            elif role == "usd":  # CR-055 — USD-at-date pair; "–" when no amount_usd
                uv = usd_by_key.get(rkey, {}).get(c["id"])
                if uv is None:
                    cell = ws.cell(row=ri, column=col_idx, value=_DASH)
                else:
                    cell = ws.cell(row=ri, column=col_idx, value=uv)
                    cell.number_format = _FMT_USD
            else:  # delta (fraction for pct; absolute ₺/number for abs)
                cell = ws.cell(row=ri, column=col_idx, value=deltas.get(c["id"]))
                cell.number_format = _delta_fmt(c)
            cell.font = _ink
            if banded:
                cell.fill = _band_fill
    last_data_row = 1 + len(rows)

    # "Toplam" totals row — authoritative run_spec values (NOT a re-sum formula),
    # so the sheet total can never diverge from the trusted engine total.
    totals = (result.get("totals") or {}).get("metrics") or {}
    total_deltas = (result.get("totals") or {}).get("deltas") or {}
    trow = last_data_row + 1
    label_placed = False
    for col_idx, c, role in plan:
        cell = ws.cell(row=trow, column=col_idx)
        if role == "dim":
            cell.value = _safe_cell("Toplam") if not label_placed else None
            label_placed = True
        elif role == "metric":
            tv = totals.get(c["id"])
            if tv is None:  # CR-047 — "–" for a model-inapplicable total
                cell.value = _DASH
            else:
                cell.value = tv
                cell.number_format = _num_fmt(c["type"])
        elif role == "usd":  # CR-055 — USD-at-date total (authoritative usd-run total)
            uv = usd_totals.get(c["id"])
            if uv is None:
                cell.value = _DASH
            else:
                cell.value = uv
                cell.number_format = _FMT_USD
        else:
            cell.value = (total_deltas or {}).get(c["id"]) if total_deltas else None
            cell.number_format = _delta_fmt(c)
        cell.font = _bold
        cell.fill = _gold_fill

    # RAG conditional formatting on each Δ column (favourable green / unfavourable red)
    for col_idx, c in delta_cols:
        col = get_column_letter(col_idx)
        rng = f"{col}2:{col}{last_data_row}"
        up_fav = _favourable_up(c["id"])
        good = PatternFill("solid", fgColor=_argb(TINT["g"]))
        bad = PatternFill("solid", fgColor=_argb(TINT["r"]))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="greaterThan", formula=["0"], fill=(good if up_fav else bad)))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="lessThan", formula=["0"], fill=(bad if up_fav else good)))

    # Autofilter, freeze, widths
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{last_data_row}"
    ws.freeze_panes = "A2"
    for col_idx, c, role in plan:
        ws.column_dimensions[get_column_letter(col_idx)].width = 30 if role == "dim" else 16

    # Chart range info: categories = first dim column, values = the metric block.
    cat_col = plan[0][0] if dim_cols else None
    return {
        "sheet": ws,
        "cat_col": cat_col,
        "metric_first_col": metric_first_col,
        "metric_last_col": metric_last_col,
        "header_row": 1,
        "last_data_row": last_data_row,
        "metric_cols": metric_cols,
        "dim_cols": dim_cols,
        "time_grouped": time_grouped,
    }


# --------------------------------------------------------------------------- #
# Native charts referencing the data-sheet ranges
# --------------------------------------------------------------------------- #
# CR-046 Heneka categorical palette (mirrors report_charts._PALETTE / _DONUT_PALETTE)
# — series/slices cycle through these so native charts match the PDF deck.
_SERIES_COLORS = [NAVY, PETROL, GOLD, GREEN, "#A9B2AC"]
_AXIS_FMT = '#,##0" ₺"'   # ₺ value-axis ticks (CR-052 §3)
_CAT_CAP = 12             # cap a ranking/composition chart's categories for readability


def _color6(hex_str: str) -> str:
    """'#183047' / '183047' → '183047' (openpyxl solidFill wants 6-hex, no alpha)."""
    return hex_str.lstrip("#").upper()


def _apply_series_colors(chart, *, line: bool) -> None:
    """Recolour every data series with the Heneka palette. Defensive: a styling
    failure must NEVER break the workbook (the file-opens-clean invariant)."""
    try:
        for i, s in enumerate(chart.series):
            color = _color6(_SERIES_COLORS[i % len(_SERIES_COLORS)])
            if line:
                gp = GraphicalProperties()
                gp.line = LineProperties(solidFill=color, w=28575)  # ~2.25pt
                s.graphicalProperties = gp
            else:
                s.graphicalProperties = GraphicalProperties(solidFill=color)
    except Exception:
        pass


def _apply_pie_colors(chart, n_points: int) -> None:
    """Colour each pie slice from the Heneka palette (per-point, not per-series)."""
    try:
        if not chart.series:
            return
        ser = chart.series[0]
        for i in range(max(0, n_points)):
            color = _color6(_SERIES_COLORS[i % len(_SERIES_COLORS)])
            # openpyxl's DataPoint constructor takes ``spPr`` (graphicalProperties is a
            # read-only alias); passing graphicalProperties= raises TypeError → no colors.
            ser.data_points.append(
                DataPoint(idx=i, spPr=GraphicalProperties(solidFill=color)))
    except Exception:
        pass


# CR-054 — signed single-series colouring (net nakit / a variance) --------------- #
def _is_signed_series(values) -> bool:
    """True when a single-series chart is a SIGNED value worth per-sign colouring —
    i.e. its plotted points actually dip below zero (net nakit with a deficit month, a
    ± variance). An all-positive ranking / composition keeps the categorical palette."""
    for v in values:
        try:
            if float(v) < 0:
                return True
        except (TypeError, ValueError):
            continue   # a "–"/None cell is not a number — skip it
    return False


def _apply_signed_colors(chart, values) -> None:
    """CR-054 — colour a single-series signed chart per data point: green > 0, red < 0,
    neutral grey at exactly 0 — so the sign of net nakit / a variance reads at a glance.
    Uses per-point ``DataPoint`` on series[0] (the same ``spPr`` mechanism as the pie).
    Defensive: a styling failure must NEVER break the workbook (file-opens-clean)."""
    try:
        if not chart.series:
            return
        ser = chart.series[0]
        for i, v in enumerate(values):
            try:
                fv = float(v)
            except (TypeError, ValueError):
                continue
            color = GREEN if fv > 0 else (RED if fv < 0 else MUT)
            ser.data_points.append(
                DataPoint(idx=i, spPr=GraphicalProperties(solidFill=_color6(color))))
    except Exception:
        pass


# CR-054 — the value axis follows the CHARTED metric's type (not a hard-wired ₺) so a
# ₺ chart shows ₺ ticks and a genuine % chart shows % ticks.
def _axis_value_fmt(col_type: str) -> str:
    if col_type == "percent":
        return _FMT_PCT_VALUE   # literal % (percent-unit ticks, no ×100)
    if col_type == "number":
        return "#,##0"
    return _AXIS_FMT            # currency (default) — ₺ ticks


def _axis_value_title(col_type: str) -> str | None:
    """Unit label for the value axis: ₺ for currency, % for percent, none otherwise."""
    if col_type == "currency":
        return "₺"
    if col_type == "percent":
        return "%"
    return None


def _style_value_axis(axis, num_fmt: str, title) -> None:
    """CR-054 — the openpyxl trap: an axis with ``delete=None`` is HIDDEN by Excel, so
    ticks/labels vanish even though numFmt is set. Un-hide it, give it the type-correct
    number format and a unit title. Value axis is ALWAYS ``y_axis`` (NumericAxis) — even
    for a horizontal bar, where only the visual orientation swaps, not the axis object."""
    axis.delete = False
    axis.numFmt = num_fmt
    if title:
        axis.title = title


def _style_cat_axis(axis, title) -> None:
    """CR-054 — un-hide the category axis (``x_axis``, TextAxis) and title it from the
    charted dimension (e.g. 'Ay', 'Maliyet kategorisi')."""
    axis.delete = False
    if title:
        axis.title = title


def _style_chart(chart, title: str, *, legend: bool = True):
    # Size is set by add_chart (CR-055 content-aware); this only sets title + legend.
    chart.title = _safe_cell(title or "Grafik")
    if legend:
        try:
            chart.legend.position = "b"
        except AttributeError:
            pass
    else:
        chart.legend = None  # a single-series chart needs no legend


# CR-055 — content-aware chart sizing (no more cramped 3-bar or 30-month charts) ---- #
def _chart_size(n_categories: int, max_label_len: int, *, multi: bool, legend: bool) -> tuple:
    """Chart (width, height) in cm. Width grows with the category count and the longest
    category label so a 30-month trend / 10-category bar has room (clamped 16–40); a
    3-bar chart stays compact. Height bumps for a bottom legend / many series (7–12)."""
    width = 16.0 + 0.55 * max(0, n_categories - 4)
    if max_label_len > 12:
        width += min(6.0, 0.3 * (max_label_len - 12))
    width = max(16.0, min(width, 40.0))
    height = 8.0 + (1.5 if legend else 0.0) + (1.0 if multi else 0.0)
    height = max(7.0, min(height, 12.0))
    return round(width, 1), round(height, 1)


def _chart_rows(height_cm: float) -> int:
    """Rows to advance the Özet anchor for a chart of this height (+ a gap) so stacked
    charts never overlap. Excel's default row height ≈ 0.53 cm."""
    return int(round((height_cm or 8.0) / 0.53)) + 2


def _filter_key(spec: dict) -> tuple:
    """A stable, hashable key for a widget's filters, so the chart de-dup treats two
    widgets with the same metric+dim but DIFFERENT filters as distinct (not twins)."""
    out = []
    for f in (spec or {}).get("filters") or []:
        if isinstance(f, dict):
            out.append((str(f.get("field")), str(f.get("op")), str(f.get("value"))))
    return tuple(sorted(out))


def pick_chart(columns: list, rows: list, viz: str | None) -> str | None:
    """CR-052 — pick the chart that the data SHAPE earns, or ``None`` (a chart that
    merely restates the table beside it is clutter). Returns one of
    ``line|area|bar|clustered_bar|pie|None``:

      * **date dim** + 1 metric → ``line`` (``area`` when explicitly cumulative),
        + ≥2 metrics → ``clustered_bar`` (e.g. gelir vs gider);
      * **one category dim** + 1 metric, >3 rows → ``bar`` (horizontal ranking),
        ≤6 rows of a summable ₺ metric → ``pie`` (part-to-whole composition),
        + ≥2 metrics → ``clustered_bar`` (bütçe/gerçekleşen/taahhüt side-by-side);
      * **snapshot / single value / ≤3 rows / lookup / 0 or >1 dims** → ``None``.

    An explicit ``viz`` is honoured (``area``/``bar`` on a time series) but the shape
    overrides it to ``None`` when no chart helps, and upgrades a generic ``bar`` to
    ``clustered_bar``/``pie`` when the shape fits. ``columns`` are the RENDERED columns
    (snapshot metrics already dropped by ``data_sheet``)."""
    dim_cols = [c for c in columns if c.get("kind") == "dimension"]
    metric_cols = [c for c in columns if c.get("kind") == "metric"]
    n_rows = len(rows or [])
    n_metrics = len(metric_cols)

    # Need exactly one dimension to plot against and at least one metric. Zero dims is a
    # snapshot KPI; multiple dims is a cross-tab a single-axis chart only muddles.
    if len(dim_cols) != 1 or n_metrics == 0:
        return None
    # Too few points to be worth a chart — a KPI card / the table says it better.
    if n_rows < 2:
        return None

    if dim_cols[0].get("type") == "date":
        # Trend over time.
        if n_metrics >= 2:
            return "clustered_bar"
        if viz == "area":
            return "area"
        if viz == "bar":
            return "clustered_bar"   # honour an explicit column-over-time intent
        return "line"

    # One category dimension (kategori / vendor / tip / proje).
    if n_metrics >= 2:
        return "clustered_bar"       # side-by-side metric comparison
    if n_rows <= 3:
        return None                  # ≤3 categories — a KPI/table reads better
    metric = metric_cols[0]
    if metric.get("type") == "currency" and n_rows <= 6:
        return "pie"                 # part-to-whole composition, few slices
    return "bar"                     # ranking / comparison (horizontal bar)


def add_chart(ozet, ref: dict, kind: str, title: str, anchor: str) -> float:
    """Build a native chart from a data-sheet range info and anchor it on the Özet.
    ``kind`` is a ``pick_chart`` result. Categories on a non-time ranking/composition
    chart are capped (``_CAT_CAP``) for readability; a time series keeps every point so
    the line reads left→right over the whole span (CR-050 chronological order).

    Returns the chart's height in cm (CR-055) so the caller can advance the Özet anchor
    without overlap; returns 0.0 when nothing was drawn."""
    sheet = ref["sheet"]
    cat_col = ref["cat_col"]
    hr = ref["header_row"]
    last = ref["last_data_row"]
    mfirst, mlast = ref["metric_first_col"], ref["metric_last_col"]
    if cat_col is None or last <= hr:
        return 0.0
    is_time = bool(ref.get("time_grouped"))
    data_last = last if is_time else min(last, hr + _CAT_CAP)
    cats = Reference(sheet, min_col=cat_col, max_col=cat_col, min_row=hr + 1, max_row=data_last)
    multi = mlast > mfirst

    # CR-054 — value axis follows the CHARTED metric's type; category axis titled from
    # the dimension. (metric_cols/dim_cols carry the RENDERED columns, snapshot-dropped.)
    metric_cols = ref.get("metric_cols") or []
    dim_cols = ref.get("dim_cols") or []
    value_type = metric_cols[0].get("type") if metric_cols else "currency"
    value_fmt = _axis_value_fmt(value_type)
    value_title = _axis_value_title(value_type)
    cat_title = dim_cols[0].get("label") if dim_cols else None

    # CR-055 — content-aware size from the plotted category count + longest label.
    n_categories = max(0, data_last - hr)
    try:
        max_label_len = max((len(str(sheet.cell(row=rr, column=cat_col).value or ""))
                             for rr in range(hr + 1, data_last + 1)), default=0)
    except Exception:
        max_label_len = 0
    legend_shown = True if kind == "pie" else (False if kind == "bar" else multi)
    w_cm, h_cm = _chart_size(n_categories, max_label_len, multi=multi, legend=legend_shown)

    def _color_single(chart, *, line: bool):
        # A single-series chart of a signed value → per-point green/red; otherwise the
        # Heneka palette (an all-positive ranking/composition stays on-brand).
        vals = [sheet.cell(row=r, column=mfirst).value for r in range(hr + 1, data_last + 1)]
        if _is_signed_series(vals):
            _apply_signed_colors(chart, vals)
        else:
            _apply_series_colors(chart, line=line)

    if kind == "pie":
        chart = PieChart()
        data = Reference(sheet, min_col=mfirst, max_col=mfirst, min_row=hr, max_row=data_last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList(); chart.dataLabels.showPercent = True
        _apply_pie_colors(chart, data_last - hr)
        _style_chart(chart, title, legend=True)
        # Pie has no x/y axes (it uses showPercent labels) — nothing to un-hide.
        chart.width = w_cm; chart.height = h_cm
        ozet.add_chart(chart, anchor)
        return h_cm

    if kind in ("line", "area"):
        chart = LineChart() if kind == "line" else AreaChart()
        data = Reference(sheet, min_col=mfirst, max_col=mlast, min_row=hr, max_row=data_last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        if multi:
            _apply_series_colors(chart, line=(kind == "line"))
        else:
            _color_single(chart, line=(kind == "line"))
        _style_chart(chart, title, legend=multi)
    elif kind == "bar":
        chart = BarChart(); chart.type = "bar"      # horizontal ranking
        data = Reference(sheet, min_col=mfirst, max_col=mfirst, min_row=hr, max_row=data_last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        _color_single(chart, line=False)
        _style_chart(chart, title, legend=False)
    else:  # clustered_bar — multi-series columns over the dim
        chart = BarChart(); chart.type = "col"; chart.grouping = "clustered"
        data = Reference(sheet, min_col=mfirst, max_col=mlast, min_row=hr, max_row=data_last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        _apply_series_colors(chart, line=False)
        _style_chart(chart, title, legend=multi)

    # CR-054 — un-hide both axes (openpyxl leaves delete=None → Excel hides them, the
    # blank-axis bug) with the type-correct value format + titles. Value axis is always
    # y_axis (NumericAxis); category axis is x_axis (TextAxis) — true for bar & col alike.
    _style_value_axis(chart.y_axis, value_fmt, value_title)
    _style_cat_axis(chart.x_axis, cat_title)
    chart.width = w_cm; chart.height = h_cm
    ozet.add_chart(chart, anchor)
    return h_cm


# --------------------------------------------------------------------------- #
# Card builders from a result
# --------------------------------------------------------------------------- #
def _cards_from_result(result: dict, *, only_snapshot: bool = False, suffix: str = "",
                       result_usd: dict | None = None) -> list:
    """Build KPI cards from a result's totals. ``only_snapshot`` keeps just snapshot
    metrics (for the time-grouped reroute); otherwise all metric columns.

    CR-055 — when ``result_usd`` is given, each CURRENCY card carries a ``usd_value``
    (the USD-at-date total from the companion run) so the card can show ₺ + USD; a
    missing USD is None → rendered "–" (never fabricated)."""
    totals = (result.get("totals") or {}).get("metrics") or {}
    deltas = (result.get("totals") or {}).get("deltas") or {}
    unit = ((result.get("meta") or {}).get("comparison_unit")) or "pct"
    usd_totals = (result_usd.get("totals") or {}).get("metrics") or {} if result_usd else {}
    cards: list = []
    for c in result.get("columns") or []:
        if c.get("kind") != "metric":
            continue
        mid = c["id"]
        if only_snapshot and not _is_snapshot(mid):
            continue
        card = {
            "label": (c.get("label") or mid) + suffix,
            "value": totals.get(mid),
            "value_type": c.get("type"),
            "delta": (deltas or {}).get(mid) if deltas else None,
            "delta_unit": unit,
            "favourable_up": _favourable_up(mid),
        }
        if result_usd is not None and c.get("type") == "currency":
            card["usd_value"] = usd_totals.get(mid)   # None → "–" (no amount_usd)
        cards.append(card)
    return cards


# --------------------------------------------------------------------------- #
# The orchestrator
# --------------------------------------------------------------------------- #
def _usd_footnote(ws, row: int, results_usd: dict) -> int:
    """CR-055 — write the USD honesty footnote (rate-at-date, CR-014) on the Özet at
    ``row``, plus a 'no FX for N rows' note when any companion run reported
    ``usd_missing_count`` > 0. Returns the next free row."""
    span = 8
    last_col = get_column_letter(span)
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1, value=_safe_cell(_USD_FOOTNOTE))
    c.font = _footnote_font
    c.alignment = Alignment(horizontal="left", indent=1)
    row += 1
    missing = sum((r.get("meta") or {}).get("usd_missing_count") or 0
                  for r in (results_usd or {}).values() if isinstance(r, dict))
    if missing:
        ws.merge_cells(f"A{row}:{last_col}{row}")
        c2 = ws.cell(row=row, column=1,
                     value=_safe_cell(f"{missing} kayıt için USD kuru yok — bu değerler “–” gösterilir."))
        c2.font = _footnote_font
        c2.alignment = Alignment(horizontal="left", indent=1)
        row += 1
    return row


def build_workbook(widgets: list, results: dict, title: str, *,
                   company: str | None = None, period: str | None = None,
                   single_report: bool = False, results_usd: dict | None = None) -> bytes:
    """Render a decision-grade workbook → bytes. ``widgets`` is the ordered widget
    list (section-grouped by the caller); ``results`` maps widget id → run_spec result
    or {"unavailable": True}. Raises ``APIError(422, NO_DATA)`` when nothing is
    renderable (no empty workbook).

    CR-055 — ``results_usd`` (optional): the SAME widgets run with basis.currency='usd'
    (a USD-at-date companion, ``run_both_currencies``). When given, every currency KPI
    shows a ₺ headline + USD secondary line, currency table columns get a paired $
    column, and an Özet footnote states the rate-at-date basis. When None, ₺-only —
    identical to the pre-CR-055 output (premade cost/cashflow, which have no runnable
    USD spec, pass None and are unchanged)."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet; Özet is inserted at the front later

    used_names: set = set()
    cards: list = []
    chart_candidates: list = []  # CR-052 — {sig, ref, kind, title, explicit}, deduped below
    rendered = False

    # Period from the result windows (no query). CR-049 — prefer a CONCRETE span (a
    # label carrying a year, e.g. "2018 – 2020" / "Ocak 2026 – …") over a bare "Tüm
    # zamanlar", so a multi-widget all-time report whose first widget is a snapshot
    # KPI still headlines the real span a later time-series widget covers.
    if period is None:
        fallback = None
        for w in widgets:
            res = results.get(w.get("id"))
            if isinstance(res, dict) and res.get("meta"):
                label = _period_label(res)
                if fallback is None:
                    fallback = label
                if any(ch.isdigit() for ch in label):
                    period = label
                    break
        period = period or fallback

    for w in widgets:
        wtype = w.get("type")
        if wtype == "text":
            continue
        res = results.get(w.get("id"))
        res_usd = (results_usd or {}).get(w.get("id"))   # CR-055 — USD companion (or None)

        if wtype == "kpi":
            if _is_renderable(res):
                cards.extend(_cards_from_result(res, result_usd=res_usd))
                rendered = True
            continue

        if not _is_renderable(res):
            continue

        columns = res.get("columns") or []
        if single_report:
            # Özet-lite: headline KPI cards from the report's totals (all metrics).
            cards = _cards_from_result(res, result_usd=res_usd) + cards
            rendered = True
        elif _is_time_grouped(columns):
            # Snapshot reroute: a snapshot metric in a time-grouped table → KPI card.
            snap_cards = _cards_from_result(res, only_snapshot=True, suffix=" (tüm proje)",
                                            result_usd=res_usd)
            if snap_cards:
                # A card-only Özet built from authoritative run_spec totals is a valid
                # decision dashboard, so this counts as rendered even if the table that
                # follows drops to nothing (all-snapshot table → no data sheet).
                cards.extend(snap_cards)
                rendered = True

        ref = data_sheet(wb, res, w.get("title") or "Veri", used_names, result_usd=res_usd)
        if ref is None:
            # All metrics were snapshot → already rerouted to cards above; no tabular
            # sheet to write for this widget.
            continue
        rendered = True

        # CR-052 — let EVERY rendered table earn a chart by its data shape (not one
        # per table). pick_chart returns None for shapes a chart can't improve, so the
        # Özet stays a curated dashboard. An explicit chart widget's viz is honoured.
        spec = w.get("spec") or {}
        viz = spec.get("viz")
        sheet_columns = (ref["dim_cols"] or []) + (ref["metric_cols"] or [])
        kind = pick_chart(sheet_columns, res.get("rows") or [], viz)
        if kind:
            # Signature includes FILTERS so two same-metric+dim widgets that differ only
            # by filter (e.g. cost-by-month for project A vs B) are NOT collapsed — only
            # a true twin (a table restating a chart's exact data) de-dups.
            sig = (
                tuple(sorted(c["id"] for c in ref["metric_cols"])),
                tuple(c["id"] for c in ref["dim_cols"]),
                _filter_key(spec),
            )
            chart_candidates.append({
                "sig": sig, "ref": ref, "kind": kind,
                "title": w.get("title") or title, "explicit": wtype == "chart",
            })

    if not rendered:
        raise APIError(422, "NO_DATA", "Dışa aktarılacak veri yok")

    # CR-052 de-dup: one chart per (metrics, dims) signature — don't chart a table that
    # a twin chart widget already visualizes. An explicit chart widget wins the twin.
    chosen: dict = {}
    for cand in chart_candidates:
        prev = chosen.get(cand["sig"])
        if prev is None or (cand["explicit"] and not prev["explicit"]):
            chosen[cand["sig"]] = cand
    chart_specs = [(c["ref"], c["kind"], c["title"]) for c in chosen.values()]

    # Build the Özet sheet at the front, now that data sheets exist for chart refs.
    ozet = wb.create_sheet("Özet", 0)
    ozet.sheet_view.showGridLines = False
    next_row = header_band(ozet, title, company, period or "Tüm zamanlar")
    next_row = kpi_cards(ozet, cards, next_row + 1)
    # Anchor charts below the KPIs, stacked. CR-055 — advance by each chart's ACTUAL
    # (content-aware) height so a tall / many-category chart never overlaps the next.
    anchor_row = next_row + 1
    for ref, kind, ctitle in chart_specs:
        height_cm = add_chart(ozet, ref, kind, ctitle, f"A{anchor_row}")
        anchor_row += _chart_rows(height_cm)
    # CR-055 — honesty footnote for the USD-at-date basis (+ a note when rows had no FX).
    if results_usd is not None:
        anchor_row = _usd_footnote(ozet, anchor_row + 1, results_usd)
    for col in range(1, 9):
        ozet.column_dimensions[get_column_letter(col)].width = 18
    ozet.freeze_panes = "A6"
    ozet.print_area = f"A1:H{max(anchor_row, next_row)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Single-report convenience wrapper (xlsx branch of studio_export)
# --------------------------------------------------------------------------- #
def build_single_report(result: dict, title: str, viz: str | None, *,
                        company: str | None = None, result_usd: dict | None = None) -> bytes:
    """The single-report xlsx: one synthetic widget through the same engine, giving an
    Özet-lite (title band + headline KPIs + the report's own chart) + the data sheet.

    CR-055 — ``result_usd`` (the same report run with basis.currency='usd') adds the USD
    companion (₺+USD KPIs, paired $ columns, footnote). None → ₺-only (unchanged)."""
    wtype = "kpi" if viz == "kpi" else ("chart" if viz in ("line", "area", "bar") else "table")
    widget = {"id": "r", "type": wtype, "title": title, "spec": {"viz": viz}}
    if wtype == "kpi":
        # A kpi-viz report: render its totals as cards AND keep a data sheet via the
        # table path, so single_report still surfaces the table. Treat as report.
        widget["type"] = "report"
    return build_workbook([widget], {"r": result}, title, company=company,
                          period=_period_label(result), single_report=True,
                          results_usd=({"r": result_usd} if result_usd is not None else None))
