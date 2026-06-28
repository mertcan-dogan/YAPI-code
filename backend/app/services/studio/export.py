"""CR-033 — Report Studio export (pdf / xlsx / csv).

Turns a ``run_spec`` result (§2 shape) into a downloadable file. STRICTLY
READ-ONLY: it consumes an already-computed result dict and builds an in-memory
file — it issues no queries and never writes a row. The flat table is:

    header = [column.label for every column]
    each data row = dim label (dimension cols) / metric value (metric cols)
    a final "Toplam" row from result["totals"]["metrics"]

xlsx mirrors the openpyxl Workbook→bytes pattern in app/api/audit.py; pdf reuses
the Turkish-capable ReportLab helpers in app/services/reports.py
(``register_turkish_fonts`` / ``_styles`` / ``_data_table`` / ``_render_story``) —
no new PDF/Excel dependency is introduced. Chart-as-image is deferred.
"""
import csv
import io
import re
from datetime import datetime, timezone

from fastapi.responses import Response

from app.responses import APIError

CSV_BOM = "﻿"  # Excel needs the UTF-8 BOM to detect Turkish text correctly.

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_CSV_MEDIA = "text/csv; charset=utf-8"
_PDF_MEDIA = "application/pdf"

# ASCII transliteration for a safe Content-Disposition filename.
_TR_MAP = str.maketrans({
    "ş": "s", "Ş": "s", "ğ": "g", "Ğ": "g", "ı": "i", "İ": "i",
    "ö": "o", "Ö": "o", "ü": "u", "Ü": "u", "ç": "c", "Ç": "c",
})


def _slug(title: str) -> str:
    base = (title or "rapor").translate(_TR_MAP).lower()
    base = re.sub(r"[^a-z0-9]+", "-", base).strip("-")
    return base or "rapor"


# --------------------------------------------------------------------------- #
# Flatten the result into header / data rows / totals row
# --------------------------------------------------------------------------- #
def _flat_table(result: dict) -> tuple[list, list[list], list]:
    columns = result.get("columns") or []
    header = [c.get("label") or c.get("id") for c in columns]

    data_rows: list[list] = []
    for row in result.get("rows") or []:
        dims = row.get("dims") or {}
        metrics = row.get("metrics") or {}
        data_rows.append([
            dims.get(c["id"]) if c.get("kind") == "dimension" else metrics.get(c["id"])
            for c in columns
        ])

    totals = (result.get("totals") or {}).get("metrics") or {}
    totals_row: list = []
    label_placed = False
    for c in columns:
        if c.get("kind") == "dimension":
            totals_row.append("Toplam" if not label_placed else "")
            label_placed = True
        else:
            totals_row.append(totals.get(c["id"]))
    # KPI-style export (no dimension column): label the first cell so the totals
    # row is still distinguishable.
    if columns and not label_placed:
        totals_row[0] = totals_row[0]  # leave metric value; header already names it
    return header, data_rows, totals_row


# --------------------------------------------------------------------------- #
# Cell coercion per format
# --------------------------------------------------------------------------- #
# Formula / CSV-injection guard shared by EVERY spreadsheet + csv writer (the
# per-report _xlsx/_csv and the dashboard deck _dashboard_xlsx).
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(v):
    """Neutralize spreadsheet formula / CSV injection. A *text* cell beginning with a
    formula trigger (``= + - @`` or a leading TAB/CR) is prefixed with a single
    apostrophe so Excel/Sheets/LibreOffice render it as literal text rather than
    evaluate it (``=WEBSERVICE(...)`` / ``=HYPERLINK(...)`` / legacy DDE). Exports only
    ever hold the caller's own company data, but dimension labels are user-authored
    (project / vendor / supplier names), so a lower-privilege user could otherwise
    plant a payload that runs when a colleague opens the file.

    Only applied to genuine strings — callers must pass numbers as native int/float
    (never pre-stringified) so a legitimate negative value like ``-5`` is not mistaken
    for a formula. Numbers and other non-strings pass through unchanged."""
    if isinstance(v, str) and v[:1] in _FORMULA_TRIGGERS:
        return "'" + v
    return v


def _csv_cell(v) -> str:
    # Guard only original text (dimension labels); numbers are stringified WITHOUT the
    # guard so a negative metric (-5.0) stays a number, not "'-5.0".
    if v is None:
        return ""
    if isinstance(v, str):
        return _safe_cell(v)
    return str(v)


def _pdf_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else f"{v:.2f}"
    return str(v)


# --------------------------------------------------------------------------- #
# xlsx
# --------------------------------------------------------------------------- #
def _xlsx(header, data_rows, totals_row) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor"
    # None/int/float render natively; text cells are formula-guarded.
    ws.append([_safe_cell(c) for c in header])
    for r in data_rows:
        ws.append([_safe_cell(c) for c in r])
    if totals_row:
        ws.append([_safe_cell(c) for c in totals_row])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# csv (UTF-8 with BOM for Excel/Türkçe)
# --------------------------------------------------------------------------- #
def _csv(header, data_rows, totals_row) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([_csv_cell(v) for v in header])
    for r in data_rows:
        writer.writerow([_csv_cell(v) for v in r])
    if totals_row:
        writer.writerow([_csv_cell(v) for v in totals_row])
    return (CSV_BOM + buf.getvalue()).encode("utf-8")


# --------------------------------------------------------------------------- #
# pdf (reuse app/services/reports.py ReportLab helpers)
# --------------------------------------------------------------------------- #
def _has_unwindowed_metric(result: dict) -> bool:
    """True if any selected metric is a whole-project snapshot (ignores the date
    window) — drives the "tüm proje, bugüne kadar" note."""
    from app.services.studio.catalog import GRAIN_CASH, GRAIN_COST_LINE, METRICS

    for c in result.get("columns") or []:
        if c.get("kind") != "metric":
            continue
        m = METRICS.get(c["id"])
        if m and m["grain"] not in (GRAIN_COST_LINE, GRAIN_CASH):
            return True
    return False


def _meta_line(result: dict) -> str:
    meta = result.get("meta") or {}
    parts: list[str] = []
    dr = meta.get("date_range") or {}
    if dr.get("from") or dr.get("to"):
        parts.append(f"Dönem: {dr.get('from') or '—'} – {dr.get('to') or '—'}")
    basis = meta.get("basis") or {}
    parts.append(f"Para birimi: {(basis.get('currency') or 'try').upper()}")
    if _has_unwindowed_metric(result):
        parts.append("Bazı metrikler tüm proje, bugüne kadar (tarih aralığından bağımsız)")
    return "  ·  ".join(parts)


# --------------------------------------------------------------------------- #
# CR-037 — chart + toolkit rendering helpers (report_theme / report_charts)
# --------------------------------------------------------------------------- #
def _chart_inputs(result: dict):
    """Flatten ``result['series']`` into (labels, [(name,[y…])…]) aligned to one
    unified, order-preserving x-axis across all series (missing point → 0). Returns
    None when there is no plottable series."""
    series = result.get("series") or []
    if not series:
        return None
    labels: list[str] = []
    seen: set[str] = set()
    for se in series:
        for pt in se.get("points") or []:
            xs = "" if pt.get("x") is None else str(pt.get("x"))
            if xs not in seen:
                seen.add(xs)
                labels.append(xs)
    if not labels:
        return None
    aligned: list = []
    for se in series:
        ymap: dict = {}
        for pt in se.get("points") or []:
            xs = "" if pt.get("x") is None else str(pt.get("x"))
            try:
                ymap[xs] = float(pt.get("y")) if pt.get("y") is not None else 0.0
            except (TypeError, ValueError):
                ymap[xs] = 0.0
        aligned.append((str(se.get("name") or se.get("metric") or ""), [ymap.get(l, 0.0) for l in labels]))
    return labels, aligned


def _studio_chart_img(viz: str, result: dict, dest_dir: str):
    """Render the result's series as the chart matching ``viz`` → PNG path, or None.
    bar → grouped bars; line/area → line (1 series) else grouped bars (no chart_combo
    — that's bars+line and would misrepresent multiple line series)."""
    from app.services import report_charts as ch

    inp = _chart_inputs(result)
    if not inp:
        return None
    labels, series = inp
    if viz == "bar":
        return ch.chart_grouped_bar(labels, series, dest_dir)
    if viz in ("line", "area"):
        if len(series) == 1:
            return ch.chart_line(labels, series[0][1], dest_dir, fill=(viz == "area"))
        return ch.chart_grouped_bar(labels, series, dest_dir)
    return None


def _kpi_items(result: dict) -> list:
    """(metric label, formatted total) pairs from result totals — for kpi cards."""
    totals = (result.get("totals") or {}).get("metrics") or {}
    out = []
    for c in result.get("columns") or []:
        if c.get("kind") == "metric":
            out.append((str(c.get("label") or c.get("id")), _pdf_cell(totals.get(c["id"]))))
    return out


def _kpi_flowables(result: dict) -> list:
    """KPI cards from totals, tiled ≤4 per row (kpi viz / kpi widget)."""
    from reportlab.platypus import Spacer

    from app.services.report_theme import kpirow

    items = _kpi_items(result)
    out: list = []
    for i in range(0, len(items), 4):
        out.append(kpirow(items[i:i + 4], colw=4.07))
        out.append(Spacer(1, 8))
    return out


def _studio_dtable(header, data_rows, totals_row, columns):
    """Toolkit data table (NAVY header, zebra, optional bold totals). Metric columns
    right-aligned. dtable() escapes every cell, so user-authored dim labels are safe."""
    from reportlab.lib.units import cm

    from app.services.report_theme import dtable

    ncols = len(header) or 1
    colw = [(17.5 / ncols) * cm] * ncols
    aligns = [2 if (i < len(columns) and columns[i].get("kind") == "metric") else 0 for i in range(ncols)]
    rows = [[_pdf_cell(v) for v in r] for r in data_rows]
    totals = [_pdf_cell(v) for v in totals_row] if totals_row else None
    return dtable(header, rows, colw, aligns=aligns, totals=totals)


def _studio_furniture(eyebrow: str, generated_at: str):
    """Themed page painter (light BG, running eyebrow, footer) — the report_theme look."""
    def paint(c, doc):
        from reportlab.lib.colors import HexColor
        from reportlab.lib.pagesizes import A4

        from app.services.report_theme import (
            BG, FNT, HAIR, MUT, register_lato_fonts, LATO_REGULAR, LATO_SEMIBOLD,
        )

        register_lato_fonts()
        Wd, H = A4
        M = 40
        c.saveState()
        c.setFillColor(HexColor(BG)); c.rect(0, 0, Wd, H, fill=1, stroke=0)
        c.setFillColor(HexColor(MUT)); c.setFont(LATO_SEMIBOLD, 7)
        c.drawString(M, H - 30, eyebrow)
        c.setStrokeColor(HexColor(HAIR)); c.setLineWidth(0.6)
        c.line(M, H - 38, Wd - M, H - 38); c.line(M, 40, Wd - M, 40)
        c.setFillColor(HexColor(FNT)); c.setFont(LATO_REGULAR, 7.5)
        c.drawString(M, 28, f"Yapı tarafından {generated_at} tarihinde oluşturuldu")
        c.drawRightString(Wd - M, 28, f"Sayfa {doc.page}")
        c.restoreState()

    return paint


def _studio_doc(story: list, eyebrow: str, generated_at: str) -> bytes:
    """Build the story into PDF bytes with the themed page furniture."""
    from io import BytesIO

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=1.7 * cm, bottomMargin=1.6 * cm, leftMargin=1.4 * cm, rightMargin=1.4 * cm,
        title="Yapı Stüdyo",
    )
    paint = _studio_furniture(eyebrow, generated_at)
    doc.build(story, onFirstPage=paint, onLaterPages=paint)
    return buf.getvalue()


def _pdf(header, data_rows, totals_row, title: str, result: dict, viz: str | None = None) -> bytes:
    """CR-037 — report PDF on the report_theme toolkit. ``viz`` (from the report spec)
    decides the lead element: kpi → KPI cards from totals; line/area/bar → the matching
    chart in a chart card ABOVE the data table; table/other → table only. The data
    table (with totals + per-row deltas) is kept for every non-kpi viz. Read-only.
    Charts are PNGs in a tempfile.mkdtemp() cleaned up after doc.build()."""
    import shutil
    import tempfile
    from xml.sax.saxutils import escape

    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer

    from app.services.report_theme import chartcard, register_lato_fonts, s as TS, sect, LATO_REGULAR, MUT
    from app.utils.format import format_datetime_tr

    register_lato_fonts()
    tmpdir = tempfile.mkdtemp(prefix="yapi_studio_")
    CW = 16.0 * cm
    CHh = 5.0 * cm
    try:
        # sect()/chartcard() titles are NOT escaped by the toolkit → escape user title here.
        story = list(sect("RAPOR STÜDYOSU", escape(title or "Rapor")))
        meta_line = _meta_line(result)
        if meta_line:
            story += [Paragraph(escape(meta_line), TS("note", LATO_REGULAR, 8, MUT)), Spacer(1, 8)]

        if viz == "kpi":
            story += _kpi_flowables(result)
        else:
            if viz in ("line", "area", "bar"):
                img = _studio_chart_img(viz, result, tmpdir)
                if img:
                    story += [chartcard(escape(title or "Grafik"), img, CW, CHh), Spacer(1, 10)]
            story.append(_studio_dtable(header, data_rows, totals_row, result.get("columns") or []))

        generated_at = format_datetime_tr(datetime.now(timezone.utc))
        return _studio_doc(story, "RAPOR STÜDYOSU", generated_at)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


# --------------------------------------------------------------------------- #
# Public entrypoint
# --------------------------------------------------------------------------- #
def studio_export(result: dict, fmt: str, title: str, viz: str | None = None) -> Response:
    """Build a downloadable file (pdf/xlsx/csv) from a run_spec result. Read-only.

    ``viz`` (the report spec's visualization) drives the PDF chart (CR-037); it is
    ignored by the xlsx/csv paths, which stay byte-for-byte as before.

    Raises ``APIError(422, INVALID_FORMAT)`` for an unknown format (the router also
    guards, so this is defense-in-depth)."""
    header, data_rows, totals_row = _flat_table(result)
    slug = _slug(title)

    if fmt == "xlsx":
        content = _xlsx(header, data_rows, totals_row)
        media, ext = _XLSX_MEDIA, "xlsx"
    elif fmt == "csv":
        content = _csv(header, data_rows, totals_row)
        media, ext = _CSV_MEDIA, "csv"
    elif fmt == "pdf":
        content = _pdf(header, data_rows, totals_row, title, result, viz=viz)
        media, ext = _PDF_MEDIA, "pdf"
    else:
        raise APIError(422, "INVALID_FORMAT", "Geçersiz dışa aktarma biçimi (pdf, xlsx veya csv)")

    return Response(
        content=content,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{slug}.{ext}"'},
    )


# --------------------------------------------------------------------------- #
# CR-034 — dashboard (pano) deck export
# --------------------------------------------------------------------------- #
# Excel forbids these in a sheet name and caps it at 31 chars; names must also be
# unique. (Kept separate from studio_export so that path stays byte-for-byte.)
_XLSX_FORBIDDEN = re.compile(r"[\[\]:*?/\\]")


def _ordered_widgets(widgets: list) -> list:
    """Section-grouped, then dashboard-array order: sections appear in first-seen
    order, and within a section widgets keep their array order. Widgets without a
    section group under a stable bucket in array order."""
    buckets: dict = {}
    order: list = []
    for w in widgets or []:
        sec = w.get("section")
        if sec not in buckets:
            buckets[sec] = []
            order.append(sec)
        buckets[sec].append(w)
    ordered: list = []
    for sec in order:
        ordered.extend(buckets[sec])
    return ordered


def _is_renderable(res) -> bool:
    """A widget result carries a table iff it is a dict that is neither empty nor
    the ``{"unavailable": True}`` status sentinel."""
    return isinstance(res, dict) and bool(res) and not res.get("unavailable")


def _sheet_name(title: str, used: set) -> str:
    """Sanitised, ≤31-char, de-duplicated openpyxl sheet name."""
    base = _XLSX_FORBIDDEN.sub(" ", (title or "Sayfa")).strip() or "Sayfa"
    base = base[:31]
    name = base
    i = 1
    used_cf = {u.casefold() for u in used}
    while name.casefold() in used_cf:
        suffix = f" ({i})"
        name = base[: 31 - len(suffix)] + suffix
        i += 1
    return name


def _dashboard_pdf(ordered: list, results: dict, title: str) -> bytes:
    """CR-037 — dashboard deck PDF on the report_theme toolkit, each widget rendered as
    the right element in reading order: kpi → KPI cards; chart → its chart (the widget
    spec's viz) in a chart card; table → data table; text → paragraph; report → its chart
    (inferred from the result's series) + table; unavailable → a calm placeholder. A deck
    with no renderable widget still returns a header-only PDF (never an error). Read-only;
    charts are PNGs in a tempfile.mkdtemp() cleaned up after doc.build()."""
    import shutil
    import tempfile
    from xml.sax.saxutils import escape

    from reportlab.platypus import Paragraph, Spacer

    from reportlab.lib.units import cm

    from app.services.report_theme import (
        chartcard, register_lato_fonts, s as TS, sect,
        INK, MUT, NAVY, LATO_BOLD, LATO_REGULAR, LATO_SEMIBOLD,
    )
    from app.utils.format import format_datetime_tr

    register_lato_fonts()
    tmpdir = tempfile.mkdtemp(prefix="yapi_pano_")
    CW = 16.0 * cm
    CHh = 5.0 * cm
    try:
        story = list(sect("PANO", escape(title or "Pano")))

        last_section = object()  # sentinel so the first (even None) section prints once
        for w in ordered:
            wtype = w.get("type")
            wid = w.get("id")
            section = w.get("section")
            if section and section != last_section:
                story.append(Paragraph(escape(str(section).upper()), TS("sec", LATO_SEMIBOLD, 8.5, MUT, spaceBefore=6)))
            last_section = section

            if wtype == "text":
                content = w.get("content")
                if content:
                    story.append(Paragraph(escape(str(content)), TS("body", LATO_REGULAR, 9, INK, leading=13)))
                    story.append(Spacer(1, 6))
                continue

            res = results.get(wid)
            wtitle = escape(w.get("title") or "")
            if isinstance(res, dict) and res.get("unavailable"):
                story.append(Paragraph(f"<b>{wtitle}</b>", TS("wt", LATO_BOLD, 9.5, NAVY)))
                story.append(Paragraph("Bu rapor artık kullanılamıyor", TS("note", LATO_REGULAR, 8, MUT)))
                story.append(Spacer(1, 8))
                continue
            if not _is_renderable(res):
                continue

            story.append(Paragraph(f"<b>{wtitle}</b>", TS("wt", LATO_BOLD, 9.5, NAVY)))
            header, data_rows, totals_row = _flat_table(res)
            columns = res.get("columns") or []

            if wtype == "kpi":
                story += _kpi_flowables(res)
            elif wtype == "chart":
                viz = (w.get("spec") or {}).get("viz") or "line"
                img = _studio_chart_img(viz, res, tmpdir)
                if img:
                    story.append(chartcard(wtitle, img, CW, CHh))
                else:
                    story.append(_studio_dtable(header, data_rows, totals_row, columns))
            elif wtype == "report":
                # The deck payload carries no resolved viz for a report widget → infer
                # the chart family from the result's series (1 → line, ≥2 → grouped bars);
                # always keep the table beneath it.
                if res.get("series"):
                    img = _studio_chart_img("line", res, tmpdir)
                    if img:
                        story += [chartcard(wtitle, img, CW, CHh), Spacer(1, 6)]
                story.append(_studio_dtable(header, data_rows, totals_row, columns))
            else:  # table (and any unknown data widget) → table
                story.append(_studio_dtable(header, data_rows, totals_row, columns))
            story.append(Spacer(1, 12))

        generated_at = format_datetime_tr(datetime.now(timezone.utc))
        return _studio_doc(story, "PANO", generated_at)
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def _dashboard_xlsx(ordered: list, results: dict) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    used_names: set = set()
    sheet_count = 0
    for w in ordered:
        if w.get("type") == "text":
            continue
        res = results.get(w.get("id"))
        if not _is_renderable(res):
            continue
        header, data_rows, totals_row = _flat_table(res)
        name = _sheet_name(w.get("title"), used_names)
        used_names.add(name)
        ws = default_ws if sheet_count == 0 else wb.create_sheet(title=name)
        if sheet_count == 0:
            ws.title = name
        ws.append([_safe_cell(c) for c in header])
        for r in data_rows:
            ws.append([_safe_cell(c) for c in r])
        if totals_row:
            ws.append([_safe_cell(c) for c in totals_row])
        sheet_count += 1

    if sheet_count == 0:
        raise APIError(422, "NO_DATA", "Dışa aktarılacak veri yok")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def studio_export_dashboard(widgets: list, results: dict, title: str, fmt: str) -> Response:
    """Build a dashboard deck file (pdf/xlsx) from already-computed batch results.

    READ-ONLY: consumes ``results`` ({widget_id: run_spec-result-or-status}) and the
    ``widgets`` array (each a dict with id/type/title/section/content), grouped by
    section then dashboard-array order. ``studio_export`` and all its helpers are
    left untouched.

    * ``pdf`` — title + each renderable widget (text → paragraph; unavailable →
      a note; data/report → title + flattened table). No renderable widget ⇒ still a
      title-only PDF (never an error).
    * ``xlsx`` — one sheet per data/report widget that has a result; zero data
      sheets ⇒ ``APIError(422)`` (no empty workbook).
    * ``csv`` — not offered for a multi-widget pano ⇒ ``APIError(422)``.
    """
    ordered = _ordered_widgets(widgets)
    slug = _slug(title)

    if fmt == "pdf":
        content = _dashboard_pdf(ordered, results, title)
        media, ext = _PDF_MEDIA, "pdf"
    elif fmt == "xlsx":
        content = _dashboard_xlsx(ordered, results)
        media, ext = _XLSX_MEDIA, "xlsx"
    elif fmt == "csv":
        raise APIError(422, "INVALID_FORMAT", "Pano CSV olarak dışa aktarılamaz (pdf veya xlsx)")
    else:
        raise APIError(422, "INVALID_FORMAT", "Geçersiz dışa aktarma biçimi (pdf veya xlsx)")

    return Response(
        content=content,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{slug}.{ext}"'},
    )
