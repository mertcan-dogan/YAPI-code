"""PDF report generation via ReportLab (Section 4.9, CR-004-A).

Switched from WeasyPrint to ReportLab so reports render on Windows without the
GTK/Pango/Cairo system libraries WeasyPrint requires. The Turkish content, colour
palette, company logo, title/date, page numbers and generated-by footer are
preserved; the monthly management pack keeps its 7-page structure.

Data gathering (``build_*_data``) is kept free of ReportLab imports so it stays
unit-testable, and the actual rendering lives in ``_*_pdf`` helpers that tests can
stub. The AI summary text still comes from the same Claude API calls.
"""
import os
from datetime import datetime, timezone

from sqlalchemy.orm import Session

from app.calculations.money import D
from app.constants import COST_CATEGORIES
from app.models.company import Company
from app.models.project import Project
from app.services.financials import project_financials
from app.utils.format import (
    format_currency_tr,
    format_date_tr,
    format_datetime_tr,
    format_number_tr,
    format_pct_tr,
)

# CR-005-A: Türkçe karakterleri (ş, ğ, ı, İ, ü, ö, ç) destekleyen Unicode fontlar.
# ReportLab'ın varsayılan Helvetica/Times fontları Latin-1 ile sınırlı olduğundan
# Türkçe karakterler PDF'te ■ olarak görünüyordu. Bu üç font (app/fonts/) ile
# tüm metin/tablo/grafiklerde Türkçe karakterler doğru render edilir.
FONTS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "fonts")
FONT_NORMAL = "DejaVuSans"
FONT_BOLD = "DejaVuSans-Bold"
FONT_OBLIQUE = "DejaVuSans-Oblique"

_fonts_registered = False


def register_turkish_fonts():
    """Register the Türkçe-capable TTF fonts with ReportLab (idempotent)."""
    global _fonts_registered
    if _fonts_registered:
        return
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    pdfmetrics.registerFont(TTFont(FONT_NORMAL, os.path.join(FONTS_DIR, "DejaVuSans.ttf")))
    pdfmetrics.registerFont(TTFont(FONT_BOLD, os.path.join(FONTS_DIR, "DejaVuSans-Bold.ttf")))
    pdfmetrics.registerFont(TTFont(FONT_OBLIQUE, os.path.join(FONTS_DIR, "DejaVuSans-Oblique.ttf")))
    # Map the bold/italic variants so <b>/<i> inline markup and bold styles resolve.
    from reportlab.pdfbase.pdfmetrics import registerFontFamily

    registerFontFamily(
        FONT_NORMAL, normal=FONT_NORMAL, bold=FONT_BOLD,
        italic=FONT_OBLIQUE, boldItalic=FONT_BOLD,
    )
    _fonts_registered = True


# Yapı colour palette (unchanged from the WeasyPrint version).
PRIMARY = "#1B2B4B"
ACCENT = "#F59E0B"
BORDER = "#E2E8F0"
MUTED = "#64748B"

STATUS_LABELS = {"red": "Kritik", "amber": "Dikkat", "green": "İyi"}
RAG_COLORS = {"red": "#EF4444", "amber": "#F59E0B", "green": "#10B981"}
# CR-006-A: light tints used as cell BACKGROUND for colour-coded table cells.
RAG_TINTS = {"red": "#FEE2E2", "amber": "#FEF3C7", "green": "#DCFCE7", "gray": "#F1F5F9"}
NAVY = "#1E3A5F"   # CR-006-A cover/header band colour (per spec)

# CR-004-F: every AI-written output carries this disclaimer; the management pack
# embeds it in the PDF footer because pages 1 and 7 are AI-generated.
AI_DISCLAIMER = (
    "Bu içerik yapay zeka tarafından oluşturulmuştur ve hatalar içerebilir. "
    "Önemli finansal kararlar almadan önce lütfen doğrulayın."
)

# CR-036: the eleven management-pack sections (the new "Aylık Yönetim Raporu").
# Kept as data so the structure is verifiable without rendering a PDF. A section
# is rendered ONLY when its data exists (honesty rule §0); otherwise omitted.
SECTION_TITLES = [
    "1. Kapak",
    "2. Yönetici Özeti",
    "3. Finansal Performans",
    "4. Taahhüt & Maliyet Maruziyeti",
    "5. Kur & Döviz",
    "6. Kritik Projeler",
    "7. Nakit & İşletme Sermayesi",
    "8. Tedarikçi & Taşeron",
    "9. Satış, m² & Getiri",
    "10. Veri Güvence & Anomali",
    "11. Risk & Aksiyon Planı",
]

# CR-036 Heneka palette as PLAIN hex strings (mirrors report_theme). Defined here
# so build_management_pack_data can colour KPI tuples WITHOUT importing
# report_theme (which pulls in ReportLab) — keeping the data layer import-light.
H_NAVY = "#183047"
H_PETROL = "#0E625B"
H_GOLD = "#B9852A"
H_MUT = "#65726F"
H_RED = "#B94D45"
H_AMBER = "#C98E24"
H_GREEN = "#27815F"

# CR-036 §10: assurance finding-type → Turkish label (for the anomaly chart/table).
_ALERT_TYPE_LABELS = {
    "duplicate_cost": "Mükerrer maliyet",
    "duplicate_invoice": "Mükerrer hakediş",
    "cost_outlier": "Olağandışı maliyet",
    "kdv_mismatch": "KDV uyumsuzluğu",
    "hakedis_over_contract": "Hakediş > sözleşme",
    "missing_fx": "Eksik kur kaydı",
    "unlinked_vendor": "Bağlanmamış maliyet",
    "nonpositive_amount": "Sıfır/negatif tutar",
}
_SEVERITY_LABELS = {"high": "Yüksek", "medium": "Orta", "low": "Düşük"}
_SEVERITY_RAG = {"high": "r", "medium": "a", "low": "g"}


# ---------------------------------------------------------------------------
# Project status report
# ---------------------------------------------------------------------------
def build_project_report_data(db: Session, project: Project, company: Company) -> dict:
    """Gather everything the project report needs (no ReportLab import)."""
    f = project_financials(db, project)
    now = datetime.now(timezone.utc)
    categories = [
        {
            "label": COST_CATEGORIES.get(c["cost_category"], c["cost_category"]),
            "revised": format_currency_tr(c["revised_budget_try"]),
            "invoiced": format_currency_tr(c["invoiced_try"]),
            "forecast": format_currency_tr(c["forecast_final"]),
            "variance": format_currency_tr(c["variance_try"]),
            "status": c["status"],
            "status_label": STATUS_LABELS.get(c["status"], ""),
        }
        for c in f["categories"]
    ]
    # CR-031-C: revenue-model-aware Satışlar & Kar/Zarar summary, surfaced in the
    # report data. Revenue is sell-side (sales+landowner) OR hakediş per
    # revenue_model — never both (§0.2). Read-only over cost.
    from app.services import sales as sales_service

    p = sales_service.project_pnl(db, project)
    kur = p["fx_effect"]
    sales_pnl = {
        "revenue_source": p["revenue_source"],
        "revenue": format_currency_tr(p["revenue_try"]),
        "cost": format_currency_tr(p["cost_try"]),
        "financing": format_currency_tr(p["financing_try"]),
        "net_excl_financing": format_currency_tr(p["net_excl_financing_try"]),
        "net_incl_financing": format_currency_tr(p["net_incl_financing_try"]),
        "margin_pct": format_pct_tr(p["margin_pct"]) if p["margin_pct"] is not None else "—",
        "fx_effect": format_currency_tr(kur["fx_effect_try"]) if kur["fx_effect_try"] is not None else "—",
    }

    return {
        "company_name": company.name,
        "logo_url": company.logo_url,
        "report_title": "Proje Durum Raporu",
        "report_date": format_date_tr(now.date()),
        "generated_at": format_datetime_tr(now),
        "project_name": project.name,
        "client_name": project.client_name,
        "contract_value": format_currency_tr(f["contract_value_try"]),
        "total_actual": format_currency_tr(f["total_actual_with_vat_try"]),
        "forecast_final": format_currency_tr(f["forecast_final_cost_try"]),
        "margin_pct": format_pct_tr(f["margin_pct"]),
        "rag_status": f["rag_status"],
        "categories": categories,
        "total_invoiced": format_currency_tr(f["total_invoiced_try"]),
        "total_collected": format_currency_tr(f["total_collected_try"]),
        "total_outstanding": format_currency_tr(f["total_outstanding_try"]),
        "total_retention": format_currency_tr(f["total_retention_try"]),
        "net_cash": format_currency_tr(f["net_cash_position_try"]),
        "sales_pnl": sales_pnl,
    }


def render_project_report(db: Session, project: Project, company: Company) -> bytes:
    return _project_report_pdf(build_project_report_data(db, project, company))


# ---------------------------------------------------------------------------
# Monthly Management Pack (CR-003-K) — 7 pages
# ---------------------------------------------------------------------------
def build_management_pack_data(db: Session, company: Company, period_label: str) -> dict:
    """Gather the CR-036 "Aylık Yönetim Raporu" data (no ReportLab/matplotlib
    import; unit-testable). Produces numeric, render-agnostic keys for all 11
    sections (§2). Reuses the portfolio rollup, ``cover_kpis`` and the existing
    helpers (_active_projects, _company_overdue_payments, _subcontractor_commitments).

    HONESTY (§0): every section carries enough data for the render to decide OMIT
    (empty list / has_sell_side False / has_fx False / no active projects). No
    fabricated confidence/coverage; no EVM/EBITDA/HSE/13-week-treasury numbers.
    """
    from datetime import date

    from app.constants import PROJECT_TYPES, SELL_SIDE_REVENUE_MODELS
    from app.services import ai as ai_service
    from app.services import sales as sales_service
    from app.services.financials import forecast_at_completion, project_cashflow

    projects = _active_projects(db, company)
    project_lookup = {p.id: p.name for p in projects}

    today = date.today()
    anchor = _parse_period_anchor(period_label)
    rows = []                         # §6 critical-project rows
    # Portfolio rollup (real, summed — no EBITDA / net-debt).
    P = {k: D(0) for k in (
        "contract", "revised_budget", "actual", "actual_with_vat", "open_committed",
        "exposure", "forecast", "invoiced", "collected", "outstanding", "net_cash",
        "revenue", "cost", "net_excl", "net_incl", "fx_effect")}
    cat_totals: dict[str, dict] = {}  # per-category rollup (commitment + margin bridge)
    cash_in: dict[str, object] = {}   # monthly income (trailing months)
    cash_out: dict[str, object] = {}  # monthly expense (trailing months)
    worst_rag = "green"
    has_sell_side = any(p.revenue_model in SELL_SIDE_REVENUE_MODELS for p in projects)

    for p in projects:
        f = project_financials(db, p)
        fac = forecast_at_completion(db, p)
        pnl = sales_service.project_pnl(db, p)
        forecast_margin = float(fac["forecast_final_margin_pct"])
        target_margin = float(p.target_margin_pct) if p.target_margin_pct is not None else 0.0
        outstanding_high = D(f["total_outstanding_try"]) > D(f["contract_value_try"]) * D("0.20")
        rows.append({
            "name": p.name,
            "client": p.client_name,
            "project_type": PROJECT_TYPES.get(p.project_type, p.custom_project_type or "—"),
            "contract": format_currency_tr(f["contract_value_try"]),
            "budget": format_currency_tr(f["revised_budget_try"]),
            "actual": format_currency_tr(f["total_actual_with_vat_try"]),
            "completion": format_pct_tr(f["completion_pct"]),
            "target_margin": format_pct_tr(target_margin),
            "margin": format_pct_tr(fac["forecast_final_margin_pct"]),
            "margin_value": round(forecast_margin, 1),
            "outstanding": format_currency_tr(f["total_outstanding_try"]),
            "outstanding_high": outstanding_high,
            "commercial_note": f.get("rag_reason_tr") or "—",
            "rag": f["rag_status"],
        })
        worst_rag = _worse_rag(worst_rag, f["rag_status"])

        # Per-category rollup (CR-023 committed/open/exposure + variance for §3/§4).
        for c in f["categories"]:
            key = c["cost_category"]
            t = cat_totals.setdefault(key, {
                "revised": D(0), "open_committed": D(0), "invoiced": D(0),
                "exposure": D(0), "variance": D(0)})
            t["revised"] += D(c["revised_budget_try"])
            t["open_committed"] += D(c["open_committed_try"])
            t["invoiced"] += D(c["invoiced_try"])
            t["exposure"] += D(c["exposure_try"])
            t["variance"] += D(c["variance_try"])

        # Trailing monthly cashflow (effective in/out for past + current months).
        for m in project_cashflow(db, p, today=anchor):
            if not (m["is_past"] or m["is_current"]):
                continue
            cash_in[m["month"]] = D(cash_in.get(m["month"], D(0))) + D(m["actual_in_try"])
            cash_out[m["month"]] = D(cash_out.get(m["month"], D(0))) + D(m["actual_out_try"])

        P["contract"] += D(f["contract_value_try"])
        P["revised_budget"] += D(f["revised_budget_try"])
        P["actual"] += D(f["total_actual_try"])
        P["actual_with_vat"] += D(f["total_actual_with_vat_try"])
        P["open_committed"] += D(f["total_open_committed_try"])
        P["exposure"] += D(f["total_committed_exposure_try"])
        P["forecast"] += D(fac["forecast_final_cost_try"])
        P["invoiced"] += D(f["total_invoiced_try"])
        P["collected"] += D(f["total_collected_try"])
        P["outstanding"] += D(f["total_outstanding_try"])
        P["net_cash"] += D(f["net_cash_position_try"])
        P["revenue"] += D(pnl["revenue_try"])
        P["cost"] += D(pnl["cost_try"])
        P["net_excl"] += D(pnl["net_excl_financing_try"])
        P["net_incl"] += D(pnl["net_incl_financing_try"])
        kur = (pnl.get("fx_effect") or {}).get("fx_effect_try")
        if kur is not None:
            P["fx_effect"] += D(kur)

    rows.sort(key=lambda r: _RAG_ORDER.get(r["rag"], 0), reverse=True)

    collected_pct = _safe_pct_str(P["collected"], P["invoiced"])
    margin_excl_pct = _safe_pct_str(P["net_excl"], P["revenue"])
    margin_incl_pct = _safe_pct_str(P["net_incl"], P["revenue"])

    # ---- §2 Yönetici Özeti — executive KPI cards (6) -----------------------
    exec_kpis = [
        ("GELİR", format_currency_tr(P["revenue"]), "Dönem geliri", H_MUT, H_PETROL),
        ("MALİYET", format_currency_tr(P["cost"]), "Gerçekleşen + tahmini", H_MUT, H_PETROL),
        ("NET KÂR (FİN. HARİÇ)", format_currency_tr(P["net_excl"]),
         f"Marj {margin_excl_pct}", _good_bad(P["net_excl"] >= 0), H_PETROL),
        ("MALİYET MARUZİYETİ", format_currency_tr(P["exposure"]),
         "Gerçekleşen + açık taahhüt", H_AMBER, H_GOLD),
        ("NAKİT POZİSYONU", format_currency_tr(P["net_cash"]),
         "Tahsilat − ödeme", _good_bad(P["net_cash"] >= 0), H_PETROL),
        ("BEKLEYEN TAHSİLAT", format_currency_tr(P["outstanding"]),
         f"{collected_pct} tahsil edildi", H_AMBER, H_GOLD),
    ]

    # ---- §10 Veri Güvence (READ-ONLY — collect_findings, never scan_company) -
    assurance = _assurance_section(db, company, today)
    high_findings = assurance.pop("high_findings")

    # ---- §2 decisions + early warning (rules + high-severity findings) ------
    overdue_payments = _company_overdue_payments(db, company, today)
    decisions = _build_decisions(rows, overdue_payments, cat_totals, high_findings, project_lookup)
    early_warning = _early_warning(high_findings, decisions, overdue_payments)

    # ---- §3 Finansal Performans -------------------------------------------
    fin_kpis = [
        ("GELİR", format_currency_tr(P["revenue"]), "Dönem geliri", H_MUT, H_PETROL),
        ("MALİYET", format_currency_tr(P["cost"]), "Tahmini final maliyet", H_MUT, H_PETROL),
        ("NET KÂR (FİN. HARİÇ)", format_currency_tr(P["net_excl"]),
         f"Marj {margin_excl_pct}", _good_bad(P["net_excl"] >= 0), H_PETROL),
        ("NET KÂR (FİN. DAHİL)", format_currency_tr(P["net_incl"]),
         f"Marj {margin_incl_pct}", _good_bad(P["net_incl"] >= 0), H_PETROL),
    ]
    # Monthly trend (trailing months) for the combo chart (bar=gelir, line=net).
    months = sorted(set(cash_in) | set(cash_out))[-6:]
    monthly_trend = [{
        "month": mk,
        "label": _month_label_tr(mk),
        "income": float(cash_in.get(mk, D(0))),
        "expense": float(cash_out.get(mk, D(0))),
        "net": float(D(cash_in.get(mk, D(0))) - D(cash_out.get(mk, D(0)))),
    } for mk in months]
    margin_bridge = _margin_bridge(cat_totals, P)

    # ---- §4 Taahhüt & Maliyet Maruziyeti ----------------------------------
    commitment_kpis = [
        ("GERÇEKLEŞEN MALİYET", format_currency_tr(P["actual"]),
         f"Bütçenin {_safe_pct_str(P['actual'], P['revised_budget'])}'i", H_MUT, H_PETROL),
        ("AÇIK TAAHHÜT", format_currency_tr(P["open_committed"]),
         "Faturalanmamış sipariş", H_AMBER, H_GOLD),
        ("MALİYET MARUZİYETİ", format_currency_tr(P["exposure"]),
         "Gerçekleşen + açık", H_AMBER, H_GOLD),
        ("TAHMİNİ FİNAL", format_currency_tr(P["forecast"]),
         f"Bütçe {format_currency_tr(P['revised_budget'])}",
         _good_bad(P["forecast"] <= P["revised_budget"]), H_GOLD),
    ]
    commitment_categories, commitment_chart = _commitment_categories(cat_totals)

    # ---- §5 Kur & Döviz (omitted by render when has_fx False) --------------
    fx = _fx_section(db, company, projects, P["fx_effect"])

    # ---- §7 Nakit & İşletme Sermayesi --------------------------------------
    # AR aging — replicate app/api/projects.py:901 _ar_aging EXACTLY (same buckets
    # + outstanding-weighted DSO, over the active project ids).
    ar_aging = _ar_aging_for_company(db, list(project_lookup.keys()), today)
    dso = ar_aging["dso_days"]
    cash_kpis = [
        ("NAKİT POZİSYONU", format_currency_tr(P["net_cash"]),
         "Tahsilat − ödeme", _good_bad(P["net_cash"] >= 0), H_PETROL),
        ("TAHSİL EDİLEN", format_currency_tr(P["collected"]),
         f"{collected_pct} tahsilat", H_GREEN, H_PETROL),
        ("BEKLEYEN TAHSİLAT", format_currency_tr(P["outstanding"]),
         "Açık alacak", H_AMBER, H_GOLD),
        ("DSO", f"{dso} gün" if dso is not None else "—",
         "Ort. tahsilat süresi", H_MUT, H_PETROL),
    ]

    # ---- §8 Tedarikçi & Taşeron -------------------------------------------
    vendor = _vendor_spend(db, company)
    subcontractor_commitments = _subcontractor_commitments(db, company)
    open_sub_balance = sum((D(s["revised_d"]) - D(_strip_money(s["paid"]))
                            for s in subcontractor_commitments), D(0))
    vendor_kpis = [
        ("AKTİF TEDARİKÇİ", str(vendor["active_count"]), "Harcama yapılan", H_MUT, H_PETROL),
        ("İLK 5 YOĞUNLAŞMA", format_pct_tr(vendor["concentration_pct"]),
         "Toplam harcamanın", H_AMBER, H_GOLD),
        ("AKTİF TAŞERON", str(len(subcontractor_commitments)), "Sözleşmeli", H_MUT, H_PETROL),
        ("AÇIK TAŞERON BAKİYE", format_currency_tr(open_sub_balance),
         "Kalan hakediş", H_AMBER, H_GOLD),
    ]

    # ---- §9 Satış, m² & Getiri (omitted by render unless has_sell_side) -----
    sell_side = _sell_side_section(db, projects) if has_sell_side else None

    # ---- §11 Risk & Aksiyon Planı -----------------------------------------
    risk_register, action_plan = _risk_and_actions(decisions)

    ai_summary = ai_service.management_summary({
        "sirket": company.name,
        "donem": period_label,
        "proje_sayisi": len(projects),
        "toplam_sozlesme": str(P["contract"]),
        "toplam_gelir": str(P["revenue"]),
        "toplam_maliyet": str(P["cost"]),
        "net_kar": str(P["net_excl"]),
        "toplam_bekleyen_tahsilat": str(P["outstanding"]),
    })

    cover_kpis = {
        "active_projects": str(len(projects)),
        "total_contract": format_currency_tr(P["contract"]),
        "total_outstanding": format_currency_tr(P["outstanding"]),
        "risk_level": _RISK_LABELS[worst_rag],
        "risk_rag": worst_rag,
    }

    return {
        "company_name": company.name,
        "logo_url": company.logo_url,
        "period": period_label,
        "generated_at": format_datetime_tr(datetime.now(timezone.utc)),
        "is_sample": False,
        "section_titles": list(SECTION_TITLES),
        "cover_kpis": cover_kpis,
        # §2
        "ai_summary": ai_summary,
        "exec_kpis": exec_kpis,
        "decisions": decisions,
        "early_warning": early_warning,
        # §3
        "fin_kpis": fin_kpis,
        "monthly_trend": monthly_trend,
        "margin_bridge": margin_bridge,
        # §4
        "commitment_kpis": commitment_kpis,
        "commitment_categories": commitment_categories,
        "commitment_chart": commitment_chart,
        # §5
        "has_fx": fx["has_fx"],
        "fx_kpis": fx["fx_kpis"],
        "fx_split": fx["fx_split"],
        "fx_note": fx["fx_note"],
        # §6
        "rows": rows,
        # §7
        "cash_kpis": cash_kpis,
        "ar_aging": ar_aging,
        # §8
        "vendor_spend": vendor["top"],
        "vendor_concentration": {
            "first5_pct": vendor["concentration_pct"], "active_count": vendor["active_count"]},
        "vendor_kpis": vendor_kpis,
        "subcontractor_commitments": subcontractor_commitments,
        # §9
        "has_sell_side": bool(sell_side),
        "sell_side": sell_side,
        # §10
        "assurance": assurance,
        # §11
        "risk_register": risk_register,
        "action_plan": action_plan,
    }


# CR-006-A: cover-page risk label per portfolio RAG.
_RISK_LABELS = {"green": "Düşük", "amber": "Orta", "red": "Yüksek"}
_RAG_ORDER = {"green": 0, "amber": 1, "red": 2}


def _worse_rag(a: str, b: str) -> str:
    return a if _RAG_ORDER.get(a, 0) >= _RAG_ORDER.get(b, 0) else b


def _margin_rag(pct: float) -> str:
    """§3/§2 margin colour: < %5 kırmızı, %5–%10 amber, > %10 yeşil."""
    if pct < 5:
        return "red"
    if pct <= 10:
        return "amber"
    return "green"


# ---------------------------------------------------------------------------
# CR-036 data-layer helpers (NO ReportLab/matplotlib — render-agnostic numbers).
# ---------------------------------------------------------------------------
def _safe_pct_str(numerator, denominator) -> str:
    """Formatted ``numerator/denominator*100`` Turkish %, or '—' when undefined."""
    den = D(denominator)
    if den == 0:
        return "—"
    return format_pct_tr(float(D(numerator) / den * 100))


def _good_bad(is_good: bool) -> str:
    """Context colour for a KPI: green when good, red otherwise."""
    return H_GREEN if is_good else H_RED


def _strip_money(formatted: str):
    """Parse a Turkish currency string ('1.234,56 ₺') back to a Decimal."""
    s = str(formatted or "").replace("₺", "").replace("$", "").strip()
    s = s.replace(".", "").replace(",", ".")
    try:
        return D(s)
    except Exception:  # noqa: BLE001
        return D(0)


def _build_decisions(rows, overdue_payments, cat_totals, high_findings, project_lookup) -> list[dict]:
    """§2 "Karar Gerektiren Konular": restructure the action rules (overdue /
    budget-overrun / large-outstanding / low-margin) PLUS high-severity assurance
    findings into decision dicts {konu, gerekce, sahip, beklenen_etki, oncelik_rag}."""
    decisions: list[dict] = []

    # Rule 1 — overdue payables.
    for o in overdue_payments[:3]:
        decisions.append({
            "konu": f"{o['supplier']} — vadesi geçmiş ödeme",
            "gerekce": f"{o['amount']}, {o['days']} gün gecikti",
            "sahip": "Finans",
            "beklenen_etki": "Tedarikçi ilişkisi ve gecikme riski ↓",
            "oncelik_rag": "r" if o["severe"] else "a",
        })

    # Rule 2 — budget overruns (exposure > revised budget per category).
    overruns = sorted(
        ((COST_CATEGORIES.get(k, k), t) for k, t in cat_totals.items()
         if D(t["revised"]) > 0 and D(t["variance"]) > 0),
        key=lambda kt: D(kt[1]["variance"]), reverse=True)
    for label, t in overruns[:2]:
        decisions.append({
            "konu": f"{label} — bütçe aşımı",
            "gerekce": f"Maruziyet revize bütçeyi {format_currency_tr(t['variance'])} aşıyor",
            "sahip": "Proje Müdürü",
            "beklenen_etki": "Maliyet kontrolü ve marj koruması",
            "oncelik_rag": "r",
        })

    # Rule 3 — large outstanding receivables.
    for r in rows:
        if r["outstanding_high"]:
            decisions.append({
                "konu": f"{r['name']} — yüksek bekleyen tahsilat",
                "gerekce": f"{r['outstanding']} tahsilat bekliyor",
                "sahip": "Ticari",
                "beklenen_etki": "Nakit girişini hızlandırma",
                "oncelik_rag": "a",
            })

    # Rule 4 — low forecast margin.
    for r in rows:
        if r["margin_value"] < 10:
            decisions.append({
                "konu": f"{r['name']} — düşük marj",
                "gerekce": f"Tahmini final marj {r['margin']}",
                "sahip": "Proje Müdürü",
                "beklenen_etki": "Maliyet kontrolü gerekli",
                "oncelik_rag": "r" if r["margin_value"] < 5 else "a",
            })

    # Rule 5 — high-severity assurance findings (review, not accusation).
    for fnd in high_findings[:3]:
        pname = project_lookup.get(fnd.get("project_id"), "")
        decisions.append({
            "konu": fnd["title_tr"] + (f" — {pname}" if pname else ""),
            "gerekce": fnd["body_tr"],
            "sahip": "Finans / Güvence",
            "beklenen_etki": fnd.get("recommended_action") or "İnceleme önerilir",
            "oncelik_rag": "r",
        })

    if not decisions:
        decisions.append({
            "konu": "Acil karar gerektiren konu yok",
            "gerekce": "Bu dönemde kritik finansal risk tespit edilmedi",
            "sahip": "—",
            "beklenen_etki": "—",
            "oncelik_rag": "g",
        })
    return decisions[:8]


def _early_warning(high_findings, decisions, overdue_payments) -> dict:
    """§2 AI Erken Uyarı panel. text = top assurance finding or top decision; the
    footer carries ONLY honest tokens — 'İnsan onayı gerekir' plus an 'Eylemsizlik
    riski' phrase when derivable. NO fabricated Güven %/Veri kapsama % (§0)."""
    if high_findings:
        f = high_findings[0]
        text = f.get("reasoning") or f.get("body_tr") or f.get("title_tr")
    elif decisions and decisions[0]["oncelik_rag"] != "g":
        d = decisions[0]
        text = f"En yüksek öncelikli konu: {d['konu']}. {d['gerekce']}."
    else:
        text = "Bu dönem için acil bir erken uyarı bulgusu tespit edilmemiştir."

    tokens = ["İnsan onayı gerekir"]
    if overdue_payments:
        tokens.append(f"Eylemsizlik riski: {overdue_payments[0]['days']} gün geciken ödeme")
    elif high_findings:
        tokens.append("Eylemsizlik riski: incelenmeyen yüksek öncelikli bulgu")
    return {"text": text, "footer": "   ·   ".join(tokens)}


def _margin_bridge(cat_totals, P) -> dict:
    """§3 Marj Köprüsü from REAL per-category variances only (no narrative rows).
    Opening = bütçe marjı (sözleşme − revize bütçe); rows = top categories by
    |Σ variance|; closing = tahmini final marj (sözleşme − tahmini final maliyet)."""
    opening_pct = _safe_pct_str(P["contract"] - P["revised_budget"], P["contract"])
    closing_pct = _safe_pct_str(P["contract"] - P["forecast"], P["contract"])
    ranked = sorted(
        ((COST_CATEGORIES.get(k, k), t) for k, t in cat_totals.items()
         if D(t["variance"]) != 0),
        key=lambda kt: abs(D(kt[1]["variance"])), reverse=True)
    rows = []
    for label, t in ranked[:6]:
        var = D(t["variance"])
        rows.append({
            "kalem": label,
            "etki": format_currency_tr(-var),  # over budget => negative margin impact
            "not": "Bütçe aşımı" if var > 0 else "Tasarruf",
        })
    return {
        "opening_pct": opening_pct,
        "closing_pct": closing_pct,
        "rows": rows,
    }


def _commitment_categories(cat_totals) -> tuple[list[dict], list[dict]]:
    """§4 category detail table + stacked-bar dataset (gerçekleşen vs açık taahhüt),
    aggregated across projects. Sorted by % usage descending."""
    table: list[dict] = []
    for key, t in cat_totals.items():
        revised, invoiced, open_c = D(t["revised"]), D(t["invoiced"]), D(t["open_committed"])
        if revised <= 0 and invoiced <= 0 and open_c <= 0:
            continue
        pct_value = float(invoiced / revised * 100) if revised > 0 else 0.0
        durum = "r" if pct_value > 100 else ("a" if pct_value >= 85 else "g")
        table.append({
            "label": COST_CATEGORIES.get(key, key),
            "revised": revised, "invoiced": invoiced, "open_committed": open_c,
            "remaining": revised - D(t["exposure"]),
            "pct_value": round(pct_value, 1),
            "durum_rag": durum,
        })
    table.sort(key=lambda x: x["pct_value"], reverse=True)
    chart = [{"label": _short(r["label"], 14),
              "actual": float(r["invoiced"]), "open": float(r["open_committed"])}
             for r in table[:8]]
    return table, chart


def _fx_cost_aggregates(db: Session, company: Company) -> dict:
    """Company-scoped cost FX aggregates (mirror financials.py scope): Σ amount_try,
    Σ amount_try over rows WITH a USD snapshot, Σ amount_usd. Excludes soft-deleted /
    pending-approval / forecast rows."""
    from sqlalchemy import case, func, select

    from app.models.cost_entry import CostEntry

    total_try, fx_try, usd = db.execute(
        select(
            func.coalesce(func.sum(CostEntry.amount_try), 0),
            func.coalesce(func.sum(
                case((CostEntry.amount_usd.is_not(None), CostEntry.amount_try), else_=0)), 0),
            func.coalesce(func.sum(CostEntry.amount_usd), 0),
        ).where(
            CostEntry.company_id == company.id,
            CostEntry.is_deleted.is_(False),
            CostEntry.pending_approval.is_(False),
            CostEntry.entry_type != "forecast",
        )
    ).one()
    return {"total_try": D(total_try), "fx_try": D(fx_try), "usd": D(usd)}


def _fx_section(db: Session, company: Company, projects, fx_effect_total) -> dict:
    """§5 Kur & Döviz. ``has_fx`` is False (render omits the section) when no cost
    is tracked in USD. ``fx_sensitive_pct`` = share of cost (TRY value) that carries
    a USD snapshot; ``avg_rate`` = Σ TRY ÷ Σ USD over those rows. No fabricated
    numbers — the Yönetim Notu is a static management template."""
    from app.services import fx as fx_service

    agg = _fx_cost_aggregates(db, company)
    usd = agg["usd"]
    has_fx = usd > 0
    fx_sensitive_pct = float(agg["fx_try"] / agg["total_try"] * 100) if agg["total_try"] > 0 else 0.0
    fx_sensitive_pct = max(0.0, min(100.0, fx_sensitive_pct))
    avg_rate = float(agg["fx_try"] / usd) if usd > 0 else None
    today_rate = fx_service.latest_rate(db)

    fx_kpis = [
        ("DÖVİZE DUYARLI MALİYET", format_pct_tr(fx_sensitive_pct),
         "USD karşılığı izlenen", H_AMBER, H_GOLD),
        ("KUR ETKİSİ", format_currency_tr(fx_effect_total),
         "Maliyete yansıyan", _good_bad(fx_effect_total <= 0), H_GOLD),
        ("USD MALİYET", f"${format_number_tr(usd, 0)}", "Snapshot kurdan", H_MUT, H_PETROL),
        ("ORT. ALIM KURU", f"{format_number_tr(avg_rate, 2)} ₺/$" if avg_rate else "—",
         "Dönem ağırlıklı", H_MUT, H_PETROL),
    ]
    fx_note = (
        "Döviz maruziyeti, USD karşılığı izlenen maliyet kalemlerine dayanır. Kur "
        "hareketleri brüt marjı doğrudan etkiler. İthal ekipman ve dövizli "
        "sözleşmelerde vadeli kur koruması ve sözleşmelerde kur eskalasyon maddesi "
        "değerlendirilmesi önerilir."
    )
    return {
        "has_fx": has_fx,
        "fx_kpis": fx_kpis,
        "fx_split": {"try_pct": round(100.0 - fx_sensitive_pct, 1),
                     "fx_sensitive_pct": round(fx_sensitive_pct, 1)},
        "fx_note": fx_note,
        "today_rate": str(today_rate) if today_rate is not None else None,
    }


def _ar_aging_for_company(db: Session, project_ids, today) -> dict:
    """§7 AR aging — REPLICATES app/api/projects.py:901 ``_ar_aging`` exactly: same
    buckets (not_due / d1_30 / d31_60 / d60_plus) and the same outstanding-weighted
    DSO over the active project ids. (Kept here rather than imported from the API
    layer; due_date/invoice_date None-guards added so a report never crashes.)"""
    from sqlalchemy import select

    from app.calculations.money import money
    from app.models.client_invoice import ClientInvoice

    zero = {"not_due_try": "0", "d1_30_try": "0", "d31_60_try": "0",
            "d60_plus_try": "0", "total_outstanding_try": "0", "dso_days": None}
    if not project_ids:
        return zero
    invs = db.execute(
        select(ClientInvoice).where(
            ClientInvoice.project_id.in_(project_ids),
            ClientInvoice.is_deleted.is_(False),
        )
    ).scalars().all()
    b = {"not_due": D(0), "d1_30": D(0), "d31_60": D(0), "d60_plus": D(0)}
    total = D(0)
    weighted_age = D(0)
    for i in invs:
        out = D(i.outstanding_try)
        if out <= 0:
            continue
        total += out
        if i.due_date is not None:
            overdue = (today - i.due_date).days
            if overdue <= 0:
                b["not_due"] += out
            elif overdue <= 30:
                b["d1_30"] += out
            elif overdue <= 60:
                b["d31_60"] += out
            else:
                b["d60_plus"] += out
        else:
            b["not_due"] += out
        if i.invoice_date is not None:
            weighted_age += out * D((today - i.invoice_date).days)
    dso = int(round(float(weighted_age / total))) if total > 0 else None
    return {
        "not_due_try": str(money(b["not_due"])),
        "d1_30_try": str(money(b["d1_30"])),
        "d31_60_try": str(money(b["d31_60"])),
        "d60_plus_try": str(money(b["d60_plus"])),
        "total_outstanding_try": str(money(total)),
        "dso_days": dso,
    }


def _vendor_spend(db: Session, company: Company, top_n: int = 8) -> dict:
    """§8 vendor spend — NEW aggregation grouping cost entries by ``vendor_id``
    (company-scoped; is_deleted=False AND pending_approval=False AND entry_type !=
    'forecast', mirroring financials.py). vendor_id IS NULL is EXCLUDED from the
    ranking. Returns top-N + first-5 concentration % over total vendor spend."""
    from sqlalchemy import func, select

    from app.models.cost_entry import CostEntry
    from app.models.vendor import Vendor

    grouped = db.execute(
        select(CostEntry.vendor_id, func.coalesce(func.sum(CostEntry.total_with_vat_try), 0))
        .where(
            CostEntry.company_id == company.id,
            CostEntry.is_deleted.is_(False),
            CostEntry.pending_approval.is_(False),
            CostEntry.entry_type != "forecast",
            CostEntry.vendor_id.is_not(None),
        )
        .group_by(CostEntry.vendor_id)
    ).all()
    names = {v.id: (v.canonical_name or "Bilinmeyen Tedarikçi")
             for v in db.execute(
                 select(Vendor).where(Vendor.company_id == company.id)).scalars().all()}
    spend = [{"name": names.get(vid, "Bilinmeyen Tedarikçi"),
              "amount": format_currency_tr(D(amt)), "amount_d": D(amt)}
             for vid, amt in grouped]
    spend.sort(key=lambda x: x["amount_d"], reverse=True)
    total = sum((s["amount_d"] for s in spend), D(0))
    first5 = sum((s["amount_d"] for s in spend[:5]), D(0))
    concentration = round(float(first5 / total * 100), 1) if total > 0 else 0.0
    return {"top": spend[:top_n], "active_count": len(spend),
            "concentration_pct": concentration}


def _assurance_section(db: Session, company: Company, today) -> dict:
    """§10 Veri Güvence & Anomali — READ-ONLY: uses ``collect_findings`` +
    ``scanned_counts`` (what POST /ai/assurance/scan reads), NEVER ``scan_company``
    (which writes AIAlert rows). Groups findings by type and surfaces high-severity
    priority findings. ``high_findings`` is returned for §2/§11 reuse and popped by
    the caller before the dict is serialised into the report payload."""
    from app.services import assurance

    counts = assurance.scanned_counts(db, company.id)
    findings = assurance.collect_findings(db, company.id, today)
    scanned_total = counts["cost_entries"] + counts["client_invoices"]

    by_type: dict[str, int] = {}
    for f in findings:
        by_type[f["alert_type"]] = by_type.get(f["alert_type"], 0) + 1
    found_by_type = [
        {"type": t, "label": _ALERT_TYPE_LABELS.get(t, t),
         "count": n, "rag": ("r" if t in ("duplicate_cost", "duplicate_invoice",
                                           "hakedis_over_contract") else "a")}
        for t, n in sorted(by_type.items(), key=lambda x: -x[1])
    ]
    high = [f for f in findings if f["severity"] == "high"]
    priority = [{
        "bulgu": f["title_tr"],
        "tur": _ALERT_TYPE_LABELS.get(f["alert_type"], f["alert_type"]),
        "oneri": f.get("recommended_action") or "—",
        "durum_label": _SEVERITY_LABELS.get(f["severity"], f["severity"]),
        "durum_rag": _SEVERITY_RAG.get(f["severity"], "a"),
    } for f in (high or findings)[:8]]

    return {
        "scanned_total": scanned_total,
        "total_found": len(findings),
        "high_count": len(high),
        "found_by_type": found_by_type,
        "priority_findings": priority,
        "high_findings": high,
    }


def _sell_side_section(db: Session, projects) -> dict | None:
    """§9 Satış, m² & Getiri — ONLY for sell-side projects (revenue_model ∈
    SELL_SIDE_REVENUE_MODELS). Aggregates unit-sales P&L, per-m² economics and
    daire-tipi P&L from sales.unit_sales_pnl / project_pnl / investment_return.
    All str(Decimal) values converted; None guarded. Returns None when no sales."""
    from app.constants import SELL_SIDE_REVENUE_MODELS, UNIT_TYPES
    from app.services import sales as sales_service

    ss = [p for p in projects if p.revenue_model in SELL_SIDE_REVENUE_MODELS]
    if not ss:
        return None

    total_sold = 0
    total_units = 0
    sum_sales = D(0)
    sum_m2 = D(0)
    sum_profit = D(0)
    sum_cost = D(0)
    m2_economics: list[dict] = []
    unit_type_agg: dict[str, dict] = {}

    for p in ss:
        usp = sales_service.unit_sales_pnl(db, p)
        totals = usp["totals"]
        pnl = sales_service.project_pnl(db, p)
        total_sold += int(totals["count"])
        if p.unit_count:
            total_units += int(p.unit_count)
        sum_sales += D(totals["sale_price_try"])
        sum_m2 += D(totals["total_m2"])
        sum_profit += D(totals["pnl_try"])
        sum_cost += D(totals["cost_try"])

        m2a = pnl.get("m2_analysis") or {}
        cost_per_m2 = (m2a.get("per_net_m2") or {}).get("try") or (m2a.get("per_gross_m2") or {}).get("try")
        rev_per_m2 = totals.get("avg_price_per_m2_try")
        if cost_per_m2 is not None and rev_per_m2 is not None:
            cpm, rpm = float(cost_per_m2), float(rev_per_m2)
            m2_economics.append({
                "project": _short(p.name, 16),
                "cost_per_m2": cpm, "revenue_per_m2": rpm, "profit_per_m2": rpm - cpm})

        for a in usp["allocations"]:
            ut = a.get("unit_type") or "other"
            agg = unit_type_agg.setdefault(ut, {"count": 0, "sales": D(0), "cost": D(0), "pnl": D(0)})
            agg["count"] += 1
            agg["sales"] += D(a.get("sale_price_try") or 0)
            if a.get("unit_cost_try") is not None:
                agg["cost"] += D(a["unit_cost_try"])
            if a.get("pnl_try") is not None:
                agg["pnl"] += D(a["pnl_try"])

    roi = float(sum_profit / sum_cost * 100) if sum_cost > 0 else None
    avg_price_m2 = (sum_sales / sum_m2) if sum_m2 > 0 else None
    margin_pct = float(sum_profit / sum_sales * 100) if sum_sales > 0 else None

    # IRR is not additive across projects — show it only for a single sell-side
    # project (honest), else "—".
    irr = None
    if len(ss) == 1:
        irr = sales_service.investment_return(db, ss[0]).get("irr_try_pct")
    irr_str = f"%{format_number_tr(D(irr), 1)}" if irr is not None else "—"
    roi_str = format_pct_tr(roi) if roi is not None else "—"

    sales_kpis = [
        ("SATILAN / TOPLAM", f"{total_sold} / {total_units}" if total_units else str(total_sold),
         (f"{round(total_sold / total_units * 100)}% satış" if total_units else "Birim"), H_GREEN, H_PETROL),
        ("ORT. SATIŞ ₺/m²", format_currency_tr(avg_price_m2) if avg_price_m2 is not None else "—",
         "Satılan birimler", H_MUT, H_PETROL),
        ("BRÜT KÂR (SATIŞ)", format_currency_tr(sum_profit),
         f"Marj {format_pct_tr(margin_pct)}" if margin_pct is not None else "—",
         _good_bad(sum_profit >= 0), H_PETROL),
        ("IRR / ROI", f"{irr_str} / {roi_str}", "Yıllık / kümülatif",
         _good_bad((roi or 0) >= 0), H_PETROL),
    ]

    unit_type_pnl = []
    for ut, agg in sorted(unit_type_agg.items(), key=lambda x: -x[1]["sales"]):
        if agg["count"] == 0:
            continue
        avg_price = agg["sales"] / agg["count"]
        ut_margin = float(agg["pnl"] / agg["sales"] * 100) if agg["sales"] > 0 else None
        unit_type_pnl.append({
            "tip": UNIT_TYPES.get(ut, ut),
            "satilan": str(agg["count"]),
            "ort_fiyat": format_currency_tr(avg_price),
            "maliyet": format_currency_tr(agg["cost"]),
            "brut_kar": format_currency_tr(agg["pnl"]),
            "marj": format_pct_tr(ut_margin) if ut_margin is not None else "—",
        })

    return {
        "sales_kpis": sales_kpis,
        "m2_economics": m2_economics,
        "unit_type_pnl": unit_type_pnl,
    }


def _risk_and_actions(decisions) -> tuple[list[dict], list[dict]]:
    """§11 risk register + action plan synthesised from the §2 decisions (which
    already fold in overdue payments and high-severity findings)."""
    risk_register = []
    action_plan = []
    for d in decisions:
        if d["oncelik_rag"] == "g":
            continue
        risk_register.append({
            "risk": d["konu"],
            "duzey_label": _SEVERITY_LABELS["high"] if d["oncelik_rag"] == "r" else _SEVERITY_LABELS["medium"],
            "duzey_rag": d["oncelik_rag"],
            "etki": d["gerekce"],
            "azaltma": d["beklenen_etki"],
            "sahip": d["sahip"],
        })
        action_plan.append({
            "aksiyon": d["konu"],
            "sahip": d["sahip"],
            "beklenen_etki": d["beklenen_etki"],
            "durum_label": "Açık" if d["oncelik_rag"] == "r" else "İzle",
            "durum_rag": d["oncelik_rag"],
        })
    return risk_register[:6], action_plan[:6]


def _active_projects(db: Session, company: Company):
    from sqlalchemy import select

    return db.execute(
        select(Project).where(
            Project.company_id == company.id,
            Project.is_deleted.is_(False),
            Project.status == "active",
        )
    ).scalars().all()


def _company_overdue_payments(db: Session, company: Company, today) -> list[dict]:
    """§6/§7: unpaid cost entries whose payment_due_date is in the past."""
    from sqlalchemy import select

    from app.models.cost_entry import CostEntry

    project_names = {
        p.id: p.name for p in _active_projects(db, company)
    }
    rows = db.execute(
        select(CostEntry).where(
            CostEntry.company_id == company.id,
            CostEntry.is_deleted.is_(False),
            CostEntry.pending_approval.is_(False),
            CostEntry.entry_type != "forecast",
            CostEntry.payment_status != "paid",
            CostEntry.payment_due_date.is_not(None),
            CostEntry.payment_due_date < today,
        )
    ).scalars().all()
    out = []
    for c in rows:
        remaining = D(c.total_with_vat_try) - D(c.amount_paid_try)
        if remaining <= 0:
            continue
        days = (today - c.payment_due_date).days
        out.append({
            "supplier": c.supplier_name or "Bilinmeyen Tedarikçi",
            "project": project_names.get(c.project_id, ""),
            "amount": format_currency_tr(remaining),
            "amount_d": remaining,
            "due_date": format_date_tr(c.payment_due_date),
            "days": days,
            "severe": days >= 30,
        })
    out.sort(key=lambda r: r["days"], reverse=True)
    return out


def _subcontractor_commitments(db: Session, company: Company) -> list[dict]:
    """§6: top 5 active subcontractor commitments by revised contract value."""
    from sqlalchemy import select

    from app.calculations.subcontractor import subcontractor_revised_contract
    from app.models.cost_entry import CostEntry
    from app.models.subcontractor import Subcontractor

    subs = db.execute(
        select(Subcontractor).where(
            Subcontractor.company_id == company.id,
            Subcontractor.is_deleted.is_(False),
            Subcontractor.status == "active",
        )
    ).scalars().all()
    if not subs:
        return []
    # Paid-to-date per subcontractor (linked cost entries).
    paid_by_sub: dict = {}
    for c in db.execute(
        select(CostEntry).where(
            CostEntry.company_id == company.id,
            CostEntry.is_deleted.is_(False),
            CostEntry.subcontractor_id.is_not(None),
        )
    ).scalars().all():
        paid_by_sub[c.subcontractor_id] = D(paid_by_sub.get(c.subcontractor_id, D(0))) + D(c.amount_paid_try)

    items = []
    for s in subs:
        revised = subcontractor_revised_contract(s.contract_value_try, s.approved_variations_try)
        paid = D(paid_by_sub.get(s.id, D(0)))
        remaining = revised - paid
        pct_done = float(paid / revised * 100) if revised > 0 else 0.0
        items.append({
            "name": s.name,
            "scope": s.scope_of_work or "—",
            "contract": format_currency_tr(revised),
            "paid": format_currency_tr(paid),
            "remaining": format_currency_tr(remaining),
            "pct_done": format_pct_tr(pct_done),
            "revised_d": revised,
        })
    items.sort(key=lambda x: D(x["revised_d"]), reverse=True)
    return items[:5]


def _strip_action_lines(text) -> list[str]:
    """Turn an AI action blob into clean plain-text lines (no markdown/numbering)."""
    import re

    lines = []
    for raw in str(text or "").splitlines():
        s = raw.strip()
        if not s:
            continue
        s = re.sub(r"^\s*(\d+[\.\)]|[-*•])\s*", "", s)   # leading numbering / bullets
        s = s.replace("**", "").replace("##", "").replace("`", "").strip()
        if s:
            lines.append(s)
    return lines


def _parse_period_anchor(period_label: str):
    """Best-effort anchor date from a period string ("2026-06" -> 2026-06-15).

    Falls back to today when the label is free-form ("Bu Ay", "Haziran 2026").
    """
    from datetime import date

    try:
        parts = str(period_label).split("-")
        if len(parts) == 2:
            return date(int(parts[0]), int(parts[1]), 15)
    except (ValueError, TypeError):
        pass
    return date.today()


def render_management_pack(db: Session, company: Company, period_label: str) -> bytes:
    return _management_pack_pdf(build_management_pack_data(db, company, period_label))


# ---------------------------------------------------------------------------
# ReportLab rendering helpers (imported lazily; stubbed in tests)
# ---------------------------------------------------------------------------
def _styles():
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet

    register_turkish_fonts()
    base = getSampleStyleSheet()
    s = {
        "company": ParagraphStyle(
            "company", parent=base["Normal"], fontName=FONT_BOLD,
            fontSize=15, textColor=PRIMARY,
        ),
        "title": ParagraphStyle(
            "title", parent=base["Normal"], fontName=FONT_BOLD,
            fontSize=17, textColor=PRIMARY, alignment=TA_RIGHT,
        ),
        "meta": ParagraphStyle(
            "meta", parent=base["Normal"], fontName=FONT_NORMAL, fontSize=9, textColor=MUTED, alignment=TA_RIGHT,
        ),
        "h2": ParagraphStyle(
            "h2", parent=base["Normal"], fontName=FONT_BOLD,
            fontSize=14, textColor=PRIMARY, spaceBefore=14, spaceAfter=6,
        ),
        "h3": ParagraphStyle(
            "h3", parent=base["Normal"], fontName=FONT_BOLD,
            fontSize=11, textColor=PRIMARY, spaceBefore=8, spaceAfter=4,
        ),
        "body": ParagraphStyle(
            "body", parent=base["Normal"], fontName=FONT_NORMAL,
            fontSize=9, textColor="#1E293B", leading=14,
        ),
        "note": ParagraphStyle(
            "note", parent=base["Normal"], fontName=FONT_OBLIQUE,
            fontSize=8, textColor=MUTED, leading=12, spaceBefore=4, spaceAfter=4,
        ),
        # CR-006-A cover-page styles.
        "band_brand": ParagraphStyle(
            "band_brand", parent=base["Normal"], fontName=FONT_BOLD, fontSize=22, textColor="#FFFFFF",
        ),
        "band_title": ParagraphStyle(
            "band_title", parent=base["Normal"], fontName=FONT_BOLD, fontSize=16, textColor="#FFFFFF",
            alignment=TA_RIGHT,
        ),
        "cover_period": ParagraphStyle(
            "cover_period", parent=base["Normal"], fontName=FONT_BOLD, fontSize=28, textColor=NAVY,
        ),
        "cover_company": ParagraphStyle(
            "cover_company", parent=base["Normal"], fontName=FONT_NORMAL, fontSize=18, textColor=PRIMARY,
        ),
        "cover_meta": ParagraphStyle(
            "cover_meta", parent=base["Normal"], fontName=FONT_NORMAL, fontSize=10, textColor=MUTED,
        ),
        "cover_kpi_label": ParagraphStyle(
            "cover_kpi_label", parent=base["Normal"], fontName=FONT_NORMAL, fontSize=8, textColor=MUTED,
        ),
        "cover_kpi_value": ParagraphStyle(
            "cover_kpi_value", parent=base["Normal"], fontName=FONT_BOLD, fontSize=12, textColor=PRIMARY,
        ),
        "ai": ParagraphStyle(
            "ai", parent=base["Normal"], fontName=FONT_NORMAL, fontSize=10, textColor="#1E293B",
            leading=15, backColor="#EFF6FF", borderColor=PRIMARY, borderWidth=0,
            leftIndent=8, rightIndent=8, spaceBefore=4, spaceAfter=4,
        ),
        "disclaimer": ParagraphStyle(
            "disclaimer", parent=base["Normal"], fontName=FONT_OBLIQUE,
            fontSize=8, textColor="#94A3B8", alignment=TA_LEFT, spaceBefore=8,
        ),
        "kpi_label": ParagraphStyle(
            "kpi_label", parent=base["Normal"], fontName=FONT_NORMAL, fontSize=8, textColor=MUTED,
        ),
        "kpi_value": ParagraphStyle(
            "kpi_value", parent=base["Normal"], fontName=FONT_BOLD, fontSize=13, textColor=PRIMARY,
        ),
    }
    return s


_LOGO_MAX_BYTES = 5_000_000  # cap a remote logo download (memory-pressure guard)


def _logo_host_is_safe(logo_url: str) -> bool:
    """SSRF guard for the company-controlled logo URL: allow only http/https to a
    PUBLIC host. Reject if the host resolves to any private / loopback / link-local
    / reserved / multicast / unspecified address (e.g. 169.254.169.254 cloud
    metadata, 127.0.0.1, 10./172.16./192.168.). The logo is fetched server-side, so
    an internal URL must never be reachable. Fail-closed on any parse/DNS error."""
    import ipaddress
    import socket
    from urllib.parse import urlparse

    try:
        u = urlparse(logo_url)
        if u.scheme not in ("http", "https") or not u.hostname:
            return False
        port = u.port or (443 if u.scheme == "https" else 80)
        for *_, sockaddr in socket.getaddrinfo(u.hostname, port, proto=socket.IPPROTO_TCP):
            ip = ipaddress.ip_address(sockaddr[0])
            if (ip.is_private or ip.is_loopback or ip.is_link_local
                    or ip.is_reserved or ip.is_multicast or ip.is_unspecified):
                return False
        return True
    except Exception:  # noqa: BLE001 — any resolution failure → block the fetch
        return False


def _logo_flowable(logo_url, max_h=44.0, max_w=120.0):
    """Return a ReportLab Image for the logo, or None if it cannot be loaded.

    CR-006-D: remote (http/https) logos are downloaded into memory with a short
    timeout so the company logo embeds in the PDF. Any failure (no URL, network
    error, bad image) falls back silently to the 'YAPI' text — the logo is
    optional and must never break the whole report.
    """
    if not logo_url:
        return None
    try:
        import io

        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import Image

        source = logo_url
        if isinstance(logo_url, str) and logo_url.lower().startswith(("http://", "https://")):
            import httpx

            if not _logo_host_is_safe(logo_url):
                return None
            resp = httpx.get(logo_url, timeout=5)
            resp.raise_for_status()
            if len(resp.content) > _LOGO_MAX_BYTES:
                return None
            source = io.BytesIO(resp.content)

        reader = ImageReader(source)
        iw, ih = reader.getSize()
        # Scale to fit within max_w x max_h, preserving aspect ratio.
        ratio = (iw / ih) if ih else 1.0
        h = max_h
        w = h * ratio
        if w > max_w:
            w = max_w
            h = w / ratio if ratio else max_h
        if isinstance(source, io.BytesIO):
            source.seek(0)
        return Image(source, width=w, height=h)
    except Exception:
        # Logo is optional — never fail the whole report over it.
        return None


def _header_table(styles, logo_url, company_name, title, subtitle):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Table, TableStyle

    logo = _logo_flowable(logo_url)
    left = [logo] if logo else []
    left.append(Paragraph(company_name, styles["company"]))
    right = [Paragraph(title, styles["title"]), Paragraph(subtitle, styles["meta"])]
    t = Table([[left, right]], colWidths=[10 * cm, 7.5 * cm])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, -1), 2, colors.HexColor(PRIMARY)),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    return t


def _data_table(rows, col_widths, num_cols=(), header=True, rag_col=None, rag_values=None,
                bg_cells=None, row_bg=None):
    """Build a styled Platypus table. ``rows`` includes the header row when header=True.

    ``bg_cells`` colour-codes individual cells: a list of (col, data_row_idx, rag).
    ``row_bg`` colour-codes whole data rows: a list of (data_row_idx, rag). Both use
    the light RAG_TINTS palette so text stays readable. Padding is 6pt (CR-006-A).
    """
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    t = Table(rows, colWidths=col_widths, repeatRows=1 if header else 0)
    style = [
        ("FONTNAME", (0, 0), (-1, -1), FONT_NORMAL),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 1 if header else 0), (-1, -1), colors.HexColor("#1E293B")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor(BORDER)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    if header:
        style += [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PRIMARY)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
        ]
    for c in num_cols:
        style.append(("ALIGN", (c, 0), (c, -1), "RIGHT"))
    off = 1 if header else 0
    if row_bg:
        for i, rag in row_bg:
            r = i + off
            style.append(("BACKGROUND", (0, r), (-1, r), colors.HexColor(RAG_TINTS.get(rag, "#FFFFFF"))))
    if bg_cells:
        for col, i, rag in bg_cells:
            r = i + off
            style.append(("BACKGROUND", (col, r), (col, r), colors.HexColor(RAG_TINTS.get(rag, "#FFFFFF"))))
            style.append(("FONTNAME", (col, r), (col, r), FONT_BOLD))
    if rag_col is not None and rag_values:
        for i, rag in enumerate(rag_values):
            r = i + off
            style.append(("TEXTCOLOR", (rag_col, r), (rag_col, r), colors.HexColor(RAG_COLORS.get(rag, MUTED))))
            style.append(("FONTNAME", (rag_col, r), (rag_col, r), FONT_BOLD))
    t.setStyle(TableStyle(style))
    return t


def _footer_painter(generated_at, note=None):
    """Return an onPage callback drawing page numbers + generated-by (+ optional note)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4

    def paint(canvas, doc):
        register_turkish_fonts()
        canvas.saveState()
        width, _ = A4
        canvas.setFont(FONT_NORMAL, 8)
        canvas.setFillColor(colors.HexColor(MUTED))
        canvas.drawCentredString(width / 2, 1.0 * 28.35, f"Sayfa {doc.page}")
        canvas.setFont(FONT_NORMAL, 7)
        canvas.setFillColor(colors.HexColor("#94A3B8"))
        canvas.drawString(1.4 * 28.35, 1.0 * 28.35, f"Yapı tarafından {generated_at} tarihinde oluşturuldu")
        if note:
            canvas.setFont(FONT_OBLIQUE, 6)
            canvas.drawString(1.4 * 28.35, 0.6 * 28.35, note)
        canvas.restoreState()

    return paint


def _render_story(story, generated_at, footer_note=None) -> bytes:
    from io import BytesIO

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=1.6 * cm, bottomMargin=2 * cm, leftMargin=1.4 * cm, rightMargin=1.4 * cm,
        title="Yapı Rapor",
    )
    paint = _footer_painter(generated_at, footer_note)
    doc.build(story, onFirstPage=paint, onLaterPages=paint)
    return buf.getvalue()


def _project_report_pdf(d: dict) -> bytes:
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer

    s = _styles()
    story = [
        _header_table(s, d["logo_url"], d["company_name"], d["report_title"], d["report_date"]),
        Spacer(1, 10),
        Paragraph(f"{d['project_name']} — {d['client_name']}", s["h2"]),
    ]

    # KPI strip as a 4-column table.
    def kpi(label, value):
        return [Paragraph(label, s["kpi_label"]), Paragraph(value, s["kpi_value"])]

    kpi_row = [
        kpi("Sözleşme Değeri", d["contract_value"]),
        kpi("Gerçekleşen Maliyet", d["total_actual"]),
        kpi("Final Tahmin", d["forecast_final"]),
        kpi("Kar Marjı", d["margin_pct"]),
    ]
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    kt = Table([kpi_row], colWidths=[4.375 * cm] * 4)
    kt.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(BORDER)),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor(BORDER)),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story += [kt, Spacer(1, 6)]

    # Budget vs actual by category.
    story.append(Paragraph("Bütçe & Gerçekleşen (Kategori Bazında)", s["h2"]))
    rows = [["Kategori", "Revize Bütçe", "Faturalanan", "Final Tahmin", "Sapma", "Durum"]]
    rags = []
    for c in d["categories"]:
        rows.append([c["label"], c["revised"], c["invoiced"], c["forecast"], c["variance"], c["status_label"]])
        rags.append(c["status"])
    story.append(_data_table(
        rows, col_widths=[5 * cm, 2.7 * cm, 2.7 * cm, 2.7 * cm, 2.4 * cm, 2.1 * cm],
        num_cols=(1, 2, 3, 4), rag_col=5, rag_values=rags,
    ))

    # Income & collection summary.
    story.append(Paragraph("Gelir & Tahsilat Özeti", s["h2"]))
    summary = [
        ["İşverene Faturalanan", d["total_invoiced"]],
        ["Tahsil Edilen", d["total_collected"]],
        ["Bekleyen Tahsilat", d["total_outstanding"]],
        ["Hakediş Kesintisi", d["total_retention"]],
        ["Net Nakit Pozisyonu", d["net_cash"]],
    ]
    story.append(_data_table(summary, col_widths=[12 * cm, 5.6 * cm], num_cols=(1,), header=False))

    return _render_story(story, d["generated_at"])


_TR_MONTHS_SHORT = [
    "", "Oca", "Şub", "Mar", "Nis", "May", "Haz",
    "Tem", "Ağu", "Eyl", "Eki", "Kas", "Ara",
]


def _month_label_tr(month_key: str) -> str:
    """'2026-06' -> 'Haz 26' for compact chart axis labels."""
    try:
        y, m = month_key.split("-")
        return f"{_TR_MONTHS_SHORT[int(m)]} {y[2:]}"
    except (ValueError, IndexError):
        return month_key


def _format_ai_md(text) -> str:
    """Render lightweight Markdown (**bold**, ## başlık) as formatted Paragraph XML."""
    import re
    from xml.sax.saxutils import escape

    out_lines = []
    for raw in str(text or "").splitlines():
        line = escape(raw)
        header = re.match(r"^\s*#{1,6}\s+(.*)$", line)
        if header:
            out_lines.append(f"<b>{header.group(1).strip()}</b>")
            continue
        line = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", line)
        out_lines.append(line)
    return "<br/>".join(out_lines)


def _logo_image_reader(logo_url):
    """Return (ImageReader, width_pt, height_pt) for the cover logo, or None.

    Mirrors ``_logo_flowable`` (remote logos fetched in-memory with a short
    timeout) but yields a canvas-drawable ImageReader. Never raises — a logo is
    optional and must never break the report.
    """
    if not logo_url:
        return None
    try:
        import io

        from reportlab.lib.utils import ImageReader

        source = logo_url
        if isinstance(logo_url, str) and logo_url.lower().startswith(("http://", "https://")):
            import httpx

            if not _logo_host_is_safe(logo_url):
                return None
            resp = httpx.get(logo_url, timeout=5)
            resp.raise_for_status()
            if len(resp.content) > _LOGO_MAX_BYTES:
                return None
            source = io.BytesIO(resp.content)
        reader = ImageReader(source)
        iw, ih = reader.getSize()
        ratio = (iw / ih) if ih else 1.0
        h = 34.0
        w = h * ratio
        if w > 150.0:
            w = 150.0
            h = (w / ratio) if ratio else 34.0
        return reader, w, h
    except Exception:  # noqa: BLE001 — logo is optional, never fail the report
        return None


def _tr_upper(text: str) -> str:
    """Turkish-aware uppercase: lowercase 'i' → dotted 'İ' (and 'ı' → 'I') before
    upper(), so a company name like 'Şirket' renders 'ŞİRKET' on the cover, not the
    Latin-mangled 'ŞIRKET'."""
    return (text or "").replace("ı", "I").replace("i", "İ").upper()


def _mgmt_cover_canvas(d):
    """``onFirstPage`` painter: full-bleed NAVY cover with the gold top rule,
    brand/logo, big title, period and a 4-KPI band (ported from the reference
    ``cover()`` 145-164). Real data — the ÖRNEK chip shows ONLY if d['is_sample'].
    """
    def paint(c, doc):
        from reportlab.lib.colors import HexColor, white
        from reportlab.lib.pagesizes import A4

        from app.services.report_theme import (
            GOLD, NAVY, register_lato_fonts,
            LATO_BLACK, LATO_BOLD, LATO_MEDIUM, LATO_REGULAR, LATO_SEMIBOLD,
        )

        register_lato_fonts()
        Wd, H = A4
        M = 48
        brand = (d.get("company_name") or "YAPI").strip()
        kp = d.get("cover_kpis") or {}

        c.setFillColor(HexColor(NAVY)); c.rect(0, 0, Wd, H, fill=1, stroke=0)
        c.setFillColor(HexColor(GOLD)); c.rect(0, H - 4, Wd, 4, fill=1, stroke=0)

        c.setFillColor(HexColor(GOLD)); c.setFont(LATO_SEMIBOLD, 9)
        c.drawString(M, H - 78, f"AYLIK YÖNETİM RAPORLAMA  ·  {_tr_upper(brand)}")
        logo = _logo_image_reader(d.get("logo_url"))
        if logo:
            reader, lw, lh = logo
            try:
                c.drawImage(reader, Wd - M - lw, H - 88, width=lw, height=lh, mask="auto")
            except Exception:  # noqa: BLE001
                pass

        c.setFillColor(white); c.setFont(LATO_BLACK, 38)
        c.drawString(M, H - 300, "Aylık Yönetim Raporu")
        c.setFillColor(HexColor("#C9D6D1")); c.setFont(LATO_REGULAR, 16)
        c.drawString(M, H - 330, brand)
        c.setFillColor(HexColor(GOLD)); c.rect(M, H - 350, 64, 3, fill=1, stroke=0)
        c.setFillColor(white); c.setFont(LATO_BOLD, 14)
        c.drawString(M, H - 378, str(d.get("period", "")))

        if d.get("is_sample"):
            c.setFillColor(HexColor("#2A4257")); c.roundRect(M, H - 410, 168, 17, 4, fill=1, stroke=0)
            c.setFillColor(HexColor("#D9B36B")); c.setFont(LATO_BOLD, 7.5)
            c.drawString(M + 8, H - 405, "ÖRNEK — KURGUSAL VERİ")

        band = [
            ("AKTİF PROJE", kp.get("active_projects", "0"), ""),
            ("TOPLAM SÖZLEŞME", kp.get("total_contract", "—"), ""),
            ("BEKLEYEN TAHSİLAT", kp.get("total_outstanding", "—"), ""),
            ("PORTFÖY RİSKİ", kp.get("risk_level", "—"), ""),
        ]
        bY = 118
        colW = (Wd - 2 * M) / 4
        for i, (lab, val, ctx) in enumerate(band):
            x = M + i * colW
            if i > 0:
                c.setStrokeColor(HexColor("#2C4358")); c.setLineWidth(0.8)
                c.line(x - 14, bY - 4, x - 14, bY + 62)
            c.setFillColor(HexColor("#8FA0A8")); c.setFont(LATO_SEMIBOLD, 7)
            c.drawString(x, bY + 50, lab)
            c.setFillColor(white); c.setFont(LATO_BOLD, 12)
            c.drawString(x, bY + 26, str(val))
            if ctx:
                c.setFillColor(HexColor("#9FB0B0")); c.setFont(LATO_MEDIUM, 7.5)
                c.drawString(x, bY + 8, ctx)

        c.setStrokeColor(HexColor("#2C4358")); c.setLineWidth(0.6); c.line(M, 74, Wd - M, 74)
        c.setFillColor(HexColor("#7E8E94")); c.setFont(LATO_REGULAR, 7.5)
        c.drawString(M, 60, "GİZLİ  ·  Yönetim kullanımı içindir")
        c.drawRightString(Wd - M, 60, "Sayfa 1")

    return paint


def _mgmt_page_furniture(d, generated_at):
    """``onLaterPages`` painter: light BG, running header (eyebrow left, company ·
    period right), hairlines and footer (generated-by left, page number right).
    Ported from the reference ``later()`` 165-172 — no 'ÖRNEK VERİ' token."""
    def paint(c, doc):
        from reportlab.lib.colors import HexColor
        from reportlab.lib.pagesizes import A4

        from app.services.report_theme import (
            BG, FNT, HAIR, MUT, register_lato_fonts, LATO_REGULAR, LATO_SEMIBOLD,
        )

        register_lato_fonts()
        Wd, H = A4
        M = 48
        c.saveState()
        c.setFillColor(HexColor(BG)); c.rect(0, 0, Wd, H, fill=1, stroke=0)
        c.setFillColor(HexColor(MUT)); c.setFont(LATO_SEMIBOLD, 7)
        c.drawString(M, H - 36, "AYLIK YÖNETİM RAPORU")
        c.setFillColor(HexColor(FNT)); c.setFont(LATO_REGULAR, 8)
        c.drawRightString(Wd - M, H - 36, f"{d.get('company_name', '')}  ·  {d.get('period', '')}")
        c.setStrokeColor(HexColor(HAIR)); c.setLineWidth(0.6)
        c.line(M, H - 44, Wd - M, H - 44); c.line(M, 46, Wd - M, 46)
        c.setFillColor(HexColor(FNT)); c.setFont(LATO_REGULAR, 7.5)
        c.drawString(M, 34, f"Yapı tarafından {generated_at} tarihinde oluşturuldu")
        c.drawRightString(Wd - M, 34, f"Sayfa {doc.page}")
        c.restoreState()

    return paint


def _management_pack_pdf(d: dict) -> bytes:
    """CR-036 render of the "Aylık Yönetim Raporu" using the report_theme/
    report_charts toolkit. Each section is rendered ONLY when its data exists
    (honesty rule §0); §5 omitted when has_fx is False, §9 when has_sell_side is
    False. Charts are PNGs in a tempfile.mkdtemp() cleaned up after doc.build()."""
    import shutil
    import tempfile
    from io import BytesIO

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer

    from app.services import report_charts as ch
    from app.services.report_theme import (
        AMBER, GREEN, INK, MUT, NAVY, PETROL, RED, GOLD,
        LATO_BOLD, LATO_REGULAR, aipanel, chartcard, dtable, kpirow,
        register_lato_fonts, s as TS, sect,
    )

    register_lato_fonts()
    ch.setup_matplotlib_fonts()
    tmpdir = tempfile.mkdtemp(prefix="yapi_mgmt_")

    CW = 15.6 * cm          # chart image width
    CH = 4.9 * cm           # chart image height (full)
    CHs = 4.45 * cm         # chart image height (short / hbar)
    RAG_HEX = {"r": RED, "a": AMBER, "g": GREEN}
    SEV_PILL = {"r": "Yüksek", "a": "Orta", "g": "Düşük"}
    DURUM_PILL = {"r": "Risk", "a": "İzle", "g": "İyi"}
    RAG_KEY = {"green": "g", "amber": "a", "red": "r"}

    def h3(text):
        return Paragraph(text, TS("h3", LATO_BOLD, 10, NAVY, spaceBefore=2))

    def body(text):
        return Paragraph(text, TS("body", LATO_REGULAR, 9, INK, leading=13))

    def calm(msg="Bu dönem için veri bulunmamaktadır."):
        return Paragraph(msg, TS("calm", LATO_REGULAR, 9, MUT))

    def fc(v):
        return format_currency_tr(v)

    try:
        story = [Spacer(1, 2), PageBreak()]

        # ---- §2 Yönetici Özeti --------------------------------------------
        story += sect("KARAR ODAKLI ÖZET", "Yönetici Özeti")
        ek = d.get("exec_kpis") or []
        if ek:
            story += [kpirow(ek[:3], colw=5.4), Spacer(1, 8)]
            if ek[3:]:
                story += [kpirow(ek[3:6], colw=5.4), Spacer(1, 10)]
        story += [h3("Yönetim Görüşü"), Spacer(1, 3),
                  body(_format_ai_md(d.get("ai_summary"))), Spacer(1, 10)]
        ew = d.get("early_warning") or {}
        if ew.get("text"):
            story += [aipanel(ew["text"], ew.get("footer", "İnsan onayı gerekir")), Spacer(1, 12)]
        story += [h3("Karar Gerektiren Konular"), Spacer(1, 4)]
        drows = [[x["konu"], x["gerekce"], x["sahip"], x["beklenen_etki"],
                  (SEV_PILL[x["oncelik_rag"]], x["oncelik_rag"])]
                 for x in (d.get("decisions") or [])]
        story.append(dtable(
            ["Konu", "Gerekçe", "Sahip", "Beklenen Etki", "Öncelik"], drows,
            [4.3 * cm, 4.6 * cm, 2.6 * cm, 4.2 * cm, 1.7 * cm], aligns=[0, 0, 0, 0, 1]))
        story.append(PageBreak())

        # ---- §3 Finansal Performans ---------------------------------------
        story += sect("FİNANSAL PERFORMANS", "Gelir, Maliyet ve Marj")
        fk = d.get("fin_kpis") or []
        if fk:
            story += [kpirow(fk), Spacer(1, 12)]
        mt = d.get("monthly_trend") or []
        # Omit the chart entirely when there is no real movement to plot (e.g. a
        # company with no in-window cashflow) — never a "veri yok" placeholder card.
        if mt and any((m["income"] or m["net"]) for m in mt):
            img = ch.chart_combo(
                [m["label"] for m in mt], [m["income"] for m in mt], [m["net"] for m in mt],
                tmpdir, bar_label="Gelir", line_label="Net Nakit")
            story += [chartcard("Aylık Gelir ve Net Nakit Akışı", img, CW, CH), Spacer(1, 12)]
        mb = d.get("margin_bridge") or {}
        story += [h3("Marj Köprüsü (kategori sapmalarından)"), Spacer(1, 4)]
        mb_rows = [["Bütçe Marjı (açılış)", mb.get("opening_pct", "—"), "Sözleşme − revize bütçe"]]
        for r in mb.get("rows", []):
            mb_rows.append([r["kalem"], r["etki"], r["not"]])
        story.append(dtable(
            ["Kalem", "Etki (₺) / Marj", "Not"], mb_rows,
            [6.5 * cm, 4.5 * cm, 6.4 * cm], aligns=[0, 2, 0],
            totals=["Tahmini Final Marj", mb.get("closing_pct", "—"), "Sözleşme − tahmini final"]))
        story.append(PageBreak())

        # ---- §4 Taahhüt & Maliyet Maruziyeti ------------------------------
        story += sect("MALİYET YAPISI", "Taahhüt & Maliyet Maruziyeti")
        ck = d.get("commitment_kpis") or []
        if ck:
            story += [kpirow(ck), Spacer(1, 12)]
        cc = d.get("commitment_chart") or []
        if cc:
            img = ch.chart_stacked_bar(
                [x["label"] for x in cc], [x["actual"] for x in cc], [x["open"] for x in cc],
                tmpdir, base_label="Gerçekleşen", top_label="Açık Taahhüt")
            story += [chartcard("Kategori Bazında Gerçekleşen vs Açık Taahhüt", img, CW, CH),
                      Spacer(1, 12)]
        cats = d.get("commitment_categories") or []
        story += [h3("Kategori Bütçe Detayı"), Spacer(1, 4)]
        if cats:
            crows = []
            t_rev = t_open = t_inv = t_rem = D(0)
            for c in cats:
                t_rev += D(c["revised"]); t_open += D(c["open_committed"])
                t_inv += D(c["invoiced"]); t_rem += D(c["remaining"])
                crows.append([
                    _short(c["label"], 22), fc(c["revised"]), fc(c["open_committed"]),
                    fc(c["invoiced"]), fc(c["remaining"]), format_pct_tr(c["pct_value"]),
                    (DURUM_PILL[c["durum_rag"]], c["durum_rag"])])
            totals = ["Toplam", fc(t_rev), fc(t_open), fc(t_inv), fc(t_rem),
                      _safe_pct_str(t_inv, t_rev), ""]
            story.append(dtable(
                ["Kategori", "Revize", "Açık Taahhüt", "Faturalanan", "Kalan", "% Kul.", "Durum"],
                crows, [3.5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm, 2.3 * cm, 1.6 * cm, 2.5 * cm],
                aligns=[0, 2, 2, 2, 2, 2, 1], totals=totals))
        else:
            story.append(calm())
        story.append(PageBreak())

        # ---- §5 Kur & Döviz (omitted entirely when has_fx is False) --------
        if d.get("has_fx"):
            story += sect("KUR & DÖVİZ", "Döviz Maruziyeti ve Kur Etkisi")
            fxk = d.get("fx_kpis") or []
            if fxk:
                story += [kpirow(fxk), Spacer(1, 12)]
            split = d.get("fx_split") or {}
            img = ch.chart_donut(
                ["₺ bazlı", "Dövize duyarlı"],
                [split.get("try_pct", 0), split.get("fx_sensitive_pct", 0)],
                tmpdir, colors=[PETROL, GOLD],
                center_top=f"%{format_number_tr(split.get('fx_sensitive_pct', 0), 0)}",
                center_sub="Dövize duyarlı")
            story += [chartcard("Maliyet Para Birimi Dağılımı", img, 9.0 * cm, 5.4 * cm),
                      Spacer(1, 8)]
            story += [h3("Yönetim Notu"), Spacer(1, 3), body(d.get("fx_note", "")), PageBreak()]

        # ---- §6 Kritik Projeler -------------------------------------------
        story += sect("OPERASYON", "Kritik Projeler")
        rows = d.get("rows") or []
        if rows:
            prows = []
            for r in rows:
                rag = RAG_KEY.get(r["rag"], "g")
                prows.append([
                    _short(r["name"], 22), _short(r["project_type"], 16), r["budget"],
                    r["completion"], r["margin"], r["outstanding"],
                    _short(r["commercial_note"], 28), (DURUM_PILL[rag], rag)])
            story.append(dtable(
                ["Proje", "Tip", "Bütçe", "İlerleme", "Marj", "Bekleyen", "Ticari Konu", "Durum"],
                prows, [3.0 * cm, 2.2 * cm, 2.3 * cm, 1.5 * cm, 1.4 * cm, 2.3 * cm, 3.0 * cm, 1.7 * cm],
                aligns=[0, 0, 2, 2, 2, 2, 0, 1]))
        else:
            story.append(calm("Bu dönem için aktif proje bulunmamaktadır."))
        story.append(PageBreak())

        # ---- §7 Nakit & İşletme Sermayesi ---------------------------------
        story += sect("NAKİT", "Nakit & İşletme Sermayesi")
        cak = d.get("cash_kpis") or []
        if cak:
            story += [kpirow(cak), Spacer(1, 12)]
        ar = d.get("ar_aging") or {}
        ar_vals = [float(ar.get("not_due_try", 0)), float(ar.get("d1_30_try", 0)),
                   float(ar.get("d31_60_try", 0)), float(ar.get("d60_plus_try", 0))]
        img = ch.chart_grouped_bar(
            ["Vadesi gelmemiş", "1-30 gün", "31-60 gün", "60+ gün"],
            [("Bekleyen ₺", ar_vals)], tmpdir,
            colors=[GREEN, PETROL, AMBER, RED], value_labels=True)
        story += [chartcard("Alacak Yaşlandırma (₺)", img, CW, CHs), Spacer(1, 10)]
        if mt:
            img2 = ch.chart_line([m["label"] for m in mt], [m["net"] for m in mt],
                                  tmpdir, color=NAVY, fill=True)
            story += [chartcard("Aylık Net Nakit Akışı (₺)", img2, CW, CHs), Spacer(1, 10)]
        story += [h3("Alacak Yaşlandırma Özeti"), Spacer(1, 4)]
        story.append(dtable(
            ["Vade Aralığı", "Bekleyen Tutar"],
            [["Vadesi gelmemiş", fc(ar.get("not_due_try", 0))],
             ["1-30 gün gecikmiş", fc(ar.get("d1_30_try", 0))],
             ["31-60 gün gecikmiş", fc(ar.get("d31_60_try", 0))],
             ["60+ gün gecikmiş", fc(ar.get("d60_plus_try", 0))]],
            [10.0 * cm, 7.4 * cm], aligns=[0, 2],
            totals=["Toplam Bekleyen", fc(ar.get("total_outstanding_try", 0))]))
        story.append(PageBreak())

        # ---- §8 Tedarikçi & Taşeron ---------------------------------------
        story += sect("TEDARİK", "Tedarikçi & Taşeron")
        vk = d.get("vendor_kpis") or []
        if vk:
            story += [kpirow(vk), Spacer(1, 12)]
        vs = d.get("vendor_spend") or []
        if vs:
            img = ch.chart_hbar(
                [_short(v["name"], 22) for v in vs], [float(v["amount_d"]) for v in vs],
                tmpdir, value_fmt="₺{:,.0f}")
            story += [chartcard("En Çok Harcanan Tedarikçiler (₺)", img, CW, CHs), Spacer(1, 12)]
        sc = d.get("subcontractor_commitments") or []
        if sc:
            story += [h3("Taşeron Pozisyonu"), Spacer(1, 4)]
            srows = [[_short(s["name"], 20), _short(s["scope"], 22), s["contract"],
                      s["paid"], s["remaining"], s["pct_done"]] for s in sc]
            story.append(dtable(
                ["Alt Yüklenici", "Kapsam", "Sözleşme", "Ödenen", "Kalan", "% Tamam"],
                srows, [3.4 * cm, 3.6 * cm, 2.8 * cm, 2.6 * cm, 2.6 * cm, 2.4 * cm],
                aligns=[0, 0, 2, 2, 2, 2]))
        if not vs and not sc:
            story.append(calm("Bu dönem için tedarikçi/taşeron verisi bulunmamaktadır."))
        story.append(PageBreak())

        # ---- §9 Satış, m² & Getiri (ONLY for sell-side companies) ----------
        if d.get("has_sell_side") and d.get("sell_side"):
            ss = d["sell_side"]
            story += sect("GELİŞTİRME", "Satış, m² & Getiri")
            sk = ss.get("sales_kpis") or []
            if sk:
                story += [kpirow(sk), Spacer(1, 12)]
            m2 = ss.get("m2_economics") or []
            if m2:
                img = ch.chart_grouped_bar(
                    [x["project"] for x in m2],
                    [("Maliyet/m²", [x["cost_per_m2"] for x in m2]),
                     ("Gelir/m²", [x["revenue_per_m2"] for x in m2]),
                     ("Kâr/m²", [x["profit_per_m2"] for x in m2])],
                    tmpdir, y_label="₺/m²")
                story += [chartcard("Proje Bazında m² Birim Ekonomisi (₺/m²)", img, CW, CH),
                          Spacer(1, 12)]
            ut = ss.get("unit_type_pnl") or []
            story += [h3("Daire Tipine Göre Kâr/Zarar"), Spacer(1, 4)]
            if ut:
                urows = [[u["tip"], u["satilan"], u["ort_fiyat"], u["maliyet"],
                          u["brut_kar"], u["marj"]] for u in ut]
                story.append(dtable(
                    ["Daire Tipi", "Satılan", "Ort. Fiyat", "Maliyet Payı", "Brüt Kâr", "Marj %"],
                    urows, [3.4 * cm, 2.0 * cm, 3.0 * cm, 3.2 * cm, 3.0 * cm, 2.8 * cm],
                    aligns=[0, 2, 2, 2, 2, 2]))
            else:
                story.append(calm("Henüz satış kaydı bulunmamaktadır."))
            story.append(PageBreak())

        # ---- §10 Veri Güvence & Anomali (READ-ONLY) -----------------------
        story += sect("AI GÜVENCE", "Veri Güvence & Anomali")
        a = d.get("assurance") or {}
        high_n = a.get("high_count", 0)
        total_n = a.get("total_found", 0)
        ak = [
            ("TARANAN KAYIT", str(a.get("scanned_total", 0)), "Maliyet + hakediş", MUT, PETROL),
            ("AÇIK BULGU", str(total_n), f"{high_n} yüksek öncelik",
             AMBER if total_n else GREEN, GOLD),
            ("YÜKSEK ÖNCELİK", str(high_n), "İnceleme bekliyor",
             RED if high_n else GREEN, GOLD),
        ]
        story += [kpirow(ak, colw=5.4), Spacer(1, 12)]
        fbt = a.get("found_by_type") or []
        if fbt:
            img = ch.chart_hbar(
                [_short(x["label"], 22) for x in fbt], [x["count"] for x in fbt],
                tmpdir, colors=[RAG_HEX.get(x["rag"], AMBER) for x in fbt], value_fmt="{:.0f}")
            story += [chartcard("Bulgu Türüne Göre (adet)", img, CW, CHs), Spacer(1, 12)]
        pf = a.get("priority_findings") or []
        story += [h3("Öncelikli Bulgular"), Spacer(1, 4)]
        if pf:
            frows = [[_short(x["bulgu"], 40), x["tur"], _short(x["oneri"], 36),
                      (x["durum_label"], x["durum_rag"])] for x in pf]
            story.append(dtable(
                ["Bulgu", "Tür", "Öneri", "Durum"], frows,
                [6.2 * cm, 3.2 * cm, 5.5 * cm, 2.5 * cm], aligns=[0, 0, 0, 1]))
        else:
            story.append(calm("Bu dönem için açık bulgu tespit edilmemiştir."))
        story.append(PageBreak())

        # ---- §11 Risk & Aksiyon Planı -------------------------------------
        story += sect("YÖNETİŞİM", "Risk & Aksiyon Planı")
        rr = d.get("risk_register") or []
        story += [h3("Risk Kaydı"), Spacer(1, 4)]
        if rr:
            rrows = [[_short(x["risk"], 40), (x["duzey_label"], x["duzey_rag"]),
                      _short(x["etki"], 30), _short(x["azaltma"], 30), x["sahip"]] for x in rr]
            story.append(dtable(
                ["Risk", "Düzey", "Etki", "Azaltma", "Sahip"], rrows,
                [5.0 * cm, 1.8 * cm, 3.4 * cm, 4.4 * cm, 2.8 * cm], aligns=[0, 1, 0, 0, 0]))
        else:
            story.append(calm("Bu dönem için kayda değer risk bulunmamaktadır."))
        ap = d.get("action_plan") or []
        story += [Spacer(1, 12), h3("Aksiyon Planı"), Spacer(1, 4)]
        if ap:
            arows = [[_short(x["aksiyon"], 46), x["sahip"], _short(x["beklenen_etki"], 36),
                      (x["durum_label"], x["durum_rag"])] for x in ap]
            story.append(dtable(
                ["Aksiyon", "Sahip", "Beklenen Etki", "Durum"], arows,
                [6.5 * cm, 3.0 * cm, 5.4 * cm, 2.5 * cm], aligns=[0, 0, 0, 1]))
        else:
            story.append(calm("Planlanan aksiyon bulunmamaktadır."))
        story += [Spacer(1, 10),
                  Paragraph(AI_DISCLAIMER, TS("disc", LATO_REGULAR, 7.5, MUT, leading=10))]

        buf = BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4, topMargin=2.0 * cm, bottomMargin=1.7 * cm,
            leftMargin=1.69 * cm, rightMargin=1.69 * cm,
            title="Yapı — Aylık Yönetim Raporu")
        doc.build(story, onFirstPage=_mgmt_cover_canvas(d),
                  onLaterPages=_mgmt_page_furniture(d, d["generated_at"]))
        return buf.getvalue()
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


def _short(text, n=20) -> str:
    """Truncate a label to ``n`` chars for compact axis/cell display."""
    t = str(text or "")
    return t if len(t) <= n else t[: n - 1] + "…"


def _nl2br(text) -> str:
    """Escape text for Paragraph and turn newlines into <br/>."""
    from xml.sax.saxutils import escape

    return escape(str(text or "")).replace("\n", "<br/>")


# ---------------------------------------------------------------------------
# CR-011-C — Agent analysis export (PDF + Excel)
# An agent analysis = answer text + its chart(s) + citations. Self-contained
# (the caller passes the analysis the agent already produced; no re-run).
# ---------------------------------------------------------------------------
def build_agent_analysis_data(company: Company | None, analysis: dict) -> dict:
    """Shape the analysis for rendering (no ReportLab/openpyxl import — testable)."""
    now = datetime.now(timezone.utc)
    return {
        "company_name": company.name if company else "Yapı",
        "logo_url": company.logo_url if company else None,
        "title": analysis.get("title") or "Yapı AI Analizi",
        "question": analysis.get("question"),
        "answer_markdown": analysis.get("answer_markdown") or "",
        "charts": analysis.get("charts") or [],
        "citations": analysis.get("citations") or [],
        "generated_at": format_datetime_tr(now),
        "report_date": format_date_tr(now.date()),
    }


def render_agent_analysis_pdf(company: Company | None, analysis: dict) -> bytes:
    return _agent_analysis_pdf(build_agent_analysis_data(company, analysis))


def _agent_chart_flowables(s, ch: dict) -> list:
    """Render one agent chart as a titled data table (its underlying numbers) —
    honest and dialect/markup-safe for any chart_type (line/bar/composed)."""
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph

    out = [Paragraph(_nl2br(ch.get("title") or "Grafik"), s["h3"])]
    series = ch.get("series") or []
    data = ch.get("data") or []
    x_key = ch.get("x_key") or "x"
    if not series or not data:
        out.append(Paragraph("Grafik verisi yok.", s["body"]))
        return out
    header = [x_key] + [str(se.get("label") or se.get("key") or "") for se in series]
    rows = [header]
    for pt in data:
        row = [str(pt.get(x_key, ""))]
        for se in series:
            row.append(str(pt.get(se.get("key"), "")))
        rows.append(row)
    ncol = len(header)
    out.append(_data_table(
        rows, col_widths=[17.6 / ncol * cm] * ncol, num_cols=tuple(range(1, ncol)),
    ))
    if ch.get("source_note"):
        out.append(Paragraph(_nl2br(ch["source_note"]), s["note"]))
    return out


def _agent_analysis_pdf(d: dict) -> bytes:
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer

    s = _styles()
    story = [
        _header_table(s, d["logo_url"], d["company_name"], d["title"], d["report_date"]),
        Spacer(1, 10),
    ]
    if d.get("question"):
        story.append(Paragraph(f"<b>Soru:</b> {_nl2br(d['question'])}", s["h3"]))
        story.append(Spacer(1, 4))

    story.append(Paragraph("Analiz", s["h2"]))
    story.append(Paragraph(_format_ai_md(d["answer_markdown"]), s["body"]))

    for ch in d["charts"]:
        story.append(Spacer(1, 8))
        story += _agent_chart_flowables(s, ch)

    cits = d["citations"]
    if cits:
        story.append(Spacer(1, 8))
        story.append(Paragraph("Kaynaklar", s["h2"]))
        rows = [["Kaynak", "Bağlantı"]]
        for c in cits:
            rows.append([_short(str(c.get("label", "")), 60), str(c.get("deep_link", ""))])
        story.append(_data_table(rows, col_widths=[8.8 * cm, 8.8 * cm]))

    story.append(Spacer(1, 8))
    story.append(Paragraph(AI_DISCLAIMER, s["disclaimer"]))
    return _render_story(story, d["generated_at"], footer_note=AI_DISCLAIMER)


def _safe_sheet_title(title, idx: int) -> str:
    """Excel sheet titles: <=31 chars, none of []:*?/\\, unique (idx-prefixed)."""
    import re

    t = re.sub(r"[\[\]:\*\?/\\]", " ", str(title or "")).strip()
    return (f"{idx}-{t}" if t else f"Grafik {idx}")[:31]


def render_agent_analysis_excel(company: Company | None, analysis: dict) -> bytes:
    """Render the analysis to an .xlsx workbook: an Analiz sheet (answer text),
    one sheet per chart (its data), and a Kaynaklar sheet (citations)."""
    from io import BytesIO

    from openpyxl import Workbook

    d = build_agent_analysis_data(company, analysis)
    wb = Workbook()
    ws = wb.active
    ws.title = "Analiz"
    ws.append([d["title"]])
    ws.append(["Şirket", d["company_name"]])
    ws.append(["Oluşturulma", d["generated_at"]])
    if d.get("question"):
        ws.append(["Soru", d["question"]])
    ws.append([])
    ws.append(["Analiz"])
    answer_lines = _strip_action_lines(d["answer_markdown"]) or [d["answer_markdown"]]
    for line in answer_lines:
        ws.append([line])

    for idx, ch in enumerate(d["charts"], 1):
        cws = wb.create_sheet(title=_safe_sheet_title(ch.get("title"), idx))
        series = ch.get("series") or []
        x_key = ch.get("x_key") or "x"
        cws.append([x_key] + [str(se.get("label") or se.get("key") or "") for se in series])
        for pt in (ch.get("data") or []):
            cws.append([pt.get(x_key)] + [pt.get(se.get("key")) for se in series])

    cits = d["citations"]
    if cits:
        kws = wb.create_sheet(title="Kaynaklar")
        kws.append(["Kaynak", "Tür", "Bağlantı"])
        for c in cits:
            kws.append([c.get("label"), c.get("type"), c.get("deep_link")])

    info = wb.create_sheet(title="Bilgi")
    info.append([AI_DISCLAIMER])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
